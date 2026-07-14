# -*- coding: utf-8 -*-
"""Approval Chain Resolver and Gate Replay Analyzer.

This script parses all upstream manifests, matches candidates to their blocking phases,
replays blockers, calculates SHA-256 hashes of inputs, and exports an Excel review
sheet with 5 sheets: Chain Summary, Blockers, Candidate Replay, Required Actions,
and Source Files QA.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.feasibility.export_profile_commit_gate import APPROVAL_PHRASE


# Inputs
PATH_GATE_MANIFEST = PROJECT_ROOT / "feasibility_outputs/profile_commit_gate/profile_commit_gate_manifest.json"
PATH_SIM_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_write_summary.json"
PATH_SIM_RECORDS = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_material_records.json"
PATH_MATCH_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_summary.json"
PATH_MATCH_RESULTS = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_results.json"
PATH_RESOL_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_master_review_resolution/master_review_resolution_summary.json"
PATH_RESOL_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_master_review_resolution/master_review_resolution_items.json"
PATH_PLAN_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft/final_write_plan_summary.json"
PATH_PLAN_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft/final_write_plan_items.json"
PATH_PLAN_RISKS = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft/final_write_plan_risk_register.json"

PATH_CANDIDATE_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate/write_candidate_items.json"

# Output folder
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_approval_chain_resolver"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def calculate_file_hash(path: Path) -> str:
    if not path.exists():
        return "MISSING_FILE"
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()


def write_excel_resolver_output(
    output_dir: Path,
    summary: Dict[str, Any],
    blockers: List[Dict[str, Any]],
    replay_items: List[Dict[str, Any]],
    required_actions: List[Dict[str, Any]],
    qa_files: List[Dict[str, Any]]
) -> None:
    xlsx_path = output_dir / "approval_chain_replay.xlsx"
    wb = openpyxl.Workbook()

    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt
    fill_success = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Xanh lá nhạt
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Chain Summary
    ws_sum = wb.active
    ws_sum.title = "Chain Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số chuỗi (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("top_level_status", summary.get("top_level_status", "N/A")),
        ("commit_gate_status", summary.get("commit_gate_status", "N/A")),
        ("simulation_status", summary.get("simulation_status", "N/A")),
        ("master_match_status", summary.get("master_match_status", "N/A")),
        ("master_resolution_status", summary.get("master_resolution_status", "N/A")),
        ("final_write_plan_status", summary.get("final_write_plan_status", "N/A")),
        ("total_blockers_count", summary.get("total_blockers_count", 0)),
        ("total_candidates_count", summary.get("total_candidates_count", 0)),
        ("ready_for_execution", False),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_warning
        else:
            val_cell.alignment = align_left
            if "BLOCKED" in str(val) or "REQUIRED" in str(val) or "MISSING" in str(val):
                val_cell.fill = fill_danger
            elif "READY" in str(val) or "APPROVED" in str(val):
                val_cell.fill = fill_success

    # Sheet 2: Blockers
    ws_blk = wb.create_sheet(title="Blockers")
    ws_blk.views.sheetView[0].showGridLines = True
    ws_blk.freeze_panes = "A2"
    
    headers_blk = ["blocker_id", "phase", "blocking_reason", "required_human_action", "required_command_if_any", "severity"]
    ws_blk.append(headers_blk)
    for col_idx in range(1, len(headers_blk) + 1):
        cell = ws_blk.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, blk in enumerate(blockers, start=2):
        row_data = [blk.get(k, "") for k in headers_blk]
        ws_blk.append(row_data)
        
        sev = blk.get("severity", "WARNING")
        fill_row = fill_danger if sev == "CRITICAL" else fill_warning
        
        for col_idx, k in enumerate(headers_blk, start=1):
            cell = ws_blk.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_row
            
            if k in ["blocker_id", "phase", "severity"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_blk.auto_filter.ref = f"A1:{get_column_letter(len(headers_blk))}{len(blockers) + 1}"

    # Sheet 3: Candidate Replay
    ws_rep = wb.create_sheet(title="Candidate Replay")
    ws_rep.views.sheetView[0].showGridLines = True
    ws_rep.freeze_panes = "A2"
    
    headers_rep = [
        "chain_item_id", "write_candidate_id", "final_plan_item_id", "supplier_code",
        "material_code", "current_blocking_phase", "blocking_reason", "upstream_status",
        "required_human_action", "risk_level"
    ]
    ws_rep.append(headers_rep)
    for col_idx in range(1, len(headers_rep) + 1):
        cell = ws_rep.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, rep in enumerate(replay_items, start=2):
        row_data = [rep.get(k, "") for k in headers_rep]
        ws_rep.append(row_data)
        
        phase = rep.get("current_blocking_phase", "None")
        fill_row = fill_success if phase == "None" else fill_danger
        
        for col_idx, k in enumerate(headers_rep, start=1):
            cell = ws_rep.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_row
            
            if k in ["chain_item_id", "write_candidate_id", "final_plan_item_id", "supplier_code", "current_blocking_phase", "risk_level"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_rep.auto_filter.ref = f"A1:{get_column_letter(len(headers_rep))}{len(replay_items) + 1}"

    # Sheet 4: Required Actions
    ws_act = wb.create_sheet(title="Required Actions")
    ws_act.views.sheetView[0].showGridLines = True
    ws_act.freeze_panes = "A2"
    
    headers_act = ["Thứ tự (Order)", "Khâu (Phase)", "Hành động yêu cầu (Required Human Action)", "Lệnh chạy gỡ chặn (Required Command)"]
    ws_act.append(headers_act)
    for col_idx in range(1, len(headers_act) + 1):
        cell = ws_act.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, act in enumerate(required_actions, start=2):
        row_data = [act.get(k, "") for k in ["order", "phase", "action", "command"]]
        ws_act.append(row_data)
        
        for col_idx in range(1, len(headers_act) + 1):
            cell = ws_act.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_warning
            if col_idx in [1, 2]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_act.auto_filter.ref = f"A1:{get_column_letter(len(headers_act))}{len(required_actions) + 1}"

    # Sheet 5: Source Files QA
    ws_qa = wb.create_sheet(title="Source Files QA")
    ws_qa.views.sheetView[0].showGridLines = True
    ws_qa.freeze_panes = "A2"
    
    headers_qa = ["File Name", "Path", "Status", "SHA-256 Hash", "Ready Flag check"]
    ws_qa.append(headers_qa)
    for col_idx in range(1, len(headers_qa) + 1):
        cell = ws_qa.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, qa in enumerate(qa_files, start=2):
        row_data = [
            qa.get("name"),
            qa.get("path"),
            qa.get("status"),
            qa.get("hash"),
            qa.get("ready_flag_check")
        ]
        ws_qa.append(row_data)
        
        status = qa.get("status")
        fill_row = fill_success if status == "EXIST" else fill_danger
        
        for col_idx in range(1, 6):
            cell = ws_qa.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_row
            if col_idx in [1, 3, 5]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_qa.auto_filter.ref = f"A1:{get_column_letter(len(headers_qa))}{len(qa_files) + 1}"

    # Auto-fit column widths
    for ws in [ws_sum, ws_blk, ws_rep, ws_act, ws_qa]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_unblock_checklist(required_actions: List[Dict[str, Any]]) -> str:
    lines = [
        "# Unblock Action Checklist – Gate Replay",
        "",
        "Dưới đây là danh sách các hành động phân xử bằng tay và câu lệnh chạy để gỡ chặn toàn bộ chuỗi theo đúng thứ tự luồng.",
        ""
    ]
    for act in required_actions:
        lines.append(f"### [ ] Bước {act['order']}: {act['phase']}")
        lines.append(f"- **Hành động**: {act['action']}")
        if act['command']:
            lines.append(f"- **Lệnh chạy**: `{act['command']}`")
        lines.append("")
    return "\n".join(lines)


def build_report_md(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Approval Chain Resolver Report – Phase 2I

Báo cáo Replay trạng thái chuỗi phê duyệt MEP Quotation Pipeline và quản trị rủi ro an toàn.

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN DRY-RUN**
> * Resolver này **CHƯA** thực hiện ghi thật vào database hoặc main production pipeline của dự án.
> * Trạng thái an toàn: `ready_for_execution = FALSE` và `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Trạng Thái Chuỗi Phê Duyệt (Approval Chain Status)

* Top-level Status: `{summary['top_level_status']}`
* Số lượng Blockers hoạt động: `{summary['total_blockers_count']}`

### Trạng thái các Phase chi tiết:
- Commit Gate: `{summary['commit_gate_status']}`
- Write Simulation: `{summary['simulation_status']}`
- Master Matching: `{summary['master_match_status']}`
- Master Resolution: `{summary['master_resolution_status']}`
- Final Write Plan: `{summary['final_write_plan_status']}`

## 2. Tiêu Chí Để Đạt Trạng Thái Sẵn Sàng (CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY)

Chuỗi chỉ được xem là sẵn sàng thiết kế Phase tiếp theo khi:
- [ ] Commit Gate đạt trạng thái `APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY`.
- [ ] Không còn blockers nào trong danh sách Blockers.
- [ ] Mọi tệp tin JSON/Excel được QA xác nhận an toàn.
- [ ] `ready_for_execution` và `ready_for_write_to_main_pipeline` luôn giữ là `FALSE`.

---

## 3. Các Tệp Tin Replay Cục Bộ
* Blockers JSON: `{output_dir / 'approval_chain_blockers.json'}`
* Status JSON: `{output_dir / 'approval_chain_status.json'}`
* Excel Replay: `{output_dir / 'approval_chain_replay.xlsx'}`
* Checklist gỡ chặn: `{output_dir / 'unblock_action_checklist.md'}`
"""


def check_source_files_qa() -> List[Dict[str, Any]]:
    # QA các file input trong chuỗi
    files_to_check = [
        ("Commit Gate Manifest", PATH_GATE_MANIFEST),
        ("Simulation Summary", PATH_SIM_SUMMARY),
        ("Simulation Records", PATH_SIM_RECORDS),
        ("Match Summary", PATH_MATCH_SUMMARY),
        ("Match Results", PATH_MATCH_RESULTS),
        ("Resolution Summary", PATH_RESOL_SUMMARY),
        ("Resolution Items", PATH_RESOL_ITEMS),
        ("Final Plan Summary", PATH_PLAN_SUMMARY),
        ("Final Plan Items", PATH_PLAN_ITEMS),
        ("Risk Register", PATH_PLAN_RISKS),
        ("Candidate Items", PATH_CANDIDATE_ITEMS)
    ]
    
    qa_list = []
    for name, p in files_to_check:
        status = "MISSING"
        file_hash = "N/A"
        ready_flag_check = "OK"
        
        if p.exists():
            status = "EXIST"
            file_hash = calculate_file_hash(p)
            # Kiểm tra xem có chứa từ overclaim hoặc ready flag = True không
            try:
                content = p.read_text(encoding="utf-8")
                # Chặn các từ ngữ overclaim
                if "production-ready" in content.lower() or "ready for real write" in content.lower():
                    ready_flag_check = "OVERCLAIM_DETECTED"
                
                # Check JSON flags
                if p.suffix == ".json":
                    data = json.loads(content)
                    if isinstance(data, dict):
                        # Quét đệ quy tìm ready_for_write_to_main_pipeline hoặc ready_for_real_write hoặc ready_for_execution
                        def scan_flags(obj):
                            if isinstance(obj, dict):
                                for k, v in obj.items():
                                    if k in ["ready_for_write_to_main_pipeline", "ready_for_real_write", "ready_for_execution"]:
                                        if v is True:
                                            return False
                                    if not scan_flags(v):
                                        return False
                            elif isinstance(obj, list):
                                for item in obj:
                                    if not scan_flags(item):
                                        return False
                            return True
                        if not scan_flags(data):
                            ready_flag_check = "READY_FLAG_TRUE_DETECTED"
            except Exception:
                ready_flag_check = "READ_ERROR"
        else:
            ready_flag_check = "MISSING_FILE"
            
        qa_list.append({
            "name": name,
            "path": str(p),
            "status": status,
            "hash": file_hash,
            "ready_flag_check": ready_flag_check
        })
    return qa_list


def resolve_approval_chain(
    gate_manifest_path: Path,
    sim_summary_path: Path,
    sim_records_path: Path,
    match_summary_path: Path,
    match_results_path: Path,
    resol_summary_path: Path,
    resol_items_path: Path,
    plan_summary_path: Path,
    plan_items_path: Path,
    plan_risks_path: Path,
    candidates_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    # 1. Đọc dữ liệu
    gate = load_json(gate_manifest_path, default={})
    sim_sum = load_json(sim_summary_path, default={})
    sim_recs = load_json(sim_records_path, default=[])
    match_sum = load_json(match_summary_path, default={})
    match_res = load_json(match_results_path, default=[])
    resol_sum = load_json(resol_summary_path, default={})
    resol_its = load_json(resol_items_path, default=[])
    plan_sum = load_json(plan_summary_path, default={})
    plan_its = load_json(plan_items_path, default=[])
    
    candidates = load_json(candidates_path, default=[])

    # Trạng thái các phase
    commit_gate_status = gate.get("commit_gate_status", "MISSING_FILE") if gate_manifest_path.exists() else "MISSING_FILE"
    simulation_status = sim_sum.get("summary", {}).get("simulation_status", "MISSING_FILE") if sim_summary_path.exists() else "MISSING_FILE"
    master_match_status = match_sum.get("summary", {}).get("proposed_status", "MISSING_FILE") if match_summary_path.exists() else "MISSING_FILE"
    master_resolution_status = resol_sum.get("summary", {}).get("proposed_status", "MISSING_FILE") if resol_summary_path.exists() else "MISSING_FILE"
    final_write_plan_status = plan_sum.get("summary", {}).get("proposed_status", "MISSING_FILE") if plan_summary_path.exists() else "MISSING_FILE"

    # Lập danh sách blockers & required actions
    blockers = []
    required_actions = []
    
    # Order order tracker
    act_order = 1

    # Khâu 1: Commit Gate
    if commit_gate_status != "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY":
        approve_cmd = f'python tools/feasibility/export_profile_commit_gate.py --approve --approved-by "<reviewer>" --approval-note "<ghi chú>" --approval-phrase "{APPROVAL_PHRASE}"'
        blockers.append({
            "blocker_id": "BLOCKER_0001",
            "phase": "Commit Gate",
            "blocking_reason": f"Commit Gate chưa được phê duyệt (Trạng thái hiện tại: {commit_gate_status}).",
            "required_human_action": "Reviewer mở tệp gate md và chạy script duyệt bằng approval phrase thích hợp.",
            "required_command_if_any": approve_cmd,
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Commit Gate",
            "action": "Người dùng phê duyệt Commit Gate bằng cách nhập approval phrase thích hợp.",
            "command": approve_cmd
        })
        act_order += 1

    # Khâu 2: Simulation
    if simulation_status == "BLOCKED_BY_COMMIT_GATE":
        blockers.append({
            "blocker_id": "BLOCKER_0002",
            "phase": "Write Simulation",
            "blocking_reason": "Sandbox simulation bị chặn do Commit Gate chưa được phê duyệt.",
            "required_human_action": "Phê duyệt Commit Gate trước, sau đó chạy lại sandbox simulation.",
            "required_command_if_any": "python tools/feasibility/run_profile_write_simulation.py",
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Write Simulation",
            "action": "Chạy sandbox simulation sau khi Commit Gate đã được phê duyệt thành công.",
            "command": "python tools/feasibility/run_profile_write_simulation.py"
        })
        act_order += 1
    elif simulation_status == "MISSING_FILE":
        blockers.append({
            "blocker_id": "BLOCKER_0003",
            "phase": "Write Simulation",
            "blocking_reason": "Chưa có file summary của sandbox simulation.",
            "required_human_action": "Chạy lệnh simulation để sinh tệp tin simulated records.",
            "required_command_if_any": "python tools/feasibility/run_profile_write_simulation.py",
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Write Simulation",
            "action": "Chạy lệnh simulation sandbox.",
            "command": "python tools/feasibility/run_profile_write_simulation.py"
        })
        act_order += 1

    # Khâu 3: Master Match
    if master_match_status == "BLOCKED_NO_SIMULATION_RECORDS":
        blockers.append({
            "blocker_id": "BLOCKER_0004",
            "phase": "Master Matching",
            "blocking_reason": "Master matching dry-run không có records simulated đầu vào.",
            "required_human_action": "Đảm bảo simulation đã chạy thành công có record, rồi chạy lại đối chiếu master.",
            "required_command_if_any": "python tools/feasibility/run_profile_master_match_dry_run.py",
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Master Matching",
            "action": "Chạy đối chiếu Master Matching dry-run.",
            "command": "python tools/feasibility/run_profile_master_match_dry_run.py"
        })
        act_order += 1
    elif master_match_status == "MISSING_FILE":
        blockers.append({
            "blocker_id": "BLOCKER_0005",
            "phase": "Master Matching",
            "blocking_reason": "Chưa có file summary của Master Matching.",
            "required_human_action": "Chạy lệnh master match để sinh tệp kết quả đối chiếu.",
            "required_command_if_any": "python tools/feasibility/run_profile_master_match_dry_run.py",
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Master Matching",
            "action": "Chạy lệnh Master Matching dry-run.",
            "command": "python tools/feasibility/run_profile_master_match_dry_run.py"
        })
        act_order += 1

    # Khâu 4: Resolution
    if master_resolution_status == "MASTER_REVIEW_RESOLUTION_REQUIRED":
        pending_resol_count = len([r for r in resol_its if r.get("human_resolution_decision") == "PENDING"])
        blockers.append({
            "blocker_id": "BLOCKER_0006",
            "phase": "Master Resolution",
            "blocking_reason": f"Tồn tại {pending_resol_count} dòng đối chiếu master ở trạng thái PENDING chưa phân xử.",
            "required_human_action": "Mở file Excel resolution, chọn quyết định phân xử cho từng dòng, và chạy lệnh xuất resolution.",
            "required_command_if_any": "python tools/feasibility/export_profile_master_review_resolution.py",
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Master Resolution",
            "action": "Mở file Excel resolution, điền quyết định phân xử và chạy lệnh đóng gói resolution.",
            "command": "python tools/feasibility/export_profile_master_review_resolution.py"
        })
        act_order += 1
    elif master_resolution_status == "MISSING_FILE":
        blockers.append({
            "blocker_id": "BLOCKER_0007",
            "phase": "Master Resolution",
            "blocking_reason": "Chưa sinh tệp template phân xử Master Review Resolution.",
            "required_human_action": "Chạy lệnh xuất resolution để tạo tệp tin phân xử.",
            "required_command_if_any": "python tools/feasibility/export_profile_master_review_resolution.py",
            "severity": "CRITICAL"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Master Resolution",
            "action": "Chạy xuất Master Review Resolution template.",
            "command": "python tools/feasibility/export_profile_master_review_resolution.py"
        })
        act_order += 1

    # Khâu 5: Final Write Plan
    if final_write_plan_status == "FINAL_WRITE_PLAN_BLOCKED":
        blockers.append({
            "blocker_id": "BLOCKER_0008",
            "phase": "Final Write Plan",
            "blocking_reason": "Bản kế hoạch ghi Final Write Plan bị chặn do có dòng blocked hoặc HIGH risk chưa xử lý.",
            "required_human_action": "Giải quyết hết các lỗi resolution và cập nhật final write plan draft.",
            "required_command_if_any": "python tools/feasibility/export_profile_final_write_plan_draft.py",
            "severity": "WARNING"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Final Write Plan",
            "action": "Chạy lại lệnh đóng gói Final Write Plan Draft để cập nhật kế hoạch sạch.",
            "command": "python tools/feasibility/export_profile_final_write_plan_draft.py"
        })
        act_order += 1
    elif final_write_plan_status == "MISSING_FILE":
        blockers.append({
            "blocker_id": "BLOCKER_0009",
            "phase": "Final Write Plan",
            "blocking_reason": "Chưa sinh tệp Final Write Plan Draft.",
            "required_human_action": "Chạy lệnh đóng gói final write plan draft.",
            "required_command_if_any": "python tools/feasibility/export_profile_final_write_plan_draft.py",
            "severity": "WARNING"
        })
        required_actions.append({
            "order": act_order,
            "phase": "Final Write Plan",
            "action": "Chạy lệnh Final Write Plan Draft.",
            "command": "python tools/feasibility/export_profile_final_write_plan_draft.py"
        })
        act_order += 1

    # Quyết định top-level status
    if len(candidates) == 0:
        top_level_status = "CHAIN_EMPTY"
    elif len(blockers) > 0:
        top_level_status = "CHAIN_BLOCKED"
    else:
        top_level_status = "CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY"

    # Lập danh sách Replay Items cho từng candidate
    replay_items = []
    
    # Tra cứu nhanh
    sim_map = {s["write_candidate_id"]: s for s in sim_recs if s.get("write_candidate_id")}
    match_map = {m["write_candidate_id"]: m for m in match_res if m.get("write_candidate_id")}
    resol_map = {r["write_candidate_id"]: r for r in resol_its if r.get("write_candidate_id")}
    plan_map = {p["write_candidate_id"]: p for p in plan_its if p.get("write_candidate_id")}

    for idx, cand in enumerate(candidates, start=1):
        chain_item_id = f"CHAIN_REP_{idx:04d}"
        cand_id = cand.get("write_candidate_id", "")
        supplier = cand.get("supplier_code", "")
        material = cand.get("material_code", "")
        
        sim = sim_map.get(cand_id, {})
        match = match_map.get(cand_id, {})
        resol = resol_map.get(cand_id, {})
        plan = plan_map.get(cand_id, {})
        
        final_plan_item_id = plan.get("final_plan_item_id")
        risk_level = plan.get("risk_level", "HIGH" if top_level_status == "CHAIN_BLOCKED" else "LOW")
        
        # Tìm xem item này bị block ở khâu nào
        current_blocking_phase = "None"
        blocking_reason = "Không bị chặn ở khâu nào."
        upstream_status = "OK"
        req_action = "Sẵn sàng cho thiết kế Phase tiếp theo."
        req_command = ""
        
        if commit_gate_status != "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY":
            current_blocking_phase = "Commit Gate"
            blocking_reason = "Gate chưa approve"
            upstream_status = "BLOCKED"
            req_action = "Duyệt Commit Gate bằng approval phrase."
            req_command = f'python tools/feasibility/export_profile_commit_gate.py --approve --approved-by "<reviewer>" --approval-note "<ghi chú>" --approval-phrase "{APPROVAL_PHRASE}"'
        elif not sim:
            current_blocking_phase = "Simulation"
            blocking_reason = "Chưa giả lập ghi simulation record."
            upstream_status = "BLOCKED"
            req_action = "Chạy sandbox simulation."
            req_command = "python tools/feasibility/run_profile_write_simulation.py"
        elif not match:
            current_blocking_phase = "Master Match"
            blocking_reason = "Chưa chạy đối chiếu master matching."
            upstream_status = "BLOCKED"
            req_action = "Chạy master match dry-run."
            req_command = "python tools/feasibility/run_profile_master_match_dry_run.py"
        elif resol and resol.get("human_resolution_decision") == "PENDING":
            current_blocking_phase = "Master Resolution"
            blocking_reason = "Quyết định phân xử đang là PENDING."
            upstream_status = "BLOCKED"
            req_action = "Người dùng điền phân xử trên file Excel resolution."
            req_command = "python tools/feasibility/export_profile_master_review_resolution.py"
        elif plan and plan.get("final_planned_action") == "PLAN_BLOCKED":
            current_blocking_phase = "Final Write Plan"
            blocking_reason = "Kế hoạch ghi bị chặn do rủi ro chưa gỡ."
            upstream_status = "BLOCKED"
            req_action = "Giải quyết hết các lỗi resolution."
            req_command = "python tools/feasibility/export_profile_final_write_plan_draft.py"

        replay_items.append({
            "chain_item_id": chain_item_id,
            "write_candidate_id": cand_id,
            "final_plan_item_id": final_plan_item_id,
            "supplier_code": supplier,
            "material_code": material,
            "current_blocking_phase": current_blocking_phase,
            "blocking_reason": blocking_reason,
            "upstream_status": upstream_status,
            "required_human_action": req_action,
            "required_command_if_any": req_command,
            "risk_level": risk_level,
            "can_auto_resolve": False
        })

    # Summary metrics
    summary = {
        "top_level_status": top_level_status,
        "commit_gate_status": commit_gate_status,
        "simulation_status": simulation_status,
        "master_match_status": master_match_status,
        "master_resolution_status": master_resolution_status,
        "final_write_plan_status": final_write_plan_status,
        "total_blockers_count": len(blockers),
        "total_candidates_count": len(candidates)
    }

    manifest = {
        "resolver_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "approval_chain_resolver",
        "top_level_status": top_level_status,
        "ready_for_execution": False,
        "ready_for_write_to_main_pipeline": False,
        "chain_summary": summary,
        "blockers": blockers,
        "replay_items": replay_items
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "approval_chain_status.json", manifest)
    write_json(output_dir / "approval_chain_blockers.json", blockers)

    # QA Files check
    qa_files = check_source_files_qa()

    # Excel output
    write_excel_resolver_output(
        output_dir=output_dir,
        summary=summary,
        blockers=blockers,
        replay_items=replay_items,
        required_actions=required_actions,
        qa_files=qa_files
    )

    # Markdown Report
    report_content = build_report_md(summary, output_dir)
    (output_dir / "approval_chain_report.md").write_text(report_content, encoding="utf-8")

    # Checklist MD
    checklist_content = build_unblock_checklist(required_actions)
    (output_dir / "unblock_action_checklist.md").write_text(checklist_content, encoding="utf-8")

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run approval chain resolver.")
    parser.add_argument("--gate-manifest", type=Path, default=PATH_GATE_MANIFEST)
    parser.add_argument("--sim-summary", type=Path, default=PATH_SIM_SUMMARY)
    parser.add_argument("--sim-records", type=Path, default=PATH_SIM_RECORDS)
    parser.add_argument("--match-summary", type=Path, default=PATH_MATCH_SUMMARY)
    parser.add_argument("--match-results", type=Path, default=PATH_MATCH_RESULTS)
    parser.add_argument("--resol-summary", type=Path, default=PATH_RESOL_SUMMARY)
    parser.add_argument("--resol-items", type=Path, default=PATH_RESOL_ITEMS)
    parser.add_argument("--plan-summary", type=Path, default=PATH_PLAN_SUMMARY)
    parser.add_argument("--plan-items", type=Path, default=PATH_PLAN_ITEMS)
    parser.add_argument("--plan-risks", type=Path, default=PATH_PLAN_RISKS)
    parser.add_argument("--candidates", type=Path, default=PATH_CANDIDATE_ITEMS)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = resolve_approval_chain(
        gate_manifest_path=args.gate_manifest,
        sim_summary_path=args.sim_summary,
        sim_records_path=args.sim_records,
        match_summary_path=args.match_summary,
        match_results_path=args.match_results,
        resol_summary_path=args.resol_summary,
        resol_items_path=args.resol_items,
        plan_summary_path=args.plan_summary,
        plan_items_path=args.plan_items,
        plan_risks_path=args.plan_risks,
        candidates_path=args.candidates,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["chain_summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
