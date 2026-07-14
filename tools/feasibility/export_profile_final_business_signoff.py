# -*- coding: utf-8 -*-
"""Final Business Sign-off Package Generator.

This script aggregates final write plan items and approval chain status to generate
a business sign-off template Excel and JSON package.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import csv
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Inputs
PATH_GATE_MANIFEST = PROJECT_ROOT / "feasibility_outputs/profile_commit_gate/profile_commit_gate_manifest.json"
PATH_SIM_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_write_summary.json"
PATH_SIM_RECORDS = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_material_records.json"
PATH_MATCH_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_summary.json"
PATH_MATCH_RESULTS = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_results.json"
PATH_PLAN_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft/final_write_plan_items.json"
PATH_PLAN_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft/final_write_plan_summary.json"
PATH_PLAN_RISKS = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft/final_write_plan_risk_register.json"
PATH_CHAIN_STATUS = PROJECT_ROOT / "feasibility_outputs/profile_approval_chain_resolver/approval_chain_status.json"
PATH_CHAIN_BLOCKERS = PROJECT_ROOT / "feasibility_outputs/profile_approval_chain_resolver/approval_chain_blockers.json"

DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_final_business_signoff"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def sha256_file(path: Path) -> str:
    if not path.exists():
        return ""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_excel_signoff_output(
    output_dir: Path,
    summary: Dict[str, Any],
    signoff_items: List[Dict[str, Any]],
    chain_summary: Dict[str, Any]
) -> None:
    xlsx_path = output_dir / "final_business_signoff_template.xlsx"
    wb = openpyxl.Workbook()

    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Đỏ cam nhạt
    fill_success = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Xanh lá nhạt
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số Sign-off (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("proposed_status", summary.get("proposed_status", "N/A")),
        ("total_signoff_items", summary.get("total_signoff_items", 0)),
        ("approved_count", summary.get("approved_count", 0)),
        ("pending_count", summary.get("pending_count", 0)),
        ("rejected_count", summary.get("rejected_count", 0)),
        ("ready_for_execution", False),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_pending
        else:
            val_cell.alignment = align_left
            if "PENDING" in str(val) or "BLOCKED" in str(val) or "STALE" in str(val):
                val_cell.fill = fill_danger
            elif "APPROVED" in str(val):
                val_cell.fill = fill_success

    # Sheet 2: Sign-off Items
    ws_items = wb.create_sheet(title="Sign-off Items")
    ws_items.views.sheetView[0].showGridLines = True
    ws_items.freeze_panes = "A2"
    
    headers_items = [
        "signoff_item_id", "final_plan_item_id", "write_candidate_id", "supplier_code",
        "material_code", "normalized_description", "unit", "quantity", "unit_price",
        "amount", "planned_action", "risk_level", "human_decision", "human_note",
        "reviewed_by", "reviewed_at"
    ]
    ws_items.append(headers_items)
    for col_idx in range(1, len(headers_items) + 1):
        cell = ws_items.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Dropdown Validation cho human_decision
    dv = DataValidation(
        type="list",
        formula1='"PENDING,APPROVE_FOR_EXECUTOR_DESIGN,REJECT,NEEDS_CORRECTION,NEEDS_SOURCE_REVIEW"',
        allow_blank=False
    )
    dv.error = "Lựa chọn quyết định không hợp lệ"
    dv.errorTitle = "Lỗi nhập liệu"
    dv.prompt = "Vui lòng chọn quyết định phê duyệt"
    dv.promptTitle = "Quyết định phê duyệt"
    if len(signoff_items) > 0:
        ws_items.add_data_validation(dv)
        # Chỉ định dropdown cho cột human_decision (Cột M)
        dv.add(f"M2:M{len(signoff_items) + 1}")

    for r_idx, item in enumerate(signoff_items, start=2):
        row_data = [item.get(k, "") for k in headers_items]
        ws_items.append(row_data)
        
        dec = item.get("human_decision", "PENDING")
        fill_row = fill_pending if dec == "PENDING" else (fill_success if dec == "APPROVE_FOR_EXECUTOR_DESIGN" else fill_danger)
        
        for col_idx, k in enumerate(headers_items, start=1):
            cell = ws_items.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_row
            
            if k in ["quantity", "unit_price", "amount"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["signoff_item_id", "final_plan_item_id", "write_candidate_id", "supplier_code", "planned_action", "risk_level", "human_decision"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_items.auto_filter.ref = f"A1:{get_column_letter(len(headers_items))}{len(signoff_items) + 1}"

    # Sheet 3: Decision Options
    ws_opt = wb.create_sheet(title="Decision Options")
    ws_opt.views.sheetView[0].showGridLines = True
    ws_opt.freeze_panes = "A2"
    
    headers_opt = ["Quyết định (Decision Option)", "Giải thích nghĩa (Explanation)", "Ràng buộc dữ liệu (Validation Rule)"]
    ws_opt.append(headers_opt)
    for col_idx in range(1, len(headers_opt) + 1):
        cell = ws_opt.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    options_data = [
        ("PENDING", "Quyết định mặc định ban đầu, đang chờ reviewer phê duyệt.", "Không được ghi nhận approved."),
        ("APPROVE_FOR_EXECUTOR_DESIGN", "Chấp thuận ghi nhận vật tư này cho khâu thiết kế executor ghi thật.", "Được ghi nhận approved."),
        ("REJECT", "Từ chối ghi nhận vật tư này.", "Bắt buộc điền Human Note lý do từ chối."),
        ("NEEDS_CORRECTION", "Yêu cầu chỉnh sửa tọa độ/cấu hình parser upstream.", "Bắt buộc điền Human Note lý do cần sửa."),
        ("NEEDS_SOURCE_REVIEW", "Yêu cầu rà soát đối chiếu tài liệu PDF gốc.", "Bắt buộc điền Human Note lý do chi tiết.")
    ]
    
    for r_idx, opt in enumerate(options_data, start=2):
        ws_opt.append(opt)
        for col_idx in range(1, 4):
            cell = ws_opt.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_bold
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_opt.auto_filter.ref = f"A1:C{len(options_data) + 1}"

    # Sheet 4: Source Evidence
    ws_ev = wb.create_sheet(title="Source Evidence")
    ws_ev.views.sheetView[0].showGridLines = True
    ws_ev.freeze_panes = "A2"
    
    headers_ev = ["signoff_item_id", "supplier_code", "material_code", "provenance", "evidence_text", "master_match_result"]
    ws_ev.append(headers_ev)
    for col_idx in range(1, len(headers_ev) + 1):
        cell = ws_ev.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, item in enumerate(signoff_items, start=2):
        row_data = [item.get(k, "") for k in headers_ev]
        ws_ev.append(row_data)
        
        for col_idx, k in enumerate(headers_ev, start=1):
            cell = ws_ev.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if k in ["signoff_item_id", "supplier_code", "material_code"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_ev.auto_filter.ref = f"A1:{get_column_letter(len(headers_ev))}{len(signoff_items) + 1}"

    # Sheet 5: Approval Preconditions
    ws_pre = wb.create_sheet(title="Approval Preconditions")
    ws_pre.views.sheetView[0].showGridLines = True
    ws_pre.freeze_panes = "A2"
    
    headers_pre = ["Khâu (Phase)", "Trạng thái upstream (Upstream Status)", "Check list đạt yêu cầu"]
    ws_pre.append(headers_pre)
    for col_idx in range(1, len(headers_pre) + 1):
        cell = ws_pre.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    preconditions = [
        ("Commit Gate Status", chain_summary.get("commit_gate_status", "N/A")),
        ("Simulation Status", chain_summary.get("simulation_status", "N/A")),
        ("Master Matching Status", chain_summary.get("master_match_status", "N/A")),
        ("Master Resolution Status", chain_summary.get("master_resolution_status", "N/A")),
        ("Final Write Plan Status", chain_summary.get("final_write_plan_status", "N/A"))
    ]
    
    for r_idx, pre in enumerate(preconditions, start=2):
        status = pre[1]
        is_ok = "PASS" if ("APPROVED" in status or "READY" in status or "NO_MASTER" in status) else "BLOCKED"
        ws_pre.append([pre[0], pre[1], is_ok])
        
        fill_row = fill_success if is_ok == "PASS" else fill_danger
        
        for col_idx in range(1, 4):
            cell = ws_pre.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_row
            if col_idx in [1, 3]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_pre.auto_filter.ref = f"A1:C{len(preconditions) + 1}"

    # Auto-fit column widths
    for ws in [ws_sum, ws_items, ws_opt, ws_ev, ws_pre]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_guide_md(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Final Business Sign-off Guide – Phase 2J

Tài liệu hướng dẫn phê duyệt nghiệp vụ cuối cùng trước khi chuyển sang Phase thiết kế write executor ghi thật.

---

## 1. Hướng Dẫn Review Quyết Định Trên Excel

Reviewer thực hiện mở tệp Excel `final_business_signoff_template.xlsx` và điền quyết định tại cột `human_decision` (Cột M) của sheet `Sign-off Items`:
- **APPROVE_FOR_EXECUTOR_DESIGN**: Duyệt vật tư này chuyển tiếp sang Phase thiết kế executor ghi thật.
- **REJECT / NEEDS_CORRECTION / NEEDS_SOURCE_REVIEW**: Từ chối hoặc yêu cầu sửa đổi parser upstream. **Bắt buộc điền Human Note lý do tương ứng (Cột N).**

## 2. Tiêu Chí Để Đạt Trạng Thái APPROVED

Gói Business Sign-off chỉ được xem là phê duyệt hoàn toàn khi:
- [ ] 100% các dòng sign-off được duyệt ở trạng thái `APPROVE_FOR_EXECUTOR_DESIGN`.
- [ ] Mọi ready flags vẫn giữ chặt là `FALSE` bảo mật an toàn.

---

> [!IMPORTANT]
> **BƯỚC TIẾP THEO SAU PHASE 2J**
> * Bước tiếp theo chỉ được bắt đầu khi đã có **target contract thật** của main pipeline/database.
"""


def build_report_md(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Final Business Sign-off Report – Phase 2J

Báo cáo kết quả đóng gói Final Business Sign-off Package của MEP Quotation Pipeline.

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN DRY-RUN**
> * Sign-off Package này **CHƯA** thực hiện ghi thật vào database hoặc main production pipeline của dự án.
> * Trạng thái an toàn: `ready_for_execution = FALSE` và `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Trạng Thái Tổng Thể (Sign-off Status)

* Proposed Status: `{summary['proposed_status']}`
* Tổng số Sign-off Items: `{summary['total_signoff_items']}`
* Đã duyệt (Approved): `{summary['approved_count']}`
* Chờ duyệt (Pending): `{summary['pending_count']}`
* Từ chối (Rejected): `{summary['rejected_count']}`

## 2. Các Tệp Tin Kết Xuất Cục Bộ
* Sign-off Items JSON: `{output_dir / 'final_business_signoff_items.json'}`
* Summary JSON: `{output_dir / 'final_business_signoff_summary.json'}`
* CSV Template: `{output_dir / 'final_business_signoff_template.csv'}`
* Excel Template: `{output_dir / 'final_business_signoff_template.xlsx'}`
* Sign-off Report: `{output_dir / 'final_business_signoff_report.md'}`
* Sign-off Guide: `{output_dir / 'final_business_signoff_guide.md'}`
"""


def run_business_signoff_pipeline(
    plan_items_path: Path,
    plan_summary_path: Path,
    plan_risks_path: Path,
    chain_status_path: Path,
    chain_blockers_path: Path,
    sim_records_path: Path,
    match_results_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    # 0. Đảm bảo Phase 2J là Read-Only đối với các file upstream
    upstream_paths = [
        PATH_GATE_MANIFEST,
        PATH_SIM_SUMMARY,
        PATH_SIM_RECORDS,
        PATH_MATCH_SUMMARY,
        PATH_PLAN_SUMMARY,
        PATH_PLAN_ITEMS,
        PATH_CHAIN_STATUS
    ]
    initial_hashes = {str(p): sha256_file(p) for p in upstream_paths}

    # 1. Đọc dữ liệu đầu vào
    plan_items = load_json(plan_items_path, default=[])
    chain_status = load_json(chain_status_path, default={})
    chain_summary = chain_status.get("chain_summary", {})

    # Tính toán SHA-256 hashes của input hiện hành
    hash_plan = sha256_file(plan_items_path)
    hash_chain = sha256_file(chain_status_path)

    # 2. Kiểm tra blocker từ approval chain
    top_level_status = chain_status.get("top_level_status", "CHAIN_BLOCKED")
    total_blockers = chain_summary.get("total_blockers_count", 0)

    # 3. Lọc plan items hợp lệ (PLAN_INSERT / PLAN_UPDATE / PLAN_SKIP)
    signoff_items = []
    signoff_item_idx = 1
    
    # Đọc quyết định phân xử trước đó (nếu đã lưu) để bảo toàn và kiểm chứng STALE
    old_summary_path = output_dir / "final_business_signoff_summary.json"
    old_items_path = output_dir / "final_business_signoff_items.json"
    
    old_manifest = load_json(old_summary_path, default={})
    old_items = load_json(old_items_path, default=[])
    old_hashes = old_manifest.get("source_file_hashes", {})
    
    is_stale = False
    if old_manifest and (old_hashes.get("final_write_plan_items_json") != hash_plan or old_hashes.get("approval_chain_status_json") != hash_chain):
        is_stale = True

    # Tạo map quyết định cũ
    old_dec_map = {item["final_plan_item_id"]: item for item in old_items if item.get("final_plan_item_id")}

    for p_item in plan_items:
        act = p_item.get("final_planned_action")
        if act != "PLAN_BLOCKED":
            signoff_id = f"SIGNOFF_ITEM_{signoff_item_idx:04d}"
            
            # Khởi trị mặc định PENDING
            human_dec = "PENDING"
            human_note = ""
            rev_by = ""
            rev_at = None
            
            # Bảo toàn quyết định cũ nếu khớp
            old_item = old_dec_map.get(p_item.get("final_plan_item_id", ""))
            if old_item and not is_stale:
                human_dec = old_item.get("human_decision", "PENDING")
                human_note = old_item.get("human_note", "")
                rev_by = old_item.get("reviewed_by", "")
                rev_at = old_item.get("reviewed_at")

            # Validation rule check: REJECT, NEEDS_* bắt buộc có note
            if human_dec in ["REJECT", "NEEDS_CORRECTION", "NEEDS_SOURCE_REVIEW"] and not human_note:
                # Validation fail -> Tạm reset về PENDING
                human_dec = "PENDING"

            signoff_items.append({
                "signoff_item_id": signoff_id,
                "final_plan_item_id": p_item.get("final_plan_item_id", ""),
                "write_candidate_id": p_item.get("write_candidate_id", ""),
                "supplier_code": p_item.get("supplier_code", ""),
                "material_code": p_item.get("material_code", ""),
                "normalized_description": p_item.get("description", ""),
                "unit": p_item.get("unit", ""),
                "quantity": p_item.get("quantity", 0),
                "unit_price": p_item.get("unit_price", 0),
                "amount": p_item.get("amount", 0.0),
                "currency": p_item.get("currency", "VND"),
                "planned_action": act,
                "risk_level": p_item.get("risk_level", "LOW"),
                "provenance": p_item.get("provenance", ""),
                "evidence_text": p_item.get("evidence_text", ""),
                "master_match_result": p_item.get("matched_master_record_id"),
                "human_decision": human_dec,
                "human_note": human_note,
                "reviewed_by": rev_by,
                "reviewed_at": rev_at,
                "decision_required": True
            })
            signoff_item_idx += 1

    # Thống kê sign-off items
    total_items = len(signoff_items)
    approved_count = len([x for x in signoff_items if x["human_decision"] == "APPROVE_FOR_EXECUTOR_DESIGN"])
    pending_count = len([x for x in signoff_items if x["human_decision"] == "PENDING"])
    rejected_count = len([x for x in signoff_items if x["human_decision"] in ["REJECT", "NEEDS_CORRECTION", "NEEDS_SOURCE_REVIEW"]])

    # Xác định status
    if total_items == 0:
        proposed_status = "FINAL_BUSINESS_SIGNOFF_PENDING"
    elif top_level_status != "CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY" or total_blockers > 0:
        proposed_status = "BLOCKED_BY_APPROVAL_CHAIN"
    elif is_stale:
        proposed_status = "STALE_SIGNOFF_SOURCE"
    elif approved_count == total_items:
        proposed_status = "FINAL_BUSINESS_SIGNOFF_APPROVED"
    else:
        proposed_status = "FINAL_BUSINESS_SIGNOFF_PENDING"

    summary = {
        "total_signoff_items": total_items,
        "approved_count": approved_count,
        "pending_count": pending_count,
        "rejected_count": rejected_count,
        "proposed_status": proposed_status
    }

    manifest = {
        "signoff_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "final_business_signoff",
        "ready_for_execution": False,
        "ready_for_write_to_main_pipeline": False,
        "source_file_hashes": {
            "final_write_plan_items_json": hash_plan,
            "approval_chain_status_json": hash_chain
        },
        "summary": summary,
        "signoff_items": signoff_items
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "final_business_signoff_items.json", signoff_items)
    write_json(output_dir / "final_business_signoff_summary.json", manifest)

    # Ghi CSV phẳng
    csv_path = output_dir / "final_business_signoff_template.csv"
    headers_csv = [
        "signoff_item_id", "final_plan_item_id", "supplier_code", "material_code",
        "normalized_description", "unit", "quantity", "unit_price", "amount",
        "planned_action", "risk_level", "human_decision", "human_note"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_csv)
        writer.writeheader()
        for item in signoff_items:
            row = {k: item.get(k, "") for k in headers_csv}
            writer.writerow(row)

    # Ghi Excel
    write_excel_signoff_output(output_dir, summary, signoff_items, chain_summary)

    # Ghi Báo cáo & Hướng dẫn Markdown
    guide_content = build_guide_md(summary, output_dir)
    (output_dir / "final_business_signoff_guide.md").write_text(guide_content, encoding="utf-8")

    report_content = build_report_md(summary, output_dir)
    (output_dir / "final_business_signoff_report.md").write_text(report_content, encoding="utf-8")

    # Xác minh lại hash sau khi chạy để đảm bảo không ghi đè file upstream nào
    post_hashes = {str(p): sha256_file(p) for p in upstream_paths}
    for p in upstream_paths:
        if initial_hashes[str(p)] != post_hashes[str(p)]:
            raise RuntimeError(f"[SAFETY VIOLATION] Upstream file has been modified during Phase 2J execution: {p}")

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export final business sign-off package.")
    parser.add_argument("--plan-items", type=Path, default=PATH_PLAN_ITEMS)
    parser.add_argument("--plan-summary", type=Path, default=PATH_PLAN_SUMMARY)
    parser.add_argument("--plan-risks", type=Path, default=PATH_PLAN_RISKS)
    parser.add_argument("--chain-status", type=Path, default=PATH_CHAIN_STATUS)
    parser.add_argument("--chain-blockers", type=Path, default=PATH_CHAIN_BLOCKERS)
    parser.add_argument("--sim-records", type=Path, default=PATH_SIM_RECORDS)
    parser.add_argument("--match-results", type=Path, default=PATH_MATCH_RESULTS)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_business_signoff_pipeline(
        plan_items_path=args.plan_items,
        plan_summary_path=args.plan_summary,
        plan_risks_path=args.plan_risks,
        chain_status_path=args.chain_status,
        chain_blockers_path=args.chain_blockers,
        sim_records_path=args.sim_records,
        match_results_path=args.match_results,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
