# -*- coding: utf-8 -*-
"""Profile Commit Gate / Manual Approval Lock.

This phase creates an approval lock for write candidates. It never writes to
the main pipeline or database. The optional approval mode only marks the
candidate package as approved for the next design phase, not for direct write.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CANDIDATE_DIR = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_commit_gate"
APPROVAL_PHRASE = "APPROVE_PROFILE_WRITE_CANDIDATES_FOR_NEXT_PHASE_ONLY"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def sha256_file(path: Path) -> str:
    if not path.exists():
        return ""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def collect_artifact_hashes(candidate_dir: Path) -> Dict[str, str]:
    artifacts = [
        "write_candidate_items.json",
        "write_candidate_summary.json",
        "write_candidate_review.csv",
        "write_candidate_review.xlsx",
        "write_candidate_commit_plan.md",
    ]
    return {
        name: sha256_file(candidate_dir / name)
        for name in artifacts
    }


def determine_blocking_reasons(candidate_summary: Dict[str, Any]) -> List[str]:
    summary = candidate_summary.get("summary", candidate_summary)
    reasons: List[str] = []

    if summary.get("proposed_status") != "READY_FOR_HUMAN_COMMIT_REVIEW":
        reasons.append("write_candidate_status_not_ready")
    if int(summary.get("candidate_items_count", 0) or 0) <= 0:
        reasons.append("no_write_candidates")
    if int(summary.get("duplicate_write_key_count", 0) or 0) > 0:
        reasons.append("duplicate_write_key_requires_resolution")
    if int(summary.get("skipped_items_count", 0) or 0) > 0:
        reasons.append("skipped_candidates_require_review")
    if candidate_summary.get("ready_for_write_to_main_pipeline") is True:
        reasons.append("candidate_summary_must_not_be_ready_for_main_pipeline")
    if summary.get("ready_for_write_to_main_pipeline") is True:
        reasons.append("nested_summary_must_not_be_ready_for_main_pipeline")

    return reasons


def build_gate_manifest(
    candidate_dir: Path,
    approve: bool = False,
    approved_by: str = "",
    approval_note: str = "",
    approval_phrase: str = "",
) -> Dict[str, Any]:
    summary_path = candidate_dir / "write_candidate_summary.json"
    items_path = candidate_dir / "write_candidate_items.json"
    candidate_summary = load_json(summary_path, default={})
    candidate_items = load_json(items_path, default=[])
    source_hashes = collect_artifact_hashes(candidate_dir)
    blocking_reasons = determine_blocking_reasons(candidate_summary)

    approval_valid = (
        approve
        and not blocking_reasons
        and approved_by.strip()
        and approval_phrase == APPROVAL_PHRASE
    )

    if blocking_reasons:
        status = "BLOCKED_NEEDS_REVIEW"
    elif approval_valid:
        status = "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY"
    else:
        status = "PENDING_HUMAN_APPROVAL"

    approved_at = utc_now() if approval_valid else None
    approval = {
        "approved": bool(approval_valid),
        "approved_by": approved_by.strip() if approval_valid else "",
        "approval_note": approval_note.strip() if approval_valid else "",
        "approval_phrase": approval_phrase if approval_valid else "",
        "approved_at": approved_at,
    }

    manifest = {
        "gate_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "approval_lock",
        "source_write_candidate_summary": str(summary_path),
        "source_write_candidate_items": str(items_path),
        "source_artifact_hashes": source_hashes,
        "candidate_summary": candidate_summary.get("summary", {}),
        "candidate_items_count_observed": len(candidate_items),
        "commit_gate_status": status,
        "approval": approval,
        "ready_for_write_to_main_pipeline": False,
        "next_allowed_phase": (
            "Phase 2E write simulation design"
            if status == "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY"
            else "Human candidate review"
        ),
        "blocking_reasons": blocking_reasons,
        "safety_notes": [
            "This gate does not write to the main pipeline.",
            "This gate does not write to a database.",
            "Approval here only unlocks the next design/simulation phase.",
            "ready_for_write_to_main_pipeline remains false.",
        ],
    }
    return manifest


def write_csv(output_dir: Path, manifest: Dict[str, Any]) -> None:
    csv_path = output_dir / "commit_gate_summary.csv"
    rows = [
        ("commit_gate_status", manifest.get("commit_gate_status")),
        ("ready_for_write_to_main_pipeline", manifest.get("ready_for_write_to_main_pipeline")),
        ("next_allowed_phase", manifest.get("next_allowed_phase")),
        ("candidate_items_count", manifest.get("candidate_summary", {}).get("candidate_items_count", 0)),
        ("skipped_items_count", manifest.get("candidate_summary", {}).get("skipped_items_count", 0)),
        ("duplicate_write_key_count", manifest.get("candidate_summary", {}).get("duplicate_write_key_count", 0)),
        ("duplicate_material_code_count", manifest.get("candidate_summary", {}).get("duplicate_material_code_count", 0)),
        ("approval_approved", manifest.get("approval", {}).get("approved", False)),
        ("approved_by", manifest.get("approval", {}).get("approved_by", "")),
        ("blocking_reasons", "; ".join(manifest.get("blocking_reasons", []))),
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["metric", "value"])
        writer.writerows(rows)


def autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 80)


def write_xlsx(output_dir: Path, manifest: Dict[str, Any]) -> None:
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    danger_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ok_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws = wb.active
    ws.title = "Gate Summary"
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    rows = [
        ("commit_gate_status", manifest.get("commit_gate_status")),
        ("ready_for_write_to_main_pipeline", manifest.get("ready_for_write_to_main_pipeline")),
        ("next_allowed_phase", manifest.get("next_allowed_phase")),
        ("candidate_items_count", manifest.get("candidate_summary", {}).get("candidate_items_count", 0)),
        ("skipped_items_count", manifest.get("candidate_summary", {}).get("skipped_items_count", 0)),
        ("duplicate_write_key_count", manifest.get("candidate_summary", {}).get("duplicate_write_key_count", 0)),
        ("duplicate_material_code_count", manifest.get("candidate_summary", {}).get("duplicate_material_code_count", 0)),
        ("approval_approved", manifest.get("approval", {}).get("approved", False)),
        ("approved_by", manifest.get("approval", {}).get("approved_by", "")),
        ("blocking_reasons", "; ".join(manifest.get("blocking_reasons", []))),
    ]
    for row in rows:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2):
        row[0].font = bold_font
        row[0].border = border
        row[1].border = border
        if row[0].value == "commit_gate_status":
            if row[1].value == "BLOCKED_NEEDS_REVIEW":
                row[1].fill = danger_fill
            elif row[1].value == "PENDING_HUMAN_APPROVAL":
                row[1].fill = warning_fill
            else:
                row[1].fill = ok_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit(ws)

    ws_hash = wb.create_sheet("Artifact Hashes")
    ws_hash.append(["Artifact", "SHA256"])
    for cell in ws_hash[1]:
        cell.fill = header_fill
        cell.font = header_font
    for name, digest in manifest.get("source_artifact_hashes", {}).items():
        ws_hash.append([name, digest])
    ws_hash.freeze_panes = "A2"
    ws_hash.auto_filter.ref = ws_hash.dimensions
    autofit(ws_hash)

    ws_check = wb.create_sheet("Manual Checklist")
    ws_check.append(["Check", "Required Result", "Reviewer Result"])
    for cell in ws_check[1]:
        cell.fill = header_fill
        cell.font = header_font
    checks = [
        ("Open write_candidate_review.xlsx", "Reviewed by human", ""),
        ("candidate_items_count > 0", "True", ""),
        ("duplicate_write_key_count = 0", "True", ""),
        ("No unresolved NEEDS_INVESTIGATION", "True", ""),
        ("ready_for_write_to_main_pipeline = False", "True", ""),
        ("Backup/rollback plan documented", "True", ""),
        ("Approval phrase understood", APPROVAL_PHRASE, ""),
    ]
    for row in checks:
        ws_check.append(list(row))
    ws_check.freeze_panes = "A2"
    ws_check.auto_filter.ref = ws_check.dimensions
    autofit(ws_check)

    wb.save(output_dir / "profile_commit_gate_review.xlsx")


def build_checklist_md(manifest: Dict[str, Any]) -> str:
    blocking = manifest.get("blocking_reasons", [])
    blocking_text = "\n".join(f"- `{reason}`" for reason in blocking) if blocking else "- None"
    return f"""# Profile Commit Gate / Manual Approval Lock

## Status

- Commit gate status: `{manifest['commit_gate_status']}`
- Ready for write to main pipeline: `FALSE`
- Next allowed phase: `{manifest['next_allowed_phase']}`

## Approval

- Approved: `{manifest['approval']['approved']}`
- Approved by: `{manifest['approval']['approved_by'] or 'N/A'}`
- Approved at: `{manifest['approval']['approved_at'] or 'N/A'}`

## Blocking Reasons

{blocking_text}

## Manual Checklist Before Any Future Write Phase

- Open and review `feasibility_outputs/profile_write_candidate/write_candidate_review.xlsx`.
- Confirm `candidate_items_count > 0`.
- Confirm `duplicate_write_key_count = 0`.
- Confirm there are no unresolved `NEEDS_INVESTIGATION` or rejected candidates in the write set.
- Confirm `ready_for_write_to_main_pipeline = FALSE`.
- Confirm backup and rollback strategy exists.
- If approving for the next design/simulation phase, use the exact phrase:
  `{APPROVAL_PHRASE}`

## Safety

This phase does not write to the main pipeline, database, or official normalized package. Approval only unlocks the next design/simulation phase.
"""


def export_commit_gate(
    candidate_dir: Path,
    output_dir: Path,
    approve: bool = False,
    approved_by: str = "",
    approval_note: str = "",
    approval_phrase: str = "",
) -> Dict[str, Any]:
    manifest = build_gate_manifest(
        candidate_dir=candidate_dir,
        approve=approve,
        approved_by=approved_by,
        approval_note=approval_note,
        approval_phrase=approval_phrase,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "profile_commit_gate_manifest.json", manifest)
    write_csv(output_dir, manifest)
    write_xlsx(output_dir, manifest)
    (output_dir / "profile_commit_gate_checklist.md").write_text(
        build_checklist_md(manifest),
        encoding="utf-8",
    )
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export profile commit gate approval lock.")
    parser.add_argument("--candidate-dir", type=Path, default=DEFAULT_CANDIDATE_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--approve", action="store_true")
    parser.add_argument("--approved-by", default="")
    parser.add_argument("--approval-note", default="")
    parser.add_argument("--approval-phrase", default="")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = export_commit_gate(
        candidate_dir=args.candidate_dir,
        output_dir=args.output_dir,
        approve=args.approve,
        approved_by=args.approved_by,
        approval_note=args.approval_note,
        approval_phrase=args.approval_phrase,
    )
    print(json.dumps({
        "commit_gate_status": manifest["commit_gate_status"],
        "ready_for_write_to_main_pipeline": manifest["ready_for_write_to_main_pipeline"],
        "next_allowed_phase": manifest["next_allowed_phase"],
        "blocking_reasons": manifest["blocking_reasons"],
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
