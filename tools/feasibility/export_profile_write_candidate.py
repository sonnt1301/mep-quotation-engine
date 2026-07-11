# -*- coding: utf-8 -*-
"""Controlled Write Candidate Exporter and Duplicate Safety Planner.

This script parses the dry-run write adapter preview, performs safety checks
on duplicate write keys and material codes, assigns candidate IDs and proposed actions,
and exports Excel, CSV, JSON and Markdown summaries without writing to the main pipeline.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ADAPTER_DIR = PROJECT_ROOT / "feasibility_outputs/profile_write_adapter_dry_run"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def analyze_duplicates_and_warnings(
    items: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, int]]:
    candidates: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    
    seen_write_keys: Dict[str, List[str]] = {}
    seen_material_codes: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    
    # 1. Phân nhóm để phát hiện trùng lặp
    for idx, item in enumerate(items, start=1):
        item_id = item.get("item_id", f"ADAPTER_ITEM_{idx:04d}")
        write_key = item.get("write_key", "").strip()
        supplier = item.get("supplier_code", "").strip()
        material_code = item.get("material_code", "").strip()
        
        # Nhóm theo write_key
        if write_key:
            if write_key not in seen_write_keys:
                seen_write_keys[write_key] = []
            seen_write_keys[write_key].append(item_id)
            
        # Nhóm theo supplier + material_code
        key_tuple = (supplier, material_code)
        if supplier and material_code:
            if key_tuple not in seen_material_codes:
                seen_material_codes[key_tuple] = []
            seen_material_codes[key_tuple].append(item)

    # 2. Xử lý gán proposed_action và phát hiện trùng write_key
    duplicate_write_key_count = 0
    skipped_items_count = 0
    
    processed_write_keys = set()
    
    for idx, item in enumerate(items, start=1):
        item_id = item.get("item_id", f"ADAPTER_ITEM_{idx:04d}")
        write_key = item.get("write_key", "").strip()
        
        cand_id = f"WRITE_CANDIDATE_{idx:04d}"
        
        # Mặc định là INSERT_CANDIDATE do chưa đối chiếu database chính
        proposed_action = "INSERT_CANDIDATE"
        human_note = item.get("human_note", "")
        
        # Kiểm tra trùng write_key
        if write_key and len(seen_write_keys[write_key]) > 1:
            if write_key in processed_write_keys:
                proposed_action = "SKIP_CANDIDATE"
                human_note = f"[SYSTEM WARNING: Trùng khóa ghi write_key] {human_note}".strip()
                skipped_items_count += 1
            else:
                processed_write_keys.add(write_key)
                # Dòng đầu tiên vẫn giữ INSERT_CANDIDATE nhưng cảnh báo
                human_note = f"[SYSTEM WARNING: Dòng đầu tiên của nhóm trùng khóa ghi] {human_note}".strip()
        
        candidates.append({
            "write_candidate_id": cand_id,
            "source_adapter_item_id": item_id,
            "supplier_code": item.get("supplier_code"),
            "material_code": item.get("material_code"),
            "description": item.get("description"),
            "unit": item.get("unit"),
            "quantity": item.get("quantity"),
            "unit_price": item.get("unit_price"),
            "amount": item.get("amount"),
            "currency": item.get("currency"),
            "page_number": item.get("page_number"),
            "write_key": write_key,
            "human_decision": item.get("human_decision"),
            "human_note": human_note,
            "provenance": item.get("provenance"),
            "evidence_text": item.get("evidence_text"),
            "proposed_action": proposed_action
        })

    # Ghi nhận warnings trùng write_key
    for w_key, item_ids in seen_write_keys.items():
        if len(item_ids) > 1:
            duplicate_write_key_count += 1
            warnings.append({
                "type": "DUPLICATE_WRITE_KEY",
                "write_key": w_key,
                "material_code": "",
                "message": f"Khóa ghi '{w_key}' xuất hiện {len(item_ids)} lần trong danh sách preview. Các dòng thừa sẽ bị SKIP.",
                "items": item_ids
            })

    # 3. Kiểm tra trùng material_code nhưng khác đơn giá/mô tả
    duplicate_material_code_count = 0
    for (supplier, code), group_items in seen_material_codes.items():
        if len(group_items) > 1:
            # Kiểm tra xem có sự khác biệt về đơn giá hoặc mô tả không
            prices = {it.get("unit_price") for it in group_items}
            descs = {it.get("description", "").strip() for it in group_items}
            
            if len(prices) > 1 or len(descs) > 1:
                duplicate_material_code_count += 1
                item_ids = [it.get("item_id") for it in group_items]
                warnings.append({
                    "type": "DUPLICATE_MATERIAL_CODE_MISMATCH",
                    "write_key": "",
                    "material_code": code,
                    "message": f"Mã hàng '{code}' của {supplier} xuất hiện {len(group_items)} lần với đơn giá hoặc mô tả khác nhau. Cần kiểm tra thủ công, cấm tự động gộp dòng.",
                    "items": item_ids
                })

    counts = {
        "candidate_items_count": len([c for c in candidates if c["proposed_action"] != "SKIP_CANDIDATE"]),
        "skipped_items_count": skipped_items_count,
        "duplicate_write_key_count": duplicate_write_key_count,
        "duplicate_material_code_count": duplicate_material_code_count
    }
    return candidates, warnings, counts


def write_csv_and_xlsx(
    output_dir: Path,
    summary: Dict[str, Any],
    candidates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]]
) -> None:
    # 1. Ghi CSV write_candidate_review.csv
    csv_path = output_dir / "write_candidate_review.csv"
    headers_cand = [
        "write_candidate_id", "source_adapter_item_id", "supplier_code", "material_code",
        "description", "unit", "quantity", "unit_price", "amount", "currency",
        "page_number", "write_key", "human_decision", "human_note", "proposed_action"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_cand)
        writer.writeheader()
        for item in candidates:
            row = {k: item.get(k, "") for k in headers_cand}
            writer.writerow(row)

    # 2. Ghi Excel write_candidate_review.xlsx
    xlsx_path = output_dir / "write_candidate_review.xlsx"
    wb = openpyxl.Workbook()
    
    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Xanh dương đậm
    fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("candidate_items_count", summary.get("candidate_items_count", 0)),
        ("skipped_items_count", summary.get("skipped_items_count", 0)),
        ("duplicate_write_key_count", summary.get("duplicate_write_key_count", 0)),
        ("duplicate_material_code_count", summary.get("duplicate_material_code_count", 0)),
        ("proposed_status", summary.get("proposed_status", "N/A")),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_warning
        else:
            val_cell.alignment = align_left
            if val == "BLOCKED_NEEDS_REVIEW":
                val_cell.fill = fill_danger

    # Sheet 2: Write Candidates
    ws_cand = wb.create_sheet(title="Write Candidates")
    ws_cand.views.sheetView[0].showGridLines = True
    ws_cand.freeze_panes = "A2"
    
    ws_cand.append(headers_cand)
    for col_idx in range(1, len(headers_cand) + 1):
        cell = ws_cand.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, item in enumerate(candidates, start=2):
        row_data = [item.get(k, "") for k in headers_cand]
        ws_cand.append(row_data)
        
        p_act = item.get("proposed_action", "")
        
        for col_idx, k in enumerate(headers_cand, start=1):
            cell = ws_cand.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["quantity", "unit_price", "amount", "page_number"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["write_candidate_id", "source_adapter_item_id", "supplier_code", "currency", "proposed_action"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if p_act == "SKIP_CANDIDATE":
                cell.fill = fill_warning

    ws_cand.auto_filter.ref = f"A1:{get_column_letter(len(headers_cand))}{len(candidates) + 1}"

    # Sheet 3: Warnings / Duplicates (Dùng & thay vì / do Excel cấm ký tự /)
    ws_warn = wb.create_sheet(title="Warnings & Duplicates")
    ws_warn.views.sheetView[0].showGridLines = True
    ws_warn.freeze_panes = "A2"
    
    headers_warn = ["warning_type", "write_key", "material_code", "message", "related_item_ids"]
    ws_warn.append(headers_warn)
    for col_idx in range(1, len(headers_warn) + 1):
        cell = ws_warn.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, warn in enumerate(warnings, start=2):
        row_data = [
            warn.get("type", ""),
            warn.get("write_key", ""),
            warn.get("material_code", ""),
            warn.get("message", ""),
            ", ".join(warn.get("items", []))
        ]
        ws_warn.append(row_data)
        
        for col_idx in range(1, len(headers_warn) + 1):
            cell = ws_warn.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_danger
            if col_idx in [1, 2, 3]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_warn.auto_filter.ref = f"A1:{get_column_letter(len(headers_warn))}{len(warnings) + 1}"

    # Auto-fit column widths
    for ws in [ws_sum, ws_cand, ws_warn]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_commit_plan(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Controlled Commit Plan – Phase 2C

Tài liệu này định nghĩa kế hoạch ghi nhận có kiểm soát (Controlled Commit Plan) và phân tích rủi ro an toàn cho Write Candidates trước khi tiến hành Phase 2D (ghi thật vào main pipeline).

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN CHƯA GHI THẬT**
> * Kế hoạch này **CHƯA** thực hiện ghi bất kỳ dữ liệu nào vào database hoặc main production pipeline của dự án.
> * Trạng thái an toàn: `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Trạng Thái Của Đệ Trình (Proposed Status)

* Proposed status: `{summary['proposed_status']}`
* Ready for write to main pipeline: `FALSE`
* Lý do chưa sẵn sàng ghi thật:
{chr(10).join(f"  - {r}" for r in summary['reasons_not_ready_for_write'])}

## 2. Thống Kê Số Lượng Candidates

* Số lượng ghi dự kiến (candidate_items_count): `{summary['candidate_items_count']}`
* Số lượng bị loại bỏ (skipped_items_count): `{summary['skipped_items_count']}`
* Số lượng trùng khóa ghi (duplicate_write_key_count): `{summary['duplicate_write_key_count']}`
* Số lượng trùng mã hàng cảnh báo (duplicate_material_code_count): `{summary['duplicate_material_code_count']}`

## 3. Điều Kiện Để Tiến Hành Ghi Thật (Phase 2D Commit Criteria)

Để được phép chuyển đổi biến `ready_for_write_to_main_pipeline` sang `TRUE` và tiến hành ghi dữ liệu thật, hệ thống bắt buộc phải thỏa mãn các tiêu chí sau:

- [ ] **Tiêu chí 1: Có Candidate hợp lệ**
  - Số lượng `candidate_items_count` phải lớn hơn 0.
- [ ] **Tiêu chí 2: Triệt tiêu hoàn toàn trùng lặp khóa**
  - Số lượng `duplicate_write_key_count` phải bằng 0 (hoặc tất cả các dòng trùng lặp phải được con người duyệt qua và phân xử thủ công trong file Excel).
- [ ] **Tiêu chí 3: Không còn dòng ở trạng thái rủi ro**
  - Không có dòng nào của đệ trình đang ở trạng thái `NEEDS_INVESTIGATION` hoặc `REJECT` chưa được phân xử.
- [ ] **Tiêu chí 4: Human Approval trên file Excel**
  - Người dùng / Reviewer đã mở file [write_candidate_review.xlsx](file:///{output_dir.resolve().as_posix()}/write_candidate_review.xlsx) kiểm tra trực quan các sheet và ký duyệt chấp thuận kế hoạch.
- [ ] **Tiêu chí 5: Có phương án Backup và Phục hồi (Rollback Plan)**
  - Đã thực hiện sao lưu trạng thái database hiện tại của main pipeline trước khi chạy lệnh ghi.

---

## 4. Các Tệp Tin Preview Cục Bộ
* Tệp Candidates JSON: `{output_dir / 'write_candidate_items.json'}`
* Tệp Summary JSON: `{output_dir / 'write_candidate_summary.json'}`
* Bảng Excel Review: `{output_dir / 'write_candidate_review.xlsx'}`
"""


def run_candidate_pipeline(
    adapter_preview_path: Path,
    adapter_summary_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    preview_items = load_json(adapter_preview_path, default=[])
    adapter_summary = load_json(adapter_summary_path, default={})
    
    candidates, warnings, counts = analyze_duplicates_and_warnings(preview_items)
    
    # Quyết định proposed_status
    has_candidate = counts["candidate_items_count"] > 0
    has_duplicates = counts["duplicate_write_key_count"] > 0 or counts["duplicate_material_code_count"] > 0
    
    reasons = []
    if not has_candidate:
        reasons.append("Không có bất kỳ vật tư write candidate hợp lệ nào để ghi.")
    if has_duplicates:
        reasons.append(f"Phát hiện trùng lặp khóa ghi (trùng write_key: {counts['duplicate_write_key_count']}, trùng material_code: {counts['duplicate_material_code_count']}).")
    reasons.append("Chưa có chữ ký số hoặc Human Approval duyệt file Excel Controlled Commit.")
    
    proposed_status = (
        "READY_FOR_HUMAN_COMMIT_REVIEW"
        if has_candidate and not has_duplicates
        else "BLOCKED_NEEDS_REVIEW"
    )

    summary = {
        "candidate_items_count": counts["candidate_items_count"],
        "skipped_items_count": counts["skipped_items_count"],
        "duplicate_write_key_count": counts["duplicate_write_key_count"],
        "duplicate_material_code_count": counts["duplicate_material_code_count"],
        "proposed_status": proposed_status,
        "ready_for_write_to_main_pipeline": False,
        "reasons_not_ready_for_write": reasons
    }
    
    manifest = {
        "candidate_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "dry_run",
        "source_adapter_summary_path": str(adapter_summary_path),
        "write_target": "write_candidate_preview_only",
        "ready_for_write_to_main_pipeline": False,
        "summary": summary,
        "candidates": candidates,
        "warnings": warnings
    }
    
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "write_candidate_items.json", candidates)
    write_json(output_dir / "write_candidate_summary.json", manifest)
    
    # Kết xuất CSV và XLSX
    write_csv_and_xlsx(output_dir, summary, candidates, warnings)
    
    # Kết xuất Commit Plan
    commit_plan_content = build_commit_plan(summary, output_dir)
    (output_dir / "write_candidate_commit_plan.md").write_text(commit_plan_content, encoding="utf-8")
    
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run write candidate packager.")
    parser.add_argument("--adapter-preview", type=Path, default=DEFAULT_ADAPTER_DIR / "normalized_items_preview.json")
    parser.add_argument("--adapter-summary", type=Path, default=DEFAULT_ADAPTER_DIR / "profile_write_adapter_summary.json")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_candidate_pipeline(
        adapter_preview_path=args.adapter_preview,
        adapter_summary_path=args.adapter_summary,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
