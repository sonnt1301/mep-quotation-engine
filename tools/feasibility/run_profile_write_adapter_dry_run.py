# -*- coding: utf-8 -*-
"""Controlled write-adapter dry-run for Profile Bridge human review decisions.

This script intentionally does not write to the main pipeline, database, or any
official normalized package. It creates a reviewable normalized preview only.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_BRIDGE_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_bridge_dry_run/profile_bridge_items.json"
DEFAULT_DECISIONS = PROJECT_ROOT / "feasibility_outputs/profile_bridge_human_review/profile_bridge_review_decisions.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_write_adapter_dry_run"

ALLOWED_DECISIONS = {"APPROVE", "EDIT_AND_APPROVE", "ACCEPT_WITH_LIMITATION"}
BLOCKING_DECISIONS = {"REJECT", "NEEDS_INVESTIGATION"}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def build_review_item_key(item: Dict[str, Any]) -> str:
    supplier = str(item.get("supplier_code", "")).strip()
    page = item.get("source_page", 1)
    code = str(item.get("normalized_material_code", "")).strip()
    price = item.get("unit_price", 0)
    return f"{supplier}_P{page}_{code}_{price}"


def normalize_decisions(raw_decisions: Any) -> Dict[str, Dict[str, Any]]:
    """Support the Streamlit dict format and a future list format."""
    if isinstance(raw_decisions, dict):
        return {
            str(key): value
            for key, value in raw_decisions.items()
            if isinstance(value, dict)
        }

    if isinstance(raw_decisions, list):
        result: Dict[str, Dict[str, Any]] = {}
        for value in raw_decisions:
            if not isinstance(value, dict):
                continue
            key = value.get("review_item_key")
            if key:
                result[str(key)] = value
        return result

    return {}


def decision_type(decision: Dict[str, Any] | None) -> str:
    if not decision:
        return "UNREVIEWED"
    return str(decision.get("decision", "UNREVIEWED")).strip().upper()


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def make_write_key(item: Dict[str, Any], decision: Dict[str, Any]) -> str:
    code = decision.get("normalized_material_code_reviewed") or item.get("normalized_material_code")
    page = item.get("source_page")
    price = decision.get("unit_price_reviewed") or item.get("unit_price")
    desc = decision.get("description_reviewed") or item.get("description") or ""
    desc_sig = "".join(ch for ch in str(desc).upper() if ch.isalnum())[:24]
    return f"{item.get('supplier_code')}|{code}|P{page}|{price}|{desc_sig}"


def build_normalized_preview_item(
    item: Dict[str, Any],
    decision: Dict[str, Any],
    seq: int,
    quotation_id: str,
) -> Dict[str, Any]:
    reviewed_code = decision.get("normalized_material_code_reviewed") or item.get("normalized_material_code")
    reviewed_desc = decision.get("description_reviewed") or item.get("description")
    reviewed_unit = decision.get("unit_reviewed") or item.get("unit")
    reviewed_qty = safe_float(decision.get("quantity_reviewed"), 1.0)
    reviewed_price = safe_int(decision.get("unit_price_reviewed"), safe_int(item.get("unit_price")))
    reviewed_amount = safe_float(decision.get("amount_reviewed"), reviewed_qty * reviewed_price)
    reviewed_currency = (decision.get("currency_reviewed") or item.get("currency") or "VND").upper()

    return {
        "item_id": f"{quotation_id}_ITEM_{seq:04d}",
        "source_draft_item_id": f"PROFILE_BRIDGE_{build_review_item_key(item)}",
        "source_review_decision_id": decision.get("review_decision_id", ""),
        "supplier_code": item.get("supplier_code"),
        "material_code": reviewed_code,
        "material_name": reviewed_desc,
        "description": reviewed_desc,
        "unit": reviewed_unit,
        "quantity": reviewed_qty,
        "unit_price": reviewed_price,
        "amount": reviewed_amount,
        "currency": reviewed_currency,
        "page_number": item.get("source_page"),
        "brand": item.get("supplier_code"),
        "category": item.get("product_family"),
        "raw_text": item.get("source_evidence_text"),
        "evidence_text": item.get("source_evidence_text"),
        "confidence": 1.0 if decision_type(decision) in {"APPROVE", "EDIT_AND_APPROVE"} else 0.85,
        "reviewer": decision.get("reviewer", "human_reviewer"),
        "source_layout_name": item.get("source_layout_name"),
        "source_bridge_key": build_review_item_key(item),
        "write_key": make_write_key(item, decision),
        "human_decision": decision_type(decision),
        "human_note": decision.get("human_note", ""),
        "reviewed_at": decision.get("reviewed_at"),
        "provenance": decision.get("provenance") or item.get("provenance"),
        "warnings": (
            [{"code": "accepted_with_known_limitation", "message": decision.get("human_note", "")}]
            if decision_type(decision) == "ACCEPT_WITH_LIMITATION"
            else []
        ),
    }


def block_item(item: Dict[str, Any], reason: str, decision: Dict[str, Any] | None = None) -> Dict[str, Any]:
    return {
        "review_item_key": build_review_item_key(item),
        "supplier_code": item.get("supplier_code"),
        "source_page": item.get("source_page"),
        "normalized_material_code": item.get("normalized_material_code"),
        "unit_price": item.get("unit_price"),
        "decision": decision_type(decision),
        "block_reason": reason,
        "human_note": (decision or {}).get("human_note", ""),
    }


def convert_reviewed_items(
    bridge_items: Iterable[Dict[str, Any]],
    decisions: Dict[str, Dict[str, Any]],
    quotation_id: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, int]]:
    exported: List[Dict[str, Any]] = []
    blocked: List[Dict[str, Any]] = []
    counts = {
        "approved_count": 0,
        "edited_count": 0,
        "accepted_with_limitation_count": 0,
        "rejected_count": 0,
        "needs_investigation_count": 0,
        "unreviewed_count": 0,
    }

    for item in bridge_items:
        key = build_review_item_key(item)
        decision = decisions.get(key)
        dtype = decision_type(decision)

        if dtype == "APPROVE":
            counts["approved_count"] += 1
            exported.append(build_normalized_preview_item(item, decision or {}, len(exported) + 1, quotation_id))
        elif dtype == "EDIT_AND_APPROVE":
            counts["edited_count"] += 1
            exported.append(build_normalized_preview_item(item, decision or {}, len(exported) + 1, quotation_id))
        elif dtype == "ACCEPT_WITH_LIMITATION":
            counts["accepted_with_limitation_count"] += 1
            exported.append(build_normalized_preview_item(item, decision or {}, len(exported) + 1, quotation_id))
        elif dtype == "REJECT":
            counts["rejected_count"] += 1
            blocked.append(block_item(item, "human_rejected", decision))
        elif dtype == "NEEDS_INVESTIGATION":
            counts["needs_investigation_count"] += 1
            blocked.append(block_item(item, "needs_investigation", decision))
        else:
            counts["unreviewed_count"] += 1
            blocked.append(block_item(item, "missing_human_approval", decision))

    return exported, blocked, counts


def build_report(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Profile Write Adapter Dry-run Report

## Status

- Proposed status: `{summary['proposed_status']}`
- Ready for write to main pipeline: `FALSE`
- Write target: `normalized_json_preview_only`

## Counts

- Input bridge items: `{summary['input_bridge_items_count']}`
- Review decisions: `{summary['review_decisions_count']}`
- Exportable preview items: `{summary['exportable_items_count']}`
- Blocked items: `{summary['blocked_items_count']}`
- Approved: `{summary['approved_count']}`
- Edited and approved: `{summary['edited_count']}`
- Accepted with limitation: `{summary['accepted_with_limitation_count']}`
- Rejected: `{summary['rejected_count']}`
- Needs investigation: `{summary['needs_investigation_count']}`
- Unreviewed: `{summary['unreviewed_count']}`

## Output

- Normalized preview: `{output_dir / 'normalized_items_preview.json'}`
- Blocked items: `{output_dir / 'blocked_items.json'}`
- Summary: `{output_dir / 'profile_write_adapter_summary.json'}`

## Safety

This dry-run does not write to the main pipeline, database, or official normalized package. It only creates reviewable preview artifacts.
"""


def write_csv_and_xlsx(
    output_dir: Path,
    summary: Dict[str, Any],
    exported: List[Dict[str, Any]],
    blocked: List[Dict[str, Any]],
) -> None:
    # 1. Ghi CSV normalized_items_preview.csv
    csv_exp_path = output_dir / "normalized_items_preview.csv"
    headers_exp = [
        "item_id", "supplier_code", "material_code", "description", "unit",
        "quantity", "unit_price", "amount", "currency", "page_number",
        "human_decision", "human_note", "write_key", "provenance", "evidence_text"
    ]
    with open(csv_exp_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_exp)
        writer.writeheader()
        for item in exported:
            row = {k: item.get(k, "") for k in headers_exp}
            writer.writerow(row)

    # 2. Ghi CSV blocked_items.csv
    csv_blk_path = output_dir / "blocked_items.csv"
    headers_blk = [
        "review_item_key", "supplier_code", "source_page", "normalized_material_code",
        "unit_price", "decision", "block_reason", "human_note"
    ]
    with open(csv_blk_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_blk)
        writer.writeheader()
        for item in blocked:
            row = {k: item.get(k, "") for k in headers_blk}
            writer.writerow(row)

    # 3. Ghi Excel profile_write_adapter_review.xlsx
    xlsx_path = output_dir / "profile_write_adapter_review.xlsx"
    wb = openpyxl.Workbook()
    
    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Xanh dương đậm
    fill_warning = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Cam nhạt cho ACCEPT_WITH_LIMITATION
    fill_danger = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Đỏ nhạt cho REJECT/NEEDS_INVESTIGATION
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("input_bridge_items_count", summary.get("input_bridge_items_count", 0)),
        ("review_decisions_count", summary.get("review_decisions_count", 0)),
        ("exportable_items_count", summary.get("exportable_items_count", 0)),
        ("blocked_items_count", summary.get("blocked_items_count", 0)),
        ("approved_count", summary.get("approved_count", 0)),
        ("edited_count", summary.get("edited_count", 0)),
        ("accepted_with_limitation_count", summary.get("accepted_with_limitation_count", 0)),
        ("rejected_count", summary.get("rejected_count", 0)),
        ("needs_investigation_count", summary.get("needs_investigation_count", 0)),
        ("unreviewed_count", summary.get("unreviewed_count", 0)),
        ("proposed_status", summary.get("proposed_status", "N/A")),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_warning
        else:
            val_cell.alignment = align_left
            if val == "BLOCKED_NEEDS_HUMAN_REVIEW":
                val_cell.fill = fill_danger

    # Sheet 2: Exportable Preview
    ws_exp = wb.create_sheet(title="Exportable Preview")
    ws_exp.views.sheetView[0].showGridLines = True
    ws_exp.freeze_panes = "A2"
    
    # Ghi header
    ws_exp.append(headers_exp)
    for col_idx in range(1, len(headers_exp) + 1):
        cell = ws_exp.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Ghi data rows
    for r_idx, item in enumerate(exported, start=2):
        row_data = [item.get(k, "") for k in headers_exp]
        ws_exp.append(row_data)
        
        # Format các ô
        for col_idx, k in enumerate(headers_exp, start=1):
            cell = ws_exp.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            # Format số
            if k in ["quantity", "unit_price", "amount", "page_number"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["item_id", "supplier_code", "material_code", "currency", "human_decision"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            # Highlight
            h_dec = item.get("human_decision", "")
            if h_dec == "ACCEPT_WITH_LIMITATION":
                cell.fill = fill_warning

    ws_exp.auto_filter.ref = f"A1:{get_column_letter(len(headers_exp))}{len(exported) + 1}"

    # Sheet 3: Blocked Items
    ws_blk = wb.create_sheet(title="Blocked Items")
    ws_blk.views.sheetView[0].showGridLines = True
    ws_blk.freeze_panes = "A2"
    
    # Ghi header
    ws_blk.append(headers_blk)
    for col_idx in range(1, len(headers_blk) + 1):
        cell = ws_blk.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Ghi data rows
    for r_idx, item in enumerate(blocked, start=2):
        row_data = [item.get(k, "") for k in headers_blk]
        ws_blk.append(row_data)
        
        decision = item.get("decision", "")
        is_high_risk = decision in ["REJECT", "NEEDS_INVESTIGATION"]
        
        # Format các ô
        for col_idx, k in enumerate(headers_blk, start=1):
            cell = ws_blk.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            # Format số
            if k in ["source_page", "unit_price"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["supplier_code", "decision"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            # Highlight
            if is_high_risk:
                cell.fill = fill_danger

    ws_blk.auto_filter.ref = f"A1:{get_column_letter(len(headers_blk))}{len(blocked) + 1}"

    # Auto-fit column widths cho cả 3 sheet
    for ws in [ws_sum, ws_exp, ws_blk]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def run_adapter(
    bridge_items_path: Path,
    decisions_path: Path,
    output_dir: Path,
    quotation_id: str,
) -> Dict[str, Any]:
    bridge_items = load_json(bridge_items_path, default=[])
    raw_decisions = load_json(decisions_path, default={})
    decisions = normalize_decisions(raw_decisions)

    exported, blocked, counts = convert_reviewed_items(bridge_items, decisions, quotation_id)
    proposed_status = (
        "READY_FOR_CONTROLLED_WRITE_REVIEW"
        if exported and not counts["needs_investigation_count"]
        else "BLOCKED_NEEDS_HUMAN_REVIEW"
    )

    summary = {
        "input_bridge_items_count": len(bridge_items),
        "review_decisions_count": len(decisions),
        "exportable_items_count": len(exported),
        "blocked_items_count": len(blocked),
        **counts,
        "proposed_status": proposed_status,
    }

    manifest = {
        "adapter_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "dry_run",
        "source_bridge_items": str(bridge_items_path),
        "source_review_decisions": str(decisions_path),
        "write_target": "normalized_json_preview_only",
        "ready_for_write_to_main_pipeline": False,
        "summary": summary,
        "items": exported,
        "blocked_items": blocked,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "normalized_items_preview.json", exported)
    write_json(output_dir / "blocked_items.json", blocked)
    write_json(output_dir / "profile_write_adapter_summary.json", manifest)
    
    # Ghi các file CSV và XLSX
    write_csv_and_xlsx(output_dir, summary, exported, blocked)
    
    (output_dir / "profile_write_adapter_report.md").write_text(
        build_report(summary, output_dir),
        encoding="utf-8",
    )
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run profile write adapter dry-run.")
    parser.add_argument("--bridge-items", type=Path, default=DEFAULT_BRIDGE_ITEMS)
    parser.add_argument("--decisions", type=Path, default=DEFAULT_DECISIONS)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--quotation-id", default="PROFILE_20260710_001")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_adapter(
        bridge_items_path=args.bridge_items,
        decisions_path=args.decisions,
        output_dir=args.output_dir,
        quotation_id=args.quotation_id,
    )
    print(json.dumps(manifest["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
