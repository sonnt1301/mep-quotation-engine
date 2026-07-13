# -*- coding: utf-8 -*-
"""Final Write Plan Draft Generator.

This script aggregates data from write candidates, simulation records,
master match dry-run results, and master review resolution to generate
a final write plan draft.
"""

from __future__ import annotations

import argparse
import json
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CAND_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate/write_candidate_items.json"
DEFAULT_CAND_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate/write_candidate_summary.json"
DEFAULT_SIM_RECORDS = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_material_records.json"
DEFAULT_SIM_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_write_summary.json"
DEFAULT_MATCH_RESULTS = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_results.json"
DEFAULT_MATCH_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_summary.json"
DEFAULT_RESOL_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_master_review_resolution/master_review_resolution_items.json"
DEFAULT_RESOL_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_master_review_resolution/master_review_resolution_summary.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_final_write_plan_draft"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def process_plan(
    candidates: List[Dict[str, Any]],
    simulations: List[Dict[str, Any]],
    matches: List[Dict[str, Any]],
    resolutions: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, int]]:
    plan_items: List[Dict[str, Any]] = []
    risks: List[Dict[str, Any]] = []

    # Map tra cứu nhanh
    sim_map = {s["write_candidate_id"]: s for s in simulations if s.get("write_candidate_id")}
    match_map = {m["write_candidate_id"]: m for m in matches if m.get("write_candidate_id")}
    resol_map = {r["write_candidate_id"]: r for r in resolutions if r.get("write_candidate_id")}

    plan_insert_count = 0
    plan_update_count = 0
    plan_skip_count = 0
    plan_blocked_count = 0
    
    high_risk_count = 0
    medium_risk_count = 0
    low_risk_count = 0

    for idx, cand in enumerate(candidates, start=1):
        plan_id = f"PLAN_ITEM_{idx:04d}"
        cand_id = cand.get("write_candidate_id", "")
        supplier = cand.get("supplier_code", "")
        material = cand.get("material_code", "")
        desc = cand.get("description", "")
        unit = cand.get("unit", "")
        qty = cand.get("quantity", 0)
        price = cand.get("unit_price", 0)
        amt = cand.get("amount", 0.0)
        curr = cand.get("currency", "VND")
        write_key = cand.get("write_key", "")
        prov = cand.get("provenance", "")
        ev_text = cand.get("evidence_text", "")
        
        # Tra cứu các id liên quan
        sim = sim_map.get(cand_id, {})
        sim_rec_id = sim.get("simulation_record_id")
        sim_res = sim.get("simulated_result")
        
        match = match_map.get(cand_id, {})
        match_res_id = match.get("match_result_id")
        match_status = match.get("match_status")
        rec_act = match.get("recommended_action")
        matched_master_id = match.get("matched_master_record_id")
        match_warns = match.get("warnings", [])
        
        resol = resol_map.get(cand_id, {})
        resol_id = resol.get("resolution_item_id")
        resol_dec = resol.get("human_resolution_decision")
        
        # 1. Mapping Action logic
        final_action = "PLAN_BLOCKED"
        action_source = "default_blocked"
        
        if resol_dec and resol_dec != "PENDING":
            if resol_dec == "CONFIRM_INSERT":
                final_action = "PLAN_INSERT"
                action_source = "resolution"
            elif resol_dec == "CONFIRM_UPDATE":
                final_action = "PLAN_UPDATE"
                action_source = "resolution"
            elif resol_dec == "CONFIRM_SKIP":
                final_action = "PLAN_SKIP"
                action_source = "resolution"
            elif resol_dec == "MARK_DUPLICATE":
                final_action = "PLAN_BLOCKED"
                action_source = "resolution"
            elif resol_dec == "NEEDS_MORE_INFO":
                final_action = "PLAN_BLOCKED"
                action_source = "resolution"
            elif resol_dec == "REJECT_CANDIDATE":
                final_action = "PLAN_SKIP"
                action_source = "resolution"
        else:
            if rec_act == "WOULD_INSERT":
                final_action = "PLAN_INSERT"
                action_source = "master_match"
            elif rec_act == "WOULD_UPDATE":
                final_action = "PLAN_UPDATE"
                action_source = "master_match"
            elif rec_act == "WOULD_SKIP":
                final_action = "PLAN_SKIP"
                action_source = "master_match"
            elif rec_act == "NEEDS_MASTER_REVIEW":
                final_action = "PLAN_BLOCKED"
                action_source = "master_match"
            elif rec_act == "BLOCKED":
                final_action = "PLAN_BLOCKED"
                action_source = "master_match"

        # Tăng counts cho action
        if final_action == "PLAN_INSERT":
            plan_insert_count += 1
        elif final_action == "PLAN_UPDATE":
            plan_update_count += 1
        elif final_action == "PLAN_SKIP":
            plan_skip_count += 1
        elif final_action == "PLAN_BLOCKED":
            plan_blocked_count += 1

        # 2. Risk Evaluation logic
        risk_level = "LOW"
        risk_reasons = []
        
        # Check High Risk
        if final_action == "PLAN_BLOCKED":
            risk_level = "HIGH"
            risk_reasons.append("Bản ghi bị chặn ở khâu lập kế hoạch.")
        if rec_act == "NEEDS_MASTER_REVIEW" and (not resol_dec or resol_dec == "PENDING"):
            risk_level = "HIGH"
            risk_reasons.append("Yêu cầu đối chiếu master chưa được reviewer giải quyết.")
        if resol_dec == "NEEDS_MORE_INFO":
            risk_level = "HIGH"
            risk_reasons.append("Reviewer đánh dấu cần thêm thông tin đối chiếu master.")
        if rec_act == "BLOCKED" or sim_res == "BLOCKED":
            risk_level = "HIGH"
            risk_reasons.append("Bản ghi bị chặn ở khâu simulation/matching upstream.")
        if match_warns:
            risk_level = "HIGH"
            risk_reasons.extend(match_warns)

        # Check Medium Risk if not High
        if risk_level != "HIGH":
            if match_status in ["POSSIBLE_DUPLICATE", "POSSIBLE_UPDATE"] and resol_dec in ["CONFIRM_INSERT", "CONFIRM_UPDATE", "CONFIRM_SKIP"]:
                risk_level = "MEDIUM"
                risk_reasons.append("Đã giải quyết trùng lặp/khác biệt mã hàng qua resolution gate.")
            elif cand.get("human_decision") == "ACCEPT_WITH_LIMITATION":
                risk_level = "MEDIUM"
                risk_reasons.append("Vật tư được duyệt kèm Known Limitations ở visual review.")

        if risk_level == "LOW":
            low_risk_count += 1
        elif risk_level == "MEDIUM":
            medium_risk_count += 1
        elif risk_level == "HIGH":
            high_risk_count += 1
            # Đăng ký risk register
            risk_id = f"RISK_{len(risks) + 1:04d}"
            risks.append({
                "risk_id": risk_id,
                "final_plan_item_id": plan_id,
                "risk_level": risk_level,
                "risk_reason": "; ".join(risk_reasons),
                "suggested_resolution": "Reviewer mở tệp Excel Resolution phân xử thủ công dòng này.",
                "blocking": True
            })

        plan_items.append({
            "final_plan_item_id": plan_id,
            "write_candidate_id": cand_id,
            "simulation_record_id": sim_rec_id,
            "match_result_id": match_res_id,
            "resolution_item_id": resol_id,
            "supplier_code": supplier,
            "material_code": material,
            "description": desc,
            "unit": unit,
            "quantity": qty,
            "unit_price": price,
            "amount": amt,
            "currency": curr,
            "write_key": write_key,
            "final_planned_action": final_action,
            "action_source": action_source,
            "matched_master_record_id": matched_master_id,
            "resolution_decision": resol_dec,
            "risk_level": risk_level,
            "risk_reasons": risk_reasons,
            "provenance": prov,
            "evidence_text": ev_text,
            "ready_for_execution": False
        })

    counts = {
        "input_candidate_count": len(candidates),
        "input_simulation_count": len(simulations),
        "input_match_count": len(matches),
        "input_resolution_count": len(resolutions),
        "plan_insert_count": plan_insert_count,
        "plan_update_count": plan_update_count,
        "plan_skip_count": plan_skip_count,
        "plan_blocked_count": plan_blocked_count,
        "high_risk_count": high_risk_count,
        "medium_risk_count": medium_risk_count,
        "low_risk_count": low_risk_count
    }
    return plan_items, risks, counts


def write_csv_and_xlsx(
    output_dir: Path,
    summary: Dict[str, Any],
    plan_items: List[Dict[str, Any]],
    risks: List[Dict[str, Any]]
) -> None:
    # 1. Ghi CSV final_write_plan_review.csv
    csv_path = output_dir / "final_write_plan_review.csv"
    headers_csv = [
        "final_plan_item_id", "write_candidate_id", "supplier_code", "material_code",
        "description", "unit", "quantity", "unit_price", "amount", "final_planned_action",
        "action_source", "risk_level"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_csv)
        writer.writeheader()
        for item in plan_items:
            row = {k: item.get(k, "") for k in headers_csv}
            writer.writerow(row)

    # 2. Ghi Excel final_write_plan_review.xlsx
    xlsx_path = output_dir / "final_write_plan_review.xlsx"
    wb = openpyxl.Workbook()
    
    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fill_insert = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Xanh lá nhạt
    fill_update = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Xanh dương nhạt
    fill_skip = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # Xám nhạt
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số kế hoạch (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("proposed_status", summary.get("proposed_status", "N/A")),
        ("input_candidate_count", summary.get("input_candidate_count", 0)),
        ("input_simulation_count", summary.get("input_simulation_count", 0)),
        ("input_match_count", summary.get("input_match_count", 0)),
        ("input_resolution_count", summary.get("input_resolution_count", 0)),
        ("plan_insert_count", summary.get("plan_insert_count", 0)),
        ("plan_update_count", summary.get("plan_update_count", 0)),
        ("plan_skip_count", summary.get("plan_skip_count", 0)),
        ("plan_blocked_count", summary.get("plan_blocked_count", 0)),
        ("high_risk_count", summary.get("high_risk_count", 0)),
        ("medium_risk_count", summary.get("medium_risk_count", 0)),
        ("low_risk_count", summary.get("low_risk_count", 0)),
        ("ready_for_execution", False),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_skip
        else:
            val_cell.alignment = align_left
            if "BLOCKED" in str(val):
                val_cell.fill = fill_danger

    # Sheet 2: Final Write Plan
    ws_plan = wb.create_sheet(title="Final Write Plan")
    ws_plan.views.sheetView[0].showGridLines = True
    ws_plan.freeze_panes = "A2"
    
    ws_plan.append(headers_csv)
    for col_idx in range(1, len(headers_csv) + 1):
        cell = ws_plan.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    plan_rows_added = 0
    for item in plan_items:
        act = item.get("final_planned_action")
        if act != "PLAN_BLOCKED":
            row_data = [item.get(k, "") for k in headers_csv]
            ws_plan.append(row_data)
            plan_rows_added += 1
            r_idx = plan_rows_added + 1
            
            fill_row = fill_insert if act == "PLAN_INSERT" else (fill_update if act == "PLAN_UPDATE" else fill_skip)
            
            for col_idx, k in enumerate(headers_csv, start=1):
                cell = ws_plan.cell(row=r_idx, column=col_idx)
                cell.font = font_regular
                cell.border = thin_border
                cell.fill = fill_row
                
                if k in ["quantity", "unit_price", "amount"]:
                    cell.number_format = "#,##0"
                    cell.alignment = align_right
                elif k in ["final_plan_item_id", "write_candidate_id", "supplier_code", "final_planned_action", "risk_level"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

    ws_plan.auto_filter.ref = f"A1:{get_column_letter(len(headers_csv))}{plan_rows_added + 1}"

    # Sheet 3: Blocked Items
    ws_blk = wb.create_sheet(title="Blocked Items")
    ws_blk.views.sheetView[0].showGridLines = True
    ws_blk.freeze_panes = "A2"
    
    ws_blk.append(headers_csv)
    for col_idx in range(1, len(headers_csv) + 1):
        cell = ws_blk.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    blk_rows_added = 0
    for item in plan_items:
        act = item.get("final_planned_action")
        if act == "PLAN_BLOCKED":
            row_data = [item.get(k, "") for k in headers_csv]
            ws_blk.append(row_data)
            blk_rows_added += 1
            r_idx = blk_rows_added + 1
            
            for col_idx, k in enumerate(headers_csv, start=1):
                cell = ws_blk.cell(row=r_idx, column=col_idx)
                cell.font = font_regular
                cell.border = thin_border
                cell.fill = fill_danger
                
                if k in ["quantity", "unit_price", "amount"]:
                    cell.number_format = "#,##0"
                    cell.alignment = align_right
                elif k in ["final_plan_item_id", "write_candidate_id", "supplier_code", "final_planned_action", "risk_level"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

    ws_blk.auto_filter.ref = f"A1:{get_column_letter(len(headers_csv))}{blk_rows_added + 1}"

    # Sheet 4: Risk Register
    ws_risk = wb.create_sheet(title="Risk Register")
    ws_risk.views.sheetView[0].showGridLines = True
    ws_risk.freeze_panes = "A2"
    
    headers_risk = ["risk_id", "final_plan_item_id", "risk_level", "risk_reason", "suggested_resolution", "blocking"]
    ws_risk.append(headers_risk)
    for col_idx in range(1, len(headers_risk) + 1):
        cell = ws_risk.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, rsk in enumerate(risks, start=2):
        row_data = [rsk.get(k, "") for k in headers_risk]
        ws_risk.append(row_data)
        
        for col_idx, k in enumerate(headers_risk, start=1):
            cell = ws_risk.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.fill = fill_danger
            
            if k in ["risk_id", "final_plan_item_id", "risk_level", "blocking"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_risk.auto_filter.ref = f"A1:{get_column_letter(len(headers_risk))}{len(risks) + 1}"

    # Sheet 5: Source Trace
    ws_trace = wb.create_sheet(title="Source Trace")
    ws_trace.views.sheetView[0].showGridLines = True
    ws_trace.freeze_panes = "A2"
    
    headers_trace = [
        "final_plan_item_id", "write_candidate_id", "simulation_record_id", "match_result_id",
        "resolution_item_id", "supplier_code", "material_code", "write_key",
        "matched_master_record_id", "resolution_decision", "provenance"
    ]
    ws_trace.append(headers_trace)
    for col_idx in range(1, len(headers_trace) + 1):
        cell = ws_trace.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, item in enumerate(plan_items, start=2):
        row_data = [item.get(k, "") for k in headers_trace]
        ws_trace.append(row_data)
        
        for col_idx, k in enumerate(headers_trace, start=1):
            cell = ws_trace.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["provenance", "write_key"]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center

    ws_trace.auto_filter.ref = f"A1:{get_column_letter(len(headers_trace))}{len(plan_items) + 1}"

    # Auto-fit column widths
    for ws in [ws_sum, ws_plan, ws_blk, ws_risk, ws_trace]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_final_write_plan_md(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Final Write Plan Draft – Phase 2H

Báo cáo kế hoạch ghi nhận vật tư cuối cùng ở dạng nháp (Final Write Plan Draft) tổng hợp và phân tích rủi ro từ candidates, simulation, matching, và resolution.

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN BẢN NHÁP**
> * Kế hoạch này **CHƯA** thực hiện bất kỳ hành động ghi nào vào cơ sở dữ liệu production/main pipeline của dự án.
> * Trạng thái an toàn: `ready_for_execution = FALSE` và `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Trạng Thái Tổng Thể (Proposed Status)

* Proposed status: `{summary['proposed_status']}`
* Các lý do chặn chưa sẵn sàng ghi thật:
  - Phase Final Write Plan Draft chưa có Human Approval.
  - Tồn tại `{summary['plan_blocked_count']}` bản ghi ở trạng thái `PLAN_BLOCKED` cần phân xử.
  - Tồn tại `{summary['high_risk_count']}` lỗi HIGH risk trong Risk Register cần làm rõ.

## 2. Thống Kê Final Write Plan

* Tổng candidates: `{summary['input_candidate_count']}`
* Đề xuất thêm mới (plan_insert_count): `{summary['plan_insert_count']}`
* Đề xuất cập nhật (plan_update_count): `{summary['plan_update_count']}`
* Đề xuất bỏ qua (plan_skip_count): `{summary['plan_skip_count']}`
* Đề xuất bị chặn (plan_blocked_count): `{summary['plan_blocked_count']}`

### Thống kê mức độ rủi ro (Risk Counts):
* LOW Risk: `{summary['low_risk_count']}`
* MEDIUM Risk: `{summary['medium_risk_count']}`
* HIGH Risk: `{summary['high_risk_count']}`

---

## 3. Điều Kiện Để Chuyển Sang Phase 2I (Ghi Thật Có Kiểm Soát)

Để chuyển đổi ready flags và thực thi lệnh ghi thật ở Phase tiếp theo, bắt buộc phải thỏa mãn:
- [ ] Số lượng `plan_blocked_count` phải bằng 0.
- [ ] Số lượng `high_risk_count` phải bằng 0 hoặc tất cả các lỗi HIGH được reviewer chấp thuận.
- [ ] Cấu trúc tệp tin Excel `final_write_plan_review.xlsx` được xác minh QA thành công.
- [ ] Có phương án Rollback và sao lưu cơ sở dữ liệu chính xác thực tế.
- [ ] Đạt chữ ký số và phê duyệt từ Human Approval Gate mới.

---

## 4. Các Tệp Tin Kết Xuất Cục Bộ
* Final Write Plan Items: `{output_dir / 'final_write_plan_items.json'}`
* Summary JSON: `{output_dir / 'final_write_plan_summary.json'}`
* Risk Register JSON: `{output_dir / 'final_write_plan_risk_register.json'}`
* Excel Review Workbook: `{output_dir / 'final_write_plan_review.xlsx'}`
"""


def run_final_write_plan_pipeline(
    cand_items_path: Path,
    cand_summary_path: Path,
    sim_records_path: Path,
    sim_summary_path: Path,
    match_results_path: Path,
    match_summary_path: Path,
    resol_items_path: Path,
    resol_summary_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    candidates = load_json(cand_items_path, default=[])
    simulations = load_json(sim_records_path, default=[])
    matches = load_json(match_results_path, default=[])
    resolutions = load_json(resol_items_path, default=[])

    # Thực hiện bóc tách và matching rules
    plan_items, risks, counts = process_plan(candidates, simulations, matches, resolutions)

    # Quyết định proposed_status
    if len(candidates) == 0:
        proposed_status = "FINAL_WRITE_PLAN_EMPTY"
    elif counts["plan_blocked_count"] > 0 or counts["high_risk_count"] > 0:
        proposed_status = "FINAL_WRITE_PLAN_BLOCKED"
    else:
        proposed_status = "FINAL_WRITE_PLAN_READY_FOR_HUMAN_REVIEW"

    summary = {
        "input_candidate_count": counts["input_candidate_count"],
        "input_simulation_count": counts["input_simulation_count"],
        "input_match_count": counts["input_match_count"],
        "input_resolution_count": counts["input_resolution_count"],
        "plan_insert_count": counts["plan_insert_count"],
        "plan_update_count": counts["plan_update_count"],
        "plan_skip_count": counts["plan_skip_count"],
        "plan_blocked_count": counts["plan_blocked_count"],
        "high_risk_count": counts["high_risk_count"],
        "medium_risk_count": counts["medium_risk_count"],
        "low_risk_count": counts["low_risk_count"],
        "ready_for_execution": False,
        "ready_for_write_to_main_pipeline": False,
        "proposed_status": proposed_status
    }

    manifest = {
        "plan_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "final_write_plan_draft",
        "ready_for_write_to_main_pipeline": False,
        "summary": summary,
        "plan_items": plan_items
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "final_write_plan_items.json", plan_items)
    write_json(output_dir / "final_write_plan_summary.json", manifest)
    write_json(output_dir / "final_write_plan_risk_register.json", risks)

    # Ghi CSV và XLSX
    write_csv_and_xlsx(output_dir, summary, plan_items, risks)

    # Ghi Báo Cáo Markdown
    md_content = build_final_write_plan_md(summary, output_dir)
    (output_dir / "final_write_plan.md").write_text(md_content, encoding="utf-8")

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export final write plan draft.")
    parser.add_argument("--candidates", type=Path, default=DEFAULT_CAND_ITEMS)
    parser.add_argument("--cand-summary", type=Path, default=DEFAULT_CAND_SUMMARY)
    parser.add_argument("--simulations", type=Path, default=DEFAULT_SIM_RECORDS)
    parser.add_argument("--sim-summary", type=Path, default=DEFAULT_SIM_SUMMARY)
    parser.add_argument("--matches", type=Path, default=DEFAULT_MATCH_RESULTS)
    parser.add_argument("--match-summary", type=Path, default=DEFAULT_MATCH_SUMMARY)
    parser.add_argument("--resolutions", type=Path, default=DEFAULT_RESOL_ITEMS)
    parser.add_argument("--resol-summary", type=Path, default=DEFAULT_RESOL_SUMMARY)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_final_write_plan_pipeline(
        cand_items_path=args.candidates,
        cand_summary_path=args.cand_summary,
        sim_records_path=args.simulations,
        sim_summary_path=args.sim_summary,
        match_results_path=args.matches,
        match_summary_path=args.match_summary,
        resol_items_path=args.resolutions,
        resol_summary_path=args.resol_summary,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
