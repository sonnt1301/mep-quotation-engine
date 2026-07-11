# -*- coding: utf-8 -*-
"""Controlled Write Simulation and Sandbox Commit Simulator.

This script reads the commit gate manifest and candidates, performs simulation
checks, generates mock material records, simulated commit logs, a simulated
rollback plan, and Excel sheets, ensuring ready_for_write_to_main_pipeline = False.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_GATE_MANIFEST = PROJECT_ROOT / "feasibility_outputs/profile_commit_gate/profile_commit_gate_manifest.json"
DEFAULT_CANDIDATE_ITEMS = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate/write_candidate_items.json"
DEFAULT_CANDIDATE_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_write_candidate/write_candidate_summary.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def calculate_file_hash(path: Path) -> str:
    if not path.exists():
        return ""
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()


def write_excel_review(
    output_dir: Path,
    summary: Dict[str, Any],
    records: List[Dict[str, Any]],
    log_entries: List[Dict[str, Any]],
    rollback_actions: List[Dict[str, Any]]
) -> None:
    xlsx_path = output_dir / "simulated_write_review.xlsx"
    wb = openpyxl.Workbook()

    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Xanh dương đậm
    fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt cho WOULD_SKIP/WOULD_UPDATE
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt cho BLOCKED
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("simulation_status", summary.get("simulation_status", "N/A")),
        ("total_candidates", summary.get("total_candidates", 0)),
        ("simulated_records_count", summary.get("simulated_records_count", 0)),
        ("would_insert_count", summary.get("would_insert_count", 0)),
        ("would_update_count", summary.get("would_update_count", 0)),
        ("would_skip_count", summary.get("would_skip_count", 0)),
        ("blocked_count", summary.get("blocked_count", 0)),
        ("ready_for_real_write", False),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_warning
        else:
            val_cell.alignment = align_left
            if "BLOCKED" in str(val):
                val_cell.fill = fill_danger

    # Sheet 2: Simulated Records
    ws_rec = wb.create_sheet(title="Simulated Records")
    ws_rec.views.sheetView[0].showGridLines = True
    ws_rec.freeze_panes = "A2"
    
    headers_rec = [
        "simulation_record_id", "write_candidate_id", "supplier_code", "material_code",
        "description", "unit", "quantity", "unit_price", "amount", "currency",
        "write_key", "proposed_action", "simulated_result", "ready_for_real_write"
    ]
    
    ws_rec.append(headers_rec)
    for col_idx in range(1, len(headers_rec) + 1):
        cell = ws_rec.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, rec in enumerate(records, start=2):
        row_data = [rec.get(k, "") for k in headers_rec]
        ws_rec.append(row_data)
        
        sim_res = rec.get("simulated_result", "")
        
        for col_idx, k in enumerate(headers_rec, start=1):
            cell = ws_rec.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["quantity", "unit_price", "amount"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["simulation_record_id", "write_candidate_id", "supplier_code", "currency", "simulated_result", "ready_for_real_write"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if sim_res == "BLOCKED":
                cell.fill = fill_danger
            elif sim_res in ["WOULD_SKIP", "WOULD_UPDATE"]:
                cell.fill = fill_warning

    ws_rec.auto_filter.ref = f"A1:{get_column_letter(len(headers_rec))}{len(records) + 1}"

    # Sheet 3: Commit Log
    ws_log = wb.create_sheet(title="Commit Log")
    ws_log.views.sheetView[0].showGridLines = True
    ws_log.freeze_panes = "A2"
    
    headers_log = ["log_id", "write_candidate_id", "write_key", "action", "result", "message"]
    ws_log.append(headers_log)
    for col_idx in range(1, len(headers_log) + 1):
        cell = ws_log.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, log in enumerate(log_entries, start=2):
        row_data = [log.get(k, "") for k in headers_log]
        ws_log.append(row_data)
        
        res = log.get("result", "")
        
        for col_idx, k in enumerate(headers_log, start=1):
            cell = ws_log.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["log_id", "write_candidate_id", "result"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if res == "BLOCKED":
                cell.fill = fill_danger
            elif res in ["WOULD_SKIP", "WOULD_UPDATE"]:
                cell.fill = fill_warning

    ws_log.auto_filter.ref = f"A1:{get_column_letter(len(headers_log))}{len(log_entries) + 1}"

    # Sheet 4: Rollback Plan
    ws_roll = wb.create_sheet(title="Rollback Plan")
    ws_roll.views.sheetView[0].showGridLines = True
    ws_roll.freeze_panes = "A2"
    
    headers_roll = ["action_id", "target_record_id", "undo_action", "description"]
    ws_roll.append(headers_roll)
    for col_idx in range(1, len(headers_roll) + 1):
        cell = ws_roll.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, act in enumerate(rollback_actions, start=2):
        row_data = [act.get(k, "") for k in headers_roll]
        ws_roll.append(row_data)
        
        for col_idx, k in enumerate(headers_roll, start=1):
            cell = ws_roll.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["action_id", "target_record_id", "undo_action"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_roll.auto_filter.ref = f"A1:{get_column_letter(len(headers_roll))}{len(rollback_actions) + 1}"

    # Auto-fit column widths
    for ws in [ws_sum, ws_rec, ws_log, ws_roll]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_simulation_report(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Controlled Write Simulation Report – Phase 2E

Báo cáo kết quả chạy sandbox write simulation đối với các write candidates đã được duyệt qua Commit Gate.

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN GIẢ LẬP**
> * Toàn bộ các hành động trong báo cáo này đều là **GIẢ LẬP (WOULD_INSERT/WOULD_SKIP)**.
> * Hệ thống **CHƯA** thực hiện ghi bất kỳ dữ liệu nào vào database hoặc main production pipeline của dự án.
> * Trạng thái an toàn: `ready_for_real_write = FALSE` và `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Trạng Thái Sandbox Simulation (Simulation Status)

* Trạng thái mô phỏng: `{summary['simulation_status']}`
* Lý do chưa thể ghi thật:
{chr(10).join(f"  - {r}" for r in summary['reasons_not_ready_for_real_write'])}

## 2. Thống Kê Sandbox Commit

* Tổng số candidates: `{summary['total_candidates']}`
* Số lượng giả lập ghi mới (would_insert_count): `{summary['would_insert_count']}`
* Số lượng giả lập cập nhật (would_update_count): `{summary['would_update_count']}`
* Số lượng giả lập bỏ qua (would_skip_count): `{summary['would_skip_count']}`
* Số lượng giả lập bị chặn (blocked_count): `{summary['blocked_count']}`

---

## 3. Danh Sách Tệp Tin Kết Xuất Cục Bộ
* Simulated Material Records: `{output_dir / 'simulated_material_records.json'}`
* Simulated Commit Log: `{output_dir / 'simulated_commit_log.json'}`
* Simulated Rollback Plan: `{output_dir / 'simulated_rollback_plan.json'}`
* Excel Review Workbook: `{output_dir / 'simulated_write_review.xlsx'}`
"""


def run_write_simulation(
    gate_manifest_path: Path,
    candidate_items_path: Path,
    candidate_summary_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    gate_manifest = load_json(gate_manifest_path, default={})
    candidate_items = load_json(candidate_items_path, default=[])
    candidate_summary = load_json(candidate_summary_path, default={})
    
    gate_status = gate_manifest.get("commit_gate_status", "PENDING_HUMAN_APPROVAL")
    blocking_reasons = gate_manifest.get("blocking_reasons", [])
    
    cand_sum_inner = candidate_summary.get("summary", {})
    dup_write_keys = cand_sum_inner.get("duplicate_write_key_count", 0)
    total_candidates = cand_sum_inner.get("candidate_items_count", 0)
    
    # Kiểm tra các điều kiện an toàn
    gate_approved = gate_status == "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY"
    no_blocking_gate = len(blocking_reasons) == 0
    no_dup_candidates = dup_write_keys == 0
    has_candidates = len(candidate_items) > 0
    
    reasons = []
    if not gate_approved:
        reasons.append(f"Commit Gate chưa được phê duyệt (Trạng thái hiện tại: {gate_status}).")
    if not no_blocking_gate:
        reasons.append(f"Tồn tại lý do chặn Commit Gate: {blocking_reasons}.")
    if not no_dup_candidates:
        reasons.append(f"Phát hiện {dup_write_keys} trùng lặp write_key trong candidates.")
    if not has_candidates:
        reasons.append("Danh sách write candidates rỗng.")
    reasons.append("Chưa chuyển đổi biến môi trường và chưa kích hoạt lệnh ghi thật của Phase 2F.")

    # Xác định simulation_status
    if not gate_approved:
        simulation_status = "BLOCKED_BY_COMMIT_GATE"
    elif has_candidates and no_dup_candidates and no_blocking_gate:
        simulation_status = "SIMULATION_READY_FOR_REVIEW"
    else:
        simulation_status = "BLOCKED_NEEDS_REVIEW"

    # Giả lập ghi nhận records
    records: List[Dict[str, Any]] = []
    log_entries: List[Dict[str, Any]] = []
    rollback_actions: List[Dict[str, Any]] = []
    
    would_insert_count = 0
    would_update_count = 0
    would_skip_count = 0
    blocked_count = 0
    
    if simulation_status == "SIMULATION_READY_FOR_REVIEW":
        for idx, cand in enumerate(candidate_items, start=1):
            rec_id = f"SIM_REC_{idx:04d}"
            log_id = f"LOG_ENTRY_{idx:04d}"
            roll_id = f"ROLLBACK_ACT_{idx:04d}"
            
            p_action = cand.get("proposed_action", "INSERT_CANDIDATE")
            write_key = cand.get("write_key", "")
            cand_id = cand.get("write_candidate_id", "")
            
            # Giả lập kết quả
            if p_action == "INSERT_CANDIDATE":
                sim_res = "WOULD_INSERT"
                would_insert_count += 1
                msg = "[WARNING: not_checked_against_master_database] Giả lập thêm mới thành công."
            elif p_action == "UPDATE_CANDIDATE":
                sim_res = "WOULD_UPDATE"
                would_update_count += 1
                msg = "[WARNING: not_checked_against_master_database] Giả lập cập nhật thành công."
            elif p_action == "SKIP_CANDIDATE":
                sim_res = "WOULD_SKIP"
                would_skip_count += 1
                msg = "Giả lập bỏ qua bản ghi do trùng write_key."
            else:
                sim_res = "BLOCKED"
                blocked_count += 1
                msg = "Bản ghi bị chặn."
                
            records.append({
                "simulation_record_id": rec_id,
                "write_candidate_id": cand_id,
                "supplier_code": cand.get("supplier_code"),
                "material_code": cand.get("material_code"),
                "description": cand.get("description"),
                "unit": cand.get("unit"),
                "quantity": cand.get("quantity"),
                "unit_price": cand.get("unit_price"),
                "amount": cand.get("amount"),
                "currency": cand.get("currency"),
                "write_key": write_key,
                "proposed_action": p_action,
                "simulated_result": sim_res,
                "source_provenance": cand.get("provenance"),
                "evidence_text": cand.get("evidence_text"),
                "created_from_phase": "Phase 2E Simulation",
                "ready_for_real_write": False
            })
            
            log_entries.append({
                "log_id": log_id,
                "write_candidate_id": cand_id,
                "write_key": write_key,
                "action": "simulation_insert" if p_action == "INSERT_CANDIDATE" else "simulation_skip",
                "result": sim_res,
                "message": msg
            })
            
            if sim_res in ["WOULD_INSERT", "WOULD_UPDATE"]:
                rollback_actions.append({
                    "action_id": roll_id,
                    "target_record_id": rec_id,
                    "undo_action": "DELETE",
                    "description": f"Xóa bản ghi giả lập {rec_id} đối với write_key {write_key}."
                })
    else:
        # Nếu bị chặn thì các count sẽ bằng 0 và records/log rỗng
        pass

    summary = {
        "simulation_status": simulation_status,
        "total_candidates": len(candidate_items) if simulation_status == "SIMULATION_READY_FOR_REVIEW" else 0,
        "simulated_records_count": len(records),
        "would_insert_count": would_insert_count,
        "would_update_count": would_update_count,
        "would_skip_count": would_skip_count,
        "blocked_count": blocked_count,
        "ready_for_real_write": False,
        "ready_for_write_to_main_pipeline": False,
        "reasons_not_ready_for_real_write": reasons
    }
    
    commit_log = {
        "simulation_id": f"SIM_SESSION_{utc_now().replace(':', '-').replace('.', '-')}",
        "generated_at": utc_now(),
        "total_candidates": len(candidate_items) if simulation_status == "SIMULATION_READY_FOR_REVIEW" else 0,
        "would_insert_count": would_insert_count,
        "would_update_count": would_update_count,
        "would_skip_count": would_skip_count,
        "blocked_count": blocked_count,
        "source_commit_gate_hash": calculate_file_hash(gate_manifest_path),
        "source_candidate_hash": calculate_file_hash(candidate_items_path),
        "ready_for_real_write": False,
        "log_entries": log_entries
    }
    
    rollback_plan = {
        "rollback_plan_id": f"ROLLBACK_PLAN_{utc_now().replace(':', '-').replace('.', '-')}",
        "generated_at": utc_now(),
        "simulation_id": commit_log["simulation_id"],
        "rollback_available": True,
        "rollback_scope": "sandbox_only",
        "rollback_actions": rollback_actions,
        "note": "Đây chỉ là phương án rollback giả lập được thiết kế trong sandbox. Chưa có bất kỳ ghi dữ liệu thực tế nào được thực thi nên không cần chạy rollback thật."
    }
    
    manifest = {
        "simulation_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "simulation",
        "source_commit_gate_manifest": str(gate_manifest_path),
        "ready_for_write_to_main_pipeline": False,
        "summary": summary,
        "simulated_records": records,
        "commit_log": commit_log,
        "rollback_plan": rollback_plan
    }
    
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "simulated_material_records.json", records)
    write_json(output_dir / "simulated_commit_log.json", commit_log)
    write_json(output_dir / "simulated_rollback_plan.json", rollback_plan)
    write_json(output_dir / "simulated_write_summary.json", manifest)
    
    # Kết xuất Excel Workbook
    write_excel_review(output_dir, summary, records, log_entries, rollback_actions)
    
    # Kết xuất Báo Cáo Markdown
    report_content = build_simulation_report(summary, output_dir)
    (output_dir / "simulated_write_report.md").write_text(report_content, encoding="utf-8")
    
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run write simulation sandbox.")
    parser.add_argument("--gate-manifest", type=Path, default=DEFAULT_GATE_MANIFEST)
    parser.add_argument("--candidate-items", type=Path, default=DEFAULT_CANDIDATE_ITEMS)
    parser.add_argument("--candidate-summary", type=Path, default=DEFAULT_CANDIDATE_SUMMARY)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_write_simulation(
        gate_manifest_path=args.gate_manifest,
        candidate_items_path=args.candidate_items,
        candidate_summary_path=args.candidate_summary,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
