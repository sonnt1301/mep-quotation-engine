# -*- coding: utf-8 -*-
"""Master Data Existing Record Matching Dry-run.

This script compares simulated material records with a master data fixture,
assigns match statuses,recommended actions, warnings, and confidence scores,
and exports Excel, CSV, JSON and Markdown summaries.
"""

from __future__ import annotations

import argparse
import json
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SIM_RECORDS = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_material_records.json"
DEFAULT_SIM_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_write_simulation/simulated_write_summary.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def generate_master_fixture(path: Path) -> List[Dict[str, Any]]:
    # Dữ liệu master index giả lập phục vụ đối chiếu khô
    fixture = [
        {
            "master_record_id": "MASTER_REC_0001",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "description": "Biến tần ACS310",
            "unit": "cái",
            "unit_price": 5000000,
            "currency": "VND",
            "write_key": "ABB|ACS310|P3|5000000|BIENTAN",
            "created_at": "2026-01-01T00:00:00Z"
        },
        {
            "master_record_id": "MASTER_REC_0002",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "Biến tần ACS355",
            "unit": "cái",
            "unit_price": 7500000,  # Khác giá so với preview (7,000,000) để test POSSIBLE_UPDATE/NEEDS_MASTER_REVIEW
            "currency": "VND",
            "write_key": "ABB|ACS355|P3|7500000|BIENTAN",
            "created_at": "2026-01-01T00:00:00Z"
        },
        {
            "master_record_id": "MASTER_REC_0003",
            "supplier_code": "LS",
            "material_code": "MC-9B",
            "description": "Contactor MC-9B",
            "unit": "cái",
            "unit_price": 250000,
            "currency": "VND",
            "write_key": "LS|MC-9B|P4|250000|CONTACTOR",
            "created_at": "2026-01-01T00:00:00Z"
        }
    ]
    write_json(path, fixture)
    return fixture


def perform_matching(
    sim_records: List[Dict[str, Any]],
    master_records: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    results: List[Dict[str, Any]] = []
    
    matched_count = 0
    no_match_count = 0
    possible_duplicate_count = 0
    needs_master_review_count = 0
    would_insert_count = 0
    would_update_count = 0
    would_skip_count = 0
    blocked_count = 0

    # Chuyển đổi master index thành dict để tăng tốc độ matching
    master_by_key = {m["write_key"]: m for m in master_records if m.get("write_key")}
    master_by_code = {}
    for m in master_records:
        key = (m["supplier_code"], m["material_code"])
        if key not in master_by_code:
            master_by_code[key] = []
        master_by_code[key].append(m)

    for idx, sim in enumerate(sim_records, start=1):
        res_id = f"MATCH_RES_{idx:04d}"
        sim_id = sim.get("simulation_record_id", "")
        cand_id = sim.get("write_candidate_id", "")
        supplier = sim.get("supplier_code", "")
        material = sim.get("material_code", "")
        desc = sim.get("description", "")
        unit = sim.get("unit", "")
        price = sim.get("unit_price", 0)
        amt = sim.get("amount", 0.0)
        curr = sim.get("currency", "VND")
        write_key = sim.get("write_key", "")
        sim_res = sim.get("simulated_result", "")
        
        match_status = "NO_MATCH"
        recommended_action = "WOULD_INSERT"
        matched_master_id = None
        confidence = 0.0
        reason = "Mã hàng mới hoàn toàn, không tìm thấy trong dữ liệu master."
        warnings = []

        if sim_res == "BLOCKED":
            match_status = "BLOCKED"
            recommended_action = "BLOCKED"
            reason = "Bản ghi simulation đã bị chặn."
            blocked_count += 1
        elif sim_res == "WOULD_SKIP":
            match_status = "POSSIBLE_DUPLICATE"
            recommended_action = "WOULD_SKIP"
            reason = "Giả lập simulation đã bỏ qua dòng này (trùng khóa)."
            would_skip_count += 1
            possible_duplicate_count += 1
        else:
            # 1. Khớp theo write_key
            if write_key and write_key in master_by_key:
                master = master_by_key[write_key]
                matched_master_id = master["master_record_id"]
                match_status = "EXACT_MATCH"
                confidence = 1.0
                matched_count += 1
                
                # So sánh chi tiết
                desc_match = master.get("description", "").strip() == desc.strip()
                unit_match = master.get("unit", "").strip() == unit.strip()
                price_match = master.get("unit_price", 0) == price
                
                if desc_match and unit_match and price_match:
                    recommended_action = "WOULD_SKIP"
                    reason = "Khóa ghi trùng khớp hoàn toàn với bản ghi master."
                    would_skip_count += 1
                else:
                    recommended_action = "WOULD_UPDATE"
                    reason = "Khóa ghi trùng khớp nhưng thông tin mô tả/đơn vị có sự thay đổi."
                    would_update_count += 1
                    
            # 2. Khớp theo supplier + material_code
            elif (supplier, material) in master_by_code:
                masters = master_by_code[(supplier, material)]
                master = masters[0]  # Lấy bản ghi đầu tiên
                matched_master_id = master["master_record_id"]
                matched_count += 1
                
                # Trùng mã nhưng khác giá hoặc mô tả
                if master.get("unit_price", 0) == price:
                    match_status = "POSSIBLE_DUPLICATE"
                    confidence = 0.8
                    recommended_action = "NEEDS_MASTER_REVIEW"
                    reason = "Trùng hãng và mã hàng, trùng giá tiền nhưng khác khóa ghi."
                    possible_duplicate_count += 1
                    needs_master_review_count += 1
                    warnings.append("[SAFETY WARNING] Trùng mã hàng trùng giá nhưng cấu trúc ghi khác biệt.")
                else:
                    match_status = "POSSIBLE_UPDATE"
                    confidence = 0.6
                    recommended_action = "NEEDS_MASTER_REVIEW"
                    reason = "Trùng hãng và mã hàng nhưng khác đơn giá hoặc mô tả. Cấm tự động gộp dòng."
                    needs_master_review_count += 1
                    warnings.append("[SAFETY WARNING] Trùng mã hàng nhưng khác đơn giá. Yêu cầu làm rõ trước khi merge.")
            
            # 3. Không khớp gì
            else:
                match_status = "NO_MATCH"
                confidence = 0.0
                recommended_action = "WOULD_INSERT"
                reason = "Mã hàng mới hoàn toàn, không tìm thấy trong dữ liệu master."
                no_match_count += 1
                would_insert_count += 1

        results.append({
            "match_result_id": res_id,
            "simulation_record_id": sim_id,
            "write_candidate_id": cand_id,
            "supplier_code": supplier,
            "material_code": material,
            "description": desc,
            "unit": unit,
            "unit_price": price,
            "amount": amt,
            "currency": curr,
            "write_key": write_key,
            "simulated_result": sim_res,
            "match_status": match_status,
            "recommended_action": recommended_action,
            "matched_master_record_id": matched_master_id,
            "match_confidence": confidence,
            "match_reason": reason,
            "warnings": warnings,
            "ready_for_real_write": False
        })

    counts = {
        "input_simulated_records_count": len(sim_records),
        "matched_count": matched_count,
        "no_match_count": no_match_count,
        "possible_duplicate_count": possible_duplicate_count,
        "needs_master_review_count": needs_master_review_count,
        "would_insert_count": would_insert_count,
        "would_update_count": would_update_count,
        "would_skip_count": would_skip_count,
        "blocked_count": blocked_count
    }
    return results, counts


def write_csv_and_xlsx(
    output_dir: Path,
    summary: Dict[str, Any],
    results: List[Dict[str, Any]],
    master_records: List[Dict[str, Any]]
) -> None:
    # 1. Ghi CSV master_match_review.csv
    csv_path = output_dir / "master_match_review.csv"
    headers_csv = [
        "match_result_id", "simulation_record_id", "write_candidate_id", "supplier_code",
        "material_code", "description", "unit", "unit_price", "write_key",
        "match_status", "recommended_action", "matched_master_record_id", "match_confidence", "match_reason"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_csv)
        writer.writeheader()
        for item in results:
            row = {k: item.get(k, "") for k in headers_csv}
            writer.writerow(row)

    # 2. Ghi Excel master_match_review.xlsx
    xlsx_path = output_dir / "master_match_review.xlsx"
    wb = openpyxl.Workbook()
    
    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Chỉ số đối chiếu (Metric)", "Giá trị (Value)"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("proposed_status", summary.get("proposed_status", "N/A")),
        ("input_simulated_records_count", summary.get("input_simulated_records_count", 0)),
        ("matched_count", summary.get("matched_count", 0)),
        ("no_match_count", summary.get("no_match_count", 0)),
        ("possible_duplicate_count", summary.get("possible_duplicate_count", 0)),
        ("needs_master_review_count", summary.get("needs_master_review_count", 0)),
        ("would_insert_count", summary.get("would_insert_count", 0)),
        ("would_update_count", summary.get("would_update_count", 0)),
        ("would_skip_count", summary.get("would_skip_count", 0)),
        ("blocked_count", summary.get("blocked_count", 0)),
        ("ready_for_real_write", False),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_warning
        else:
            val_cell.alignment = align_left
            if "BLOCKED" in str(val) or "NEEDS" in str(val):
                val_cell.fill = fill_danger

    # Sheet 2: Match Results
    ws_res = wb.create_sheet(title="Match Results")
    ws_res.views.sheetView[0].showGridLines = True
    ws_res.freeze_panes = "A2"
    
    ws_res.append(headers_csv)
    for col_idx in range(1, len(headers_csv) + 1):
        cell = ws_res.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, item in enumerate(results, start=2):
        row_data = [item.get(k, "") for k in headers_csv]
        ws_res.append(row_data)
        
        rec_act = item.get("recommended_action", "")
        
        for col_idx, k in enumerate(headers_csv, start=1):
            cell = ws_res.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["unit_price", "amount"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["match_confidence"]:
                cell.number_format = "0.0%"
                cell.alignment = align_right
            elif k in ["match_result_id", "simulation_record_id", "write_candidate_id", "supplier_code", "match_status", "recommended_action", "matched_master_record_id"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if rec_act == "BLOCKED" or rec_act == "NEEDS_MASTER_REVIEW":
                cell.fill = fill_danger
            elif rec_act in ["WOULD_SKIP", "WOULD_UPDATE"]:
                cell.fill = fill_warning

    ws_res.auto_filter.ref = f"A1:{get_column_letter(len(headers_csv))}{len(results) + 1}"

    # Sheet 3: Possible Duplicates
    ws_dup = wb.create_sheet(title="Possible Duplicates")
    ws_dup.views.sheetView[0].showGridLines = True
    ws_dup.freeze_panes = "A2"
    
    ws_dup.append(headers_csv)
    for col_idx in range(1, len(headers_csv) + 1):
        cell = ws_dup.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    dup_idx = 2
    for item in results:
        rec_act = item.get("recommended_action", "")
        m_status = item.get("match_status", "")
        if rec_act == "NEEDS_MASTER_REVIEW" or m_status in ["POSSIBLE_DUPLICATE", "POSSIBLE_UPDATE"]:
            row_data = [item.get(k, "") for k in headers_csv]
            ws_dup.append(row_data)
            
            for col_idx, k in enumerate(headers_csv, start=1):
                cell = ws_dup.cell(row=dup_idx, column=col_idx)
                cell.font = font_regular
                cell.border = thin_border
                
                if k in ["unit_price", "amount"]:
                    cell.number_format = "#,##0"
                    cell.alignment = align_right
                elif k in ["match_confidence"]:
                    cell.number_format = "0.0%"
                    cell.alignment = align_right
                elif k in ["match_result_id", "simulation_record_id", "write_candidate_id", "supplier_code", "match_status", "recommended_action", "matched_master_record_id"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                
                if rec_act == "NEEDS_MASTER_REVIEW":
                    cell.fill = fill_danger
                else:
                    cell.fill = fill_warning
            dup_idx += 1

    ws_dup.auto_filter.ref = f"A1:{get_column_letter(len(headers_csv))}{dup_idx}"

    # Sheet 4: Master Index Fixture
    ws_fix = wb.create_sheet(title="Master Index Fixture")
    ws_fix.views.sheetView[0].showGridLines = True
    ws_fix.freeze_panes = "A2"
    
    headers_fix = ["master_record_id", "supplier_code", "material_code", "description", "unit", "unit_price", "currency", "write_key", "created_at"]
    ws_fix.append(headers_fix)
    for col_idx in range(1, len(headers_fix) + 1):
        cell = ws_fix.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, fix in enumerate(master_records, start=2):
        row_data = [fix.get(k, "") for k in headers_fix]
        ws_fix.append(row_data)
        
        for col_idx, k in enumerate(headers_fix, start=1):
            cell = ws_fix.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["unit_price"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["master_record_id", "supplier_code", "currency", "created_at"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws_fix.auto_filter.ref = f"A1:{get_column_letter(len(headers_fix))}{len(master_records) + 1}"

    # Auto-fit column widths
    for ws in [ws_sum, ws_res, ws_dup, ws_fix]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_master_match_report(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Master Data Existing Record Matching Report – Phase 2F

Báo cáo đối chiếu khô (Dry-run Matching) giữa các simulated material records và kho dữ liệu Master Index Fixture.

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN DRY-RUN**
> * Tài liệu này thuộc **Phase đối chiếu khô (Dry-run Matching)**.
> * Hệ thống **CHƯA** thực hiện ghi thật hoặc thay đổi bất kỳ bản ghi nào trong cơ sở dữ liệu production/main pipeline.
> * Trạng thái an toàn: `ready_for_real_write = FALSE` và `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Trạng Thái Đề Xuất (Proposed Status)

* Proposed status: `{summary['proposed_status']}`
* Lý do chặn hoặc chưa sẵn sàng ghi thật:
  - Phase đối chiếu dry-run chưa được Human Approve.
  - Tồn tại `{summary['needs_master_review_count']}` bản ghi cần rà soát thủ công (`NEEDS_MASTER_REVIEW`).
  - `{summary['possible_duplicate_count']}` bản ghi trùng lặp mã hàng/giá cần xem xét.

## 2. Thống Kê Kết Quả Matching

* Tổng simulated records: `{summary['input_simulated_records_count']}`
* Khớp dữ liệu (matched_count): `{summary['matched_count']}`
* Không khớp (no_match_count): `{summary['no_match_count']}` (Sẽ đề xuất `WOULD_INSERT`)
* Trùng lặp nghi vấn (possible_duplicate_count): `{summary['possible_duplicate_count']}`
* Yêu cầu xem xét lại (needs_master_review_count): `{summary['needs_master_review_count']}`

### Phân rã hành động đề xuất (Recommended Actions):
* WOULD_INSERT: `{summary['would_insert_count']}`
* WOULD_UPDATE: `{summary['would_update_count']}`
* WOULD_SKIP: `{summary['would_skip_count']}`
* BLOCKED: `{summary['blocked_count']}`

---

## 3. Tiêu Chí Để Tiến Hành Phase 2G (Commit Gate tiếp theo)

Để chuyển sang Phase tiếp theo, bắt buộc phải thỏa mãn:
- [ ] `needs_master_review_count` phải bằng 0 (hoặc tất cả các dòng cảnh báo được duyệt thủ công).
- [ ] `possible_duplicate_count` phải bằng 0 (hoặc được human resolve trong file Excel).
- [ ] Không tồn tại bản ghi ở trạng thái `BLOCKED` do lỗi hệ thống.
- [ ] Reviewer xác nhận file Excel [master_match_review.xlsx](file:///{output_dir.resolve().as_posix()}/master_match_review.xlsx).
"""


def run_master_match_pipeline(
    sim_records_path: Path,
    sim_summary_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    sim_records = load_json(sim_records_path, default=[])
    sim_summary = load_json(sim_summary_path, default={})
    
    # 1. Đọc hoặc sinh master fixture
    output_dir.mkdir(parents=True, exist_ok=True)
    master_fixture_path = output_dir / "master_index_fixture.json"
    if not master_fixture_path.exists():
        master_records = generate_master_fixture(master_fixture_path)
    else:
        master_records = load_json(master_fixture_path, default=[])

    # 2. Thực hiện matching
    results, counts = perform_matching(sim_records, master_records)
    
    # Quyết định proposed_status
    if len(sim_records) == 0:
        proposed_status = "BLOCKED_NO_SIMULATION_RECORDS"
    elif counts["needs_master_review_count"] > 0 or counts["possible_duplicate_count"] > 0:
        proposed_status = "BLOCKED_NEEDS_MASTER_REVIEW"
    else:
        proposed_status = "MASTER_MATCH_READY_FOR_REVIEW"

    summary = {
        "input_simulated_records_count": counts["input_simulated_records_count"],
        "matched_count": counts["matched_count"],
        "no_match_count": counts["no_match_count"],
        "possible_duplicate_count": counts["possible_duplicate_count"],
        "needs_master_review_count": counts["needs_master_review_count"],
        "would_insert_count": counts["would_insert_count"],
        "would_update_count": counts["would_update_count"],
        "would_skip_count": counts["would_skip_count"],
        "blocked_count": counts["blocked_count"],
        "ready_for_real_write": False,
        "ready_for_write_to_main_pipeline": False,
        "proposed_status": proposed_status
    }

    manifest = {
        "match_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "master_match_dry_run",
        "source_write_simulation_summary": str(sim_summary_path),
        "ready_for_write_to_main_pipeline": False,
        "summary": summary,
        "match_results": results
    }

    # Ghi file
    write_json(output_dir / "master_match_results.json", results)
    write_json(output_dir / "master_match_summary.json", manifest)
    
    # Ghi CSV và XLSX
    write_csv_and_xlsx(output_dir, summary, results, master_records)
    
    # Ghi Markdown Report
    report_content = build_master_match_report(summary, output_dir)
    (output_dir / "master_match_report.md").write_text(report_content, encoding="utf-8")

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run master data dry-run matching pipeline.")
    parser.add_argument("--sim-records", type=Path, default=DEFAULT_SIM_RECORDS)
    parser.add_argument("--sim-summary", type=Path, default=DEFAULT_SIM_SUMMARY)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_master_match_pipeline(
        sim_records_path=args.sim_records,
        sim_summary_path=args.sim_summary,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
