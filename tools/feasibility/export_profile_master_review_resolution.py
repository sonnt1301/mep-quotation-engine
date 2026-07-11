# -*- coding: utf-8 -*-
"""Master Review Resolution Package Generator.

This script filters match dry-run results to extract items requiring human
resolution, generates templates for human input (CSV and Excel with dropdown validation),
and outputs resolution packages.
"""

from __future__ import annotations

import argparse
import json
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_MATCH_RESULTS = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_results.json"
DEFAULT_MATCH_SUMMARY = PROJECT_ROOT / "feasibility_outputs/profile_master_match_dry_run/master_match_summary.json"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "feasibility_outputs/profile_master_review_resolution"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        if default is not None:
            return default
        raise FileNotFoundError(f"JSON file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    tmp_path.replace(path)


def write_excel_resolution_template(
    output_dir: Path,
    summary: Dict[str, Any],
    resolution_items: List[Dict[str, Any]]
) -> None:
    xlsx_path = output_dir / "master_review_resolution_template.xlsx"
    wb = openpyxl.Workbook()

    # Styles
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Vàng nhạt cho PENDING
    fill_danger = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Đỏ cam nhạt cho BLOCKED/NEEDS_MORE_INFO
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.append(["Metric", "Value"])
    ws_sum.cell(row=1, column=1).font = font_header
    ws_sum.cell(row=1, column=1).fill = fill_header
    ws_sum.cell(row=1, column=2).font = font_header
    ws_sum.cell(row=1, column=2).fill = fill_header
    
    summary_rows = [
        ("proposed_status", summary.get("proposed_status", "N/A")),
        ("input_match_results_count", summary.get("input_match_results_count", 0)),
        ("resolution_required_count", summary.get("resolution_required_count", 0)),
        ("pending_count", summary.get("pending_count", 0)),
        ("confirm_insert_count", summary.get("confirm_insert_count", 0)),
        ("confirm_update_count", summary.get("confirm_update_count", 0)),
        ("confirm_skip_count", summary.get("confirm_skip_count", 0)),
        ("mark_duplicate_count", summary.get("mark_duplicate_count", 0)),
        ("needs_more_info_count", summary.get("needs_more_info_count", 0)),
        ("rejected_candidate_count", summary.get("rejected_candidate_count", 0)),
        ("ready_for_write_plan", False),
        ("ready_for_write_to_main_pipeline", False)
    ]
    
    for r_idx, (metric, val) in enumerate(summary_rows, start=2):
        ws_sum.append([metric, val])
        ws_sum.cell(row=r_idx, column=1).font = font_bold
        ws_sum.cell(row=r_idx, column=1).border = thin_border
        
        val_cell = ws_sum.cell(row=r_idx, column=2)
        val_cell.font = font_regular
        val_cell.border = thin_border
        
        if isinstance(val, int):
            val_cell.number_format = "#,##0"
            val_cell.alignment = align_right
        elif isinstance(val, bool):
            val_cell.alignment = align_center
            if val is False:
                val_cell.fill = fill_warning
        else:
            val_cell.alignment = align_left
            if "REQUIRED" in str(val) or "BLOCKED" in str(val):
                val_cell.fill = fill_danger

    # Sheet 2: Resolution Items
    ws_items = wb.create_sheet(title="Resolution Items")
    ws_items.views.sheetView[0].showGridLines = True
    ws_items.freeze_panes = "A2"
    
    headers_items = [
        "resolution_item_id", "match_result_id", "simulation_record_id", "write_candidate_id",
        "supplier_code", "material_code", "description", "unit", "unit_price", "amount",
        "currency", "write_key", "match_status", "recommended_action", "matched_master_record_id",
        "match_confidence", "match_reason", "human_resolution_decision", "human_resolution_note",
        "resolved_by", "resolved_at"
    ]
    
    ws_items.append(headers_items)
    for col_idx in range(1, len(headers_items) + 1):
        cell = ws_items.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Cấu hình Data Validation cho Dropdown Decision
    # Vị trí cột human_resolution_decision là 18 (cột R)
    dv = DataValidation(
        type="list",
        formula1='"PENDING,CONFIRM_INSERT,CONFIRM_UPDATE,CONFIRM_SKIP,MARK_DUPLICATE,NEEDS_MORE_INFO,REJECT_CANDIDATE"',
        allow_blank=True
    )
    ws_items.add_data_validation(dv)

    for r_idx, item in enumerate(resolution_items, start=2):
        row_data = [item.get(k, "") for k in headers_items]
        ws_items.append(row_data)
        
        decision = item.get("human_resolution_decision", "PENDING")
        
        # Link dropdown validation vào ô decision
        decision_cell = ws_items.cell(row=r_idx, column=18)
        dv.add(decision_cell)
        
        for col_idx, k in enumerate(headers_items, start=1):
            cell = ws_items.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            
            if k in ["unit_price", "amount"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif k in ["match_confidence"]:
                cell.number_format = "0.0%"
                cell.alignment = align_right
            elif k in ["resolution_item_id", "match_result_id", "simulation_record_id", "write_candidate_id", "supplier_code", "currency", "match_status", "recommended_action", "human_resolution_decision"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            if decision == "PENDING":
                cell.fill = fill_warning
            elif decision == "NEEDS_MORE_INFO" or decision == "REJECT_CANDIDATE":
                cell.fill = fill_danger

    ws_items.auto_filter.ref = f"A1:{get_column_letter(len(headers_items))}{len(resolution_items) + 1}"

    # Sheet 3: Decision Options
    ws_opt = wb.create_sheet(title="Decision Options")
    ws_opt.views.sheetView[0].showGridLines = True
    
    ws_opt.append(["Decision Option", "Giải thích ý nghĩa", "Khi nào sử dụng"])
    ws_opt.cell(row=1, column=1).font = font_header
    ws_opt.cell(row=1, column=1).fill = fill_header
    ws_opt.cell(row=1, column=2).font = font_header
    ws_opt.cell(row=1, column=2).fill = fill_header
    ws_opt.cell(row=1, column=3).font = font_header
    ws_opt.cell(row=1, column=3).fill = fill_header
    
    options = [
        ("PENDING", "Đang chờ phân xử.", "Trạng thái mặc định ban đầu."),
        ("CONFIRM_INSERT", "Xác nhận thêm mới vật tư.", "Mã hàng mới hoàn toàn hoặc được con người xác nhận tạo mới bản ghi."),
        ("CONFIRM_UPDATE", "Xác nhận cập nhật thông tin.", "Khớp write_key nhưng có thay đổi về mô tả/đơn vị cần ghi đè."),
        ("CONFIRM_SKIP", "Xác nhận bỏ qua dòng này.", "Khóa ghi đã tồn tại đầy đủ, không cần ghi đè."),
        ("MARK_DUPLICATE", "Đánh dấu là trùng lặp rác.", "Phát hiện dòng trùng lặp trong đợt bóc tách cần loại bỏ."),
        ("NEEDS_MORE_INFO", "Cần thêm thông tin từ dự án.", "Chưa đủ cơ sở dữ liệu để ra quyết định."),
        ("REJECT_CANDIDATE", "Từ chối ghi nhận dòng này.", "Quyết định từ chối do dữ liệu lỗi hoặc sai sót.")
    ]
    
    for r_idx, opt in enumerate(options, start=2):
        ws_opt.append(opt)
        for col_idx in range(1, 4):
            cell = ws_opt.cell(row=r_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_bold
                cell.alignment = align_center

    # Sheet 4: Warnings
    ws_warn = wb.create_sheet(title="Warnings")
    ws_warn.views.sheetView[0].showGridLines = True
    ws_warn.freeze_panes = "A2"
    
    ws_warn.append(["resolution_item_id", "supplier_code", "material_code", "match_status", "warnings"])
    for col_idx in range(1, 6):
        cell = ws_warn.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    w_idx = 2
    for item in resolution_items:
        warns = item.get("warnings", [])
        if warns:
            ws_warn.append([
                item.get("resolution_item_id"),
                item.get("supplier_code"),
                item.get("material_code"),
                item.get("match_status"),
                "; ".join(warns)
            ])
            for col_idx in range(1, 6):
                cell = ws_warn.cell(row=w_idx, column=col_idx)
                cell.font = font_regular
                cell.border = thin_border
                cell.fill = fill_danger
                if col_idx in [1, 2, 4]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
            w_idx += 1

    # Auto-fit column widths
    for ws in [ws_sum, ws_items, ws_opt, ws_warn]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    wb.save(xlsx_path)


def build_resolution_guide(summary: Dict[str, Any], output_dir: Path) -> str:
    return f"""# Hướng Dẫn Phân Xử Master Review – Phase 2G

Tài liệu này hướng dẫn cách thức mở và xử lý các vấn đề trùng lặp dữ liệu master của resolution package.

---

> [!WARNING]
> **CẢNH BÁO AN TOÀN DRY-RUN**
> * Resolution Package này **CHƯA** thực hiện ghi bất kỳ dữ liệu nào vào cơ sở dữ liệu thật.
> * Trạng thái an toàn: `ready_for_write_plan = FALSE` và `ready_for_write_to_main_pipeline = FALSE`.

---

## 1. Cách Thức Tiến Hành Phân Xử

1. **Bước 1**: Mở file Excel [master_review_resolution_template.xlsx](file:///{output_dir.resolve().as_posix()}/master_review_resolution_template.xlsx) hoặc CSV [master_review_resolution_template.csv](file:///{output_dir.resolve().as_posix()}/master_review_resolution_template.csv).
2. **Bước 2**: Chuyển sang sheet `Resolution Items`.
3. **Bước 3**: Tìm đến cột `human_resolution_decision` (Cột R trong Excel) và chọn một quyết định từ dropdown:
   - **CONFIRM_INSERT**: Duyệt tạo mới hoàn toàn bản ghi vật tư này trong master index.
   - **CONFIRM_UPDATE**: Chấp thuận cập nhật đè thông tin mới (ví dụ: thay đổi mô tả/đơn vị).
   - **CONFIRM_SKIP**: Bỏ qua dòng này vì dữ liệu master đã đầy đủ.
   - **MARK_DUPLICATE**: Đánh dấu dòng là trùng lặp lỗi cần loại bỏ.
   - **NEEDS_MORE_INFO**: Chưa đủ cơ sở phân xử, cần truy vấn thêm.
   - **REJECT_CANDIDATE**: Từ chối ghi nhận dòng này.
4. **Bước 4**: Ghi chú lý do vào cột `human_resolution_note`.

## 2. Tiêu Chí Để Sang Phase 2H (Write Plan)

- [ ] Không còn dòng nào ở trạng thái quyết định `PENDING`.
- [ ] Không còn dòng nào ở trạng thái quyết định `NEEDS_MORE_INFO`.
- [ ] Đã điền đầy đủ tên người duyệt `resolved_by` và ngày duyệt `resolved_at`.
- [ ] Chạy lại lệnh đóng gói để kiểm chứng trạng thái chuyển sang `MASTER_REVIEW_RESOLUTION_READY`.
"""


def run_resolution_pipeline(
    match_results_path: Path,
    match_summary_path: Path,
    output_dir: Path
) -> Dict[str, Any]:
    match_results = load_json(match_results_path, default=[])
    match_summary = load_json(match_summary_path, default={})
    
    resolution_items: List[Dict[str, Any]] = []
    
    # Lọc những dòng cần human resolve
    for idx, item in enumerate(match_results, start=1):
        m_status = item.get("match_status", "")
        rec_act = item.get("recommended_action", "")
        warns = item.get("warnings", [])
        
        needs_resolve = (
            m_status in ["POSSIBLE_UPDATE", "POSSIBLE_DUPLICATE"]
            or rec_act in ["NEEDS_MASTER_REVIEW", "BLOCKED"]
            or len(warns) > 0
        )
        
        if needs_resolve:
            res_id = f"RES_ITEM_{len(resolution_items) + 1:04d}"
            resolution_items.append({
                "resolution_item_id": res_id,
                "match_result_id": item.get("match_result_id"),
                "simulation_record_id": item.get("simulation_record_id"),
                "write_candidate_id": item.get("write_candidate_id"),
                "supplier_code": item.get("supplier_code"),
                "material_code": item.get("material_code"),
                "description": item.get("description"),
                "unit": item.get("unit"),
                "unit_price": item.get("unit_price"),
                "amount": item.get("amount"),
                "currency": item.get("currency"),
                "write_key": item.get("write_key"),
                "match_status": m_status,
                "recommended_action": rec_act,
                "matched_master_record_id": item.get("matched_master_record_id"),
                "match_confidence": item.get("match_confidence"),
                "match_reason": item.get("match_reason"),
                "warnings": warns,
                "human_resolution_decision": "PENDING",
                "human_resolution_note": "",
                "resolved_by": "",
                "resolved_at": ""
            })

    resolution_required_count = len(resolution_items)
    pending_count = resolution_required_count
    
    # Xác định proposed_status
    if resolution_required_count == 0:
        proposed_status = "NO_MASTER_REVIEW_REQUIRED"
    else:
        proposed_status = "MASTER_REVIEW_RESOLUTION_REQUIRED"

    summary = {
        "input_match_results_count": len(match_results),
        "resolution_required_count": resolution_required_count,
        "pending_count": pending_count,
        "confirm_insert_count": 0,
        "confirm_update_count": 0,
        "confirm_skip_count": 0,
        "mark_duplicate_count": 0,
        "needs_more_info_count": 0,
        "rejected_candidate_count": 0,
        "ready_for_write_plan": False,
        "ready_for_write_to_main_pipeline": False,
        "proposed_status": proposed_status
    }

    manifest = {
        "resolution_version": "1.0.0",
        "generated_at": utc_now(),
        "mode": "master_review_resolution",
        "source_master_match_summary_path": str(match_summary_path),
        "ready_for_write_to_main_pipeline": False,
        "summary": summary,
        "resolution_items": resolution_items
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "master_review_resolution_items.json", resolution_items)
    write_json(output_dir / "master_review_resolution_summary.json", manifest)

    # Ghi CSV
    csv_path = output_dir / "master_review_resolution_template.csv"
    headers_csv = [
        "resolution_item_id", "match_result_id", "supplier_code", "material_code",
        "description", "unit_price", "match_status", "recommended_action",
        "matched_master_record_id", "match_reason", "warnings", "human_resolution_decision",
        "human_resolution_note", "resolved_by", "resolved_at"
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers_csv)
        writer.writeheader()
        for item in resolution_items:
            row = {
                "resolution_item_id": item.get("resolution_item_id"),
                "match_result_id": item.get("match_result_id"),
                "supplier_code": item.get("supplier_code"),
                "material_code": item.get("material_code"),
                "description": item.get("description"),
                "unit_price": item.get("unit_price"),
                "match_status": item.get("match_status"),
                "recommended_action": item.get("recommended_action"),
                "matched_master_record_id": item.get("matched_master_record_id"),
                "match_reason": item.get("match_reason"),
                "warnings": "; ".join(item.get("warnings", [])),
                "human_resolution_decision": item.get("human_resolution_decision"),
                "human_resolution_note": item.get("human_resolution_note"),
                "resolved_by": item.get("resolved_by"),
                "resolved_at": item.get("resolved_at")
            }
            writer.writerow(row)

    # Ghi Excel
    write_excel_resolution_template(output_dir, summary, resolution_items)

    # Ghi Guide Markdown
    guide_content = build_resolution_guide(summary, output_dir)
    (output_dir / "master_review_resolution_guide.md").write_text(guide_content, encoding="utf-8")

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export master review resolution templates.")
    parser.add_argument("--match-results", type=Path, default=DEFAULT_MATCH_RESULTS)
    parser.add_argument("--match-summary", type=Path, default=DEFAULT_MATCH_SUMMARY)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = run_resolution_pipeline(
        match_results_path=args.match_results,
        match_summary_path=args.match_summary,
        output_dir=args.output_dir
    )
    print(json.dumps(manifest["summary"], ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
