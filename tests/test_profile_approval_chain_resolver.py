# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.run_profile_approval_chain_resolver import resolve_approval_chain

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_chain_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_chain_resolver_missing_files_no_crash():
    tmp_path = make_tmp_dir("missing_no_crash")
    non_existent = tmp_path / "does_not_exist.json"
    out_dir = tmp_path / "out"

    try:
        # Chạy resolver với các tệp không tồn tại
        manifest = resolve_approval_chain(
            gate_manifest_path=non_existent,
            sim_summary_path=non_existent,
            sim_records_path=non_existent,
            match_summary_path=non_existent,
            match_results_path=non_existent,
            resol_summary_path=non_existent,
            resol_items_path=non_existent,
            plan_summary_path=non_existent,
            plan_items_path=non_existent,
            plan_risks_path=non_existent,
            candidates_path=non_existent,
            output_dir=out_dir
        )

        # File thiếu không crash, top-level status là CHAIN_EMPTY vì candidates rỗng
        assert manifest["top_level_status"] == "CHAIN_EMPTY"
        assert manifest["ready_for_execution"] is False
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert len(manifest["blockers"]) > 0

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_chain_resolver_blocked_gate():
    tmp_path = make_tmp_dir("blocked_gate")
    gate_path = tmp_path / "profile_commit_gate_manifest.json"
    candidates_path = tmp_path / "write_candidate_items.json"
    out_dir = tmp_path / "out"

    # Gate pending
    gate_data = {
        "commit_gate_status": "PENDING_HUMAN_APPROVAL",
        "blocking_reasons": []
    }
    candidates_data = [
        {
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS355"
        }
    ]

    gate_path.write_text(json.dumps(gate_data, ensure_ascii=False), encoding="utf-8")
    candidates_path.write_text(json.dumps(candidates_data, ensure_ascii=False), encoding="utf-8")

    non_existent = tmp_path / "non_existent.json"

    try:
        manifest = resolve_approval_chain(
            gate_manifest_path=gate_path,
            sim_summary_path=non_existent,
            sim_records_path=non_existent,
            match_summary_path=non_existent,
            match_results_path=non_existent,
            resol_summary_path=non_existent,
            resol_items_path=non_existent,
            plan_summary_path=non_existent,
            plan_items_path=non_existent,
            plan_risks_path=non_existent,
            candidates_path=candidates_path,
            output_dir=out_dir
        )

        assert manifest["top_level_status"] == "CHAIN_BLOCKED"
        # Blocker đầu tiên phải là Commit Gate PENDING
        assert manifest["blockers"][0]["phase"] == "Commit Gate"
        cmd = manifest["blockers"][0]["required_command_if_any"]
        assert "export_profile_commit_gate.py" in cmd
        assert "--approve" in cmd
        assert "--approved-by" in cmd
        assert "--approval-note" in cmd
        assert "APPROVE_PROFILE_WRITE_CANDIDATES_FOR_NEXT_PHASE_ONLY" in cmd
        assert "APPROVED_FOR_DESIGN" not in cmd
        
        # Candidate bị block ở Commit Gate
        assert manifest["replay_items"][0]["current_blocking_phase"] == "Commit Gate"

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_chain_resolver_clean_chain_and_excel_qa():
    tmp_path = make_tmp_dir("clean_chain")
    gate_path = tmp_path / "profile_commit_gate_manifest.json"
    sim_summary_path = tmp_path / "simulated_write_summary.json"
    sim_records_path = tmp_path / "simulated_material_records.json"
    match_summary_path = tmp_path / "master_match_summary.json"
    match_results_path = tmp_path / "master_match_results.json"
    resol_summary_path = tmp_path / "master_review_resolution_summary.json"
    resol_items_path = tmp_path / "master_review_resolution_items.json"
    plan_summary_path = tmp_path / "final_write_plan_summary.json"
    plan_items_path = tmp_path / "final_write_plan_items.json"
    candidates_path = tmp_path / "write_candidate_items.json"
    out_dir = tmp_path / "out"

    # Ghi dữ liệu sạch hoàn chỉnh (APPROVED / READY)
    gate_data = {"commit_gate_status": "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY", "blocking_reasons": []}
    sim_sum_data = {"summary": {"simulation_status": "SIMULATION_READY_FOR_REVIEW"}}
    sim_rec_data = [{"write_candidate_id": "WRITE_CAND_0001", "simulation_record_id": "SIM_0001", "simulated_result": "WOULD_INSERT"}]
    match_sum_data = {"summary": {"proposed_status": "MASTER_MATCH_READY_FOR_REVIEW"}}
    match_res_data = [{"write_candidate_id": "WRITE_CAND_0001", "match_result_id": "MATCH_0001", "match_status": "NO_MATCH", "recommended_action": "WOULD_INSERT"}]
    resol_sum_data = {"summary": {"proposed_status": "NO_MASTER_REVIEW_REQUIRED"}}
    plan_sum_data = {"summary": {"proposed_status": "FINAL_WRITE_PLAN_READY_FOR_HUMAN_REVIEW"}}
    plan_its_data = [{"write_candidate_id": "WRITE_CAND_0001", "final_plan_item_id": "PLAN_0001", "final_planned_action": "PLAN_INSERT", "risk_level": "LOW"}]
    candidates_data = [{"write_candidate_id": "WRITE_CAND_0001", "supplier_code": "ABB", "material_code": "ACS355"}]

    gate_path.write_text(json.dumps(gate_data), encoding="utf-8")
    sim_summary_path.write_text(json.dumps(sim_sum_data), encoding="utf-8")
    sim_records_path.write_text(json.dumps(sim_rec_data), encoding="utf-8")
    match_summary_path.write_text(json.dumps(match_sum_data), encoding="utf-8")
    match_results_path.write_text(json.dumps(match_res_data), encoding="utf-8")
    resol_summary_path.write_text(json.dumps(resol_sum_data), encoding="utf-8")
    resol_items_path.write_text("[]", encoding="utf-8")
    plan_summary_path.write_text(json.dumps(plan_sum_data), encoding="utf-8")
    plan_items_path.write_text(json.dumps(plan_its_data), encoding="utf-8")
    candidates_path.write_text(json.dumps(candidates_data), encoding="utf-8")

    try:
        manifest = resolve_approval_chain(
            gate_manifest_path=gate_path,
            sim_summary_path=sim_summary_path,
            sim_records_path=sim_records_path,
            match_summary_path=match_summary_path,
            match_results_path=match_results_path,
            resol_summary_path=resol_summary_path,
            resol_items_path=resol_items_path,
            plan_summary_path=plan_summary_path,
            plan_items_path=plan_items_path,
            plan_risks_path=Path(""),
            candidates_path=candidates_path,
            output_dir=out_dir
        )

        # 1. Trạng thái clean chain
        assert manifest["top_level_status"] == "CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY"
        assert manifest["ready_for_execution"] is False
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert len(manifest["blockers"]) == 0

        # Candidate không bị block ở đâu
        assert manifest["replay_items"][0]["current_blocking_phase"] == "None"

        # 2. Tự động kiểm tra cấu trúc Excel QA
        wb = openpyxl.load_workbook(out_dir / "approval_chain_replay.xlsx")
        
        # Đúng 5 sheet
        expected_sheets = ["Chain Summary", "Blockers", "Candidate Replay", "Required Actions", "Source Files QA"]
        assert list(wb.sheetnames) == expected_sheets

        for sheet_name in expected_sheets:
            ws = wb[sheet_name]
            # Header không rỗng
            assert ws.cell(row=1, column=1).value is not None
            # Freeze panes và autofilter có trên các sheet dữ liệu
            if sheet_name != "Chain Summary":
                assert ws.freeze_panes == "A2"
                assert ws.auto_filter.ref is not None

        # Số dòng replay khớp
        ws_rep = wb["Candidate Replay"]
        # 1 header + 1 dòng data = 2
        assert ws_rep.max_row == 2
        
        wb.close()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_and_blocked_adapter_read():
    script_path = PROJECT_ROOT / "tools/feasibility/run_profile_approval_chain_resolver.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    assert "profile_bridge_items.json" not in content
    assert "blocked_items.json" not in content
