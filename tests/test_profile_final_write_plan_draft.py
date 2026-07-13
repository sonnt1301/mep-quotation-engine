# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.export_profile_final_write_plan_draft import run_final_write_plan_pipeline

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_final_plan_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_final_plan_empty_input():
    tmp_path = make_tmp_dir("empty")
    cand_items = tmp_path / "write_candidate_items.json"
    sim_records = tmp_path / "simulated_material_records.json"
    match_results = tmp_path / "master_match_results.json"
    resol_items = tmp_path / "master_review_resolution_items.json"
    out_dir = tmp_path / "out"

    for p in [cand_items, sim_records, match_results, resol_items]:
        p.write_text("[]", encoding="utf-8")

    try:
        manifest = run_final_write_plan_pipeline(
            cand_items_path=cand_items,
            cand_summary_path=Path(""),
            sim_records_path=sim_records,
            sim_summary_path=Path(""),
            match_results_path=match_results,
            match_summary_path=Path(""),
            resol_items_path=resol_items,
            resol_summary_path=Path(""),
            output_dir=out_dir
        )

        assert manifest["summary"]["proposed_status"] == "FINAL_WRITE_PLAN_EMPTY"
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_execution"] is False

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_final_plan_logic_and_workbook_qa():
    tmp_path = make_tmp_dir("logic_qa")
    cand_items = tmp_path / "write_candidate_items.json"
    sim_records = tmp_path / "simulated_material_records.json"
    match_results = tmp_path / "master_match_results.json"
    resol_items = tmp_path / "master_review_resolution_items.json"
    out_dir = tmp_path / "out"

    # Simulated candidate items
    cands_data = [
        {
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "MCCB",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 5000000,
            "amount": 5000000,
            "currency": "VND",
            "provenance": "PDF Page 3",
            "evidence_text": "ACS355"
        },
        {
            "write_candidate_id": "WRITE_CAND_0002",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "description": "Biến tần ACS310",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 7000000,
            "amount": 7000000,
            "currency": "VND",
            "provenance": "PDF Page 3",
            "evidence_text": "ACS310"
        },
        {
            "write_candidate_id": "WRITE_CAND_0003",
            "supplier_code": "LS",
            "material_code": "MC-9B",
            "description": "Contactor",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 250000,
            "amount": 250000,
            "currency": "VND",
            "provenance": "PDF Page 4",
            "evidence_text": "MC-9B"
        }
    ]

    # Matches data
    matches_data = [
        {
            "write_candidate_id": "WRITE_CAND_0001",
            "match_result_id": "MATCH_0001",
            "match_status": "POSSIBLE_UPDATE",
            "recommended_action": "NEEDS_MASTER_REVIEW",
            "matched_master_record_id": "MASTER_0001",
            "warnings": ["[SAFETY WARNING] Khác giá"]
        },
        {
            "write_candidate_id": "WRITE_CAND_0002",
            "match_result_id": "MATCH_0002",
            "match_status": "POSSIBLE_DUPLICATE",
            "recommended_action": "NEEDS_MASTER_REVIEW",
            "matched_master_record_id": "MASTER_0002",
            "warnings": []
        },
        {
            "write_candidate_id": "WRITE_CAND_0003",
            "match_result_id": "MATCH_0003",
            "match_status": "POSSIBLE_DUPLICATE",
            "recommended_action": "NEEDS_MASTER_REVIEW",
            "matched_master_record_id": "MASTER_0003",
            "warnings": []
        }
    ]

    # Resolution data:
    # 0001: PENDING -> PLAN_BLOCKED, risk HIGH
    # 0002: CONFIRM_INSERT -> PLAN_INSERT
    # 0003: REJECT_CANDIDATE -> PLAN_SKIP
    resol_data = [
        {
            "write_candidate_id": "WRITE_CAND_0001",
            "resolution_item_id": "RESOL_0001",
            "human_resolution_decision": "PENDING"
        },
        {
            "write_candidate_id": "WRITE_CAND_0002",
            "resolution_item_id": "RESOL_0002",
            "human_resolution_decision": "CONFIRM_INSERT"
        },
        {
            "write_candidate_id": "WRITE_CAND_0003",
            "resolution_item_id": "RESOL_0003",
            "human_resolution_decision": "REJECT_CANDIDATE"
        }
    ]

    cand_items.write_text(json.dumps(cands_data, ensure_ascii=False), encoding="utf-8")
    sim_records.write_text("[]", encoding="utf-8")
    match_results.write_text(json.dumps(matches_data, ensure_ascii=False), encoding="utf-8")
    resol_items.write_text(json.dumps(resol_data, ensure_ascii=False), encoding="utf-8")

    try:
        manifest = run_final_write_plan_pipeline(
            cand_items_path=cand_items,
            cand_summary_path=Path(""),
            sim_records_path=sim_records,
            sim_summary_path=Path(""),
            match_results_path=match_results,
            match_summary_path=Path(""),
            resol_items_path=resol_items,
            resol_summary_path=Path(""),
            output_dir=out_dir
        )

        assert manifest["summary"]["proposed_status"] == "FINAL_WRITE_PLAN_BLOCKED"
        assert manifest["summary"]["plan_insert_count"] == 1
        assert manifest["summary"]["plan_skip_count"] == 1
        assert manifest["summary"]["plan_blocked_count"] == 1
        assert manifest["summary"]["high_risk_count"] >= 1

        # Kiểm tra file output
        assert (out_dir / "final_write_plan_items.json").exists()
        assert (out_dir / "final_write_plan_summary.json").exists()
        assert (out_dir / "final_write_plan_risk_register.json").exists()
        assert (out_dir / "final_write_plan_review.csv").exists()
        assert (out_dir / "final_write_plan_review.xlsx").exists()
        assert (out_dir / "final_write_plan.md").exists()

        # Excel QA bằng openpyxl
        wb = openpyxl.load_workbook(out_dir / "final_write_plan_review.xlsx")
        
        # 1. Đúng sheet names
        expected_sheets = ["Summary", "Final Write Plan", "Blocked Items", "Risk Register", "Source Trace"]
        assert list(wb.sheetnames) == expected_sheets
        
        # 2. Header và Freeze panes
        for sheet_name in expected_sheets:
            ws = wb[sheet_name]
            assert ws.cell(row=1, column=1).value is not None
            # Summary không cần freeze pane, các sheet khác cần
            if sheet_name != "Summary":
                assert ws.freeze_panes == "A2"
                assert ws.auto_filter.ref is not None
                
        # 3. Row count khớp JSON summary
        ws_plan = wb["Final Write Plan"]
        # Có 3 dòng trong sheet (1 header, 1 PLAN_INSERT, 1 PLAN_SKIP)
        assert ws_plan.max_row == 3
        
        ws_blk = wb["Blocked Items"]
        # Có 2 dòng trong sheet (1 header, 1 PLAN_BLOCKED)
        assert ws_blk.max_row == 2

        wb.close()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_and_blocked_adapter_read():
    script_path = PROJECT_ROOT / "tools/feasibility/export_profile_final_write_plan_draft.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    assert "profile_bridge_items.json" not in content
    assert "blocked_items.json" not in content
