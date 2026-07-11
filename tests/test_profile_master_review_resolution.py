# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.export_profile_master_review_resolution import run_resolution_pipeline

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_master_resolution_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_resolution_no_master_review_required():
    tmp_path = make_tmp_dir("no_review")
    results_json = tmp_path / "master_match_results.json"
    summary_json = tmp_path / "master_match_summary.json"
    out_dir = tmp_path / "out"

    # Dữ liệu đối chiếu sạch hoàn toàn (WOULD_INSERT không warning)
    match_data = [
        {
            "match_result_id": "MATCH_RES_0001",
            "simulation_record_id": "SIM_REC_0001",
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "CHINT",
            "material_code": "NXM-63S",
            "description": "MCCB Chint",
            "unit": "cái",
            "unit_price": 800000,
            "amount": 800000.0,
            "currency": "VND",
            "write_key": "CHINT|NXM-63S|P3|800000|MCCB",
            "simulated_result": "WOULD_INSERT",
            "match_status": "NO_MATCH",
            "recommended_action": "WOULD_INSERT",
            "matched_master_record_id": None,
            "match_confidence": 0.0,
            "match_reason": "Mã mới",
            "warnings": [],
            "ready_for_real_write": False
        }
    ]

    results_json.write_text(json.dumps(match_data, ensure_ascii=False), encoding="utf-8")
    summary_json.write_text("{}", encoding="utf-8")

    try:
        manifest = run_resolution_pipeline(results_json, summary_json, out_dir)

        # Trạng thái NO_MASTER_REVIEW_REQUIRED vì không có dòng lỗi/trùng
        assert manifest["summary"]["proposed_status"] == "NO_MASTER_REVIEW_REQUIRED"
        assert manifest["summary"]["resolution_required_count"] == 0
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_write_plan"] is False

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_resolution_required_cases():
    tmp_path = make_tmp_dir("required_cases")
    results_json = tmp_path / "master_match_results.json"
    summary_json = tmp_path / "master_match_summary.json"
    out_dir = tmp_path / "out"

    # Dữ liệu chứa 2 items cần phân xử
    match_data = [
        {
            "match_result_id": "MATCH_RES_0001",
            "simulation_record_id": "SIM_REC_0001",
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "Biến tần ACS355",
            "unit": "cái",
            "unit_price": 7000000,
            "amount": 7000000.0,
            "currency": "VND",
            "write_key": "ABB|ACS355|P3|7000000|BIENTAN",
            "simulated_result": "WOULD_INSERT",
            "match_status": "POSSIBLE_UPDATE",
            "recommended_action": "NEEDS_MASTER_REVIEW",
            "matched_master_record_id": "MASTER_REC_0002",
            "match_confidence": 0.6,
            "match_reason": "Khác giá",
            "warnings": ["[SAFETY WARNING] Khác giá"],
            "ready_for_real_write": False
        },
        {
            "match_result_id": "MATCH_RES_0002",
            "simulation_record_id": "SIM_REC_0002",
            "write_candidate_id": "WRITE_CAND_0002",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "description": "Biến tần ACS310 trùng",
            "unit": "cái",
            "unit_price": 5000000,
            "amount": 5000000.0,
            "currency": "VND",
            "write_key": "ABB|ACS310|P3|5000000|BIENTAN",
            "simulated_result": "WOULD_INSERT",
            "match_status": "POSSIBLE_DUPLICATE",
            "recommended_action": "NEEDS_MASTER_REVIEW",
            "matched_master_record_id": "MASTER_REC_0001",
            "match_confidence": 0.8,
            "match_reason": "Trùng hoàn toàn",
            "warnings": [],
            "ready_for_real_write": False
        }
    ]

    results_json.write_text(json.dumps(match_data, ensure_ascii=False), encoding="utf-8")
    summary_json.write_text("{}", encoding="utf-8")

    try:
        manifest = run_resolution_pipeline(results_json, summary_json, out_dir)

        # Trạng thái MASTER_REVIEW_RESOLUTION_REQUIRED
        assert manifest["summary"]["proposed_status"] == "MASTER_REVIEW_RESOLUTION_REQUIRED"
        assert manifest["summary"]["resolution_required_count"] == 2
        assert manifest["summary"]["pending_count"] == 2
        
        # Quyết định mặc định PENDING
        assert manifest["resolution_items"][0]["human_resolution_decision"] == "PENDING"
        assert manifest["resolution_items"][0]["resolved_by"] == ""
        
        # Đảm bảo các flags an toàn luôn = False
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_write_plan"] is False

        # Kiểm tra Excel đủ 4 sheet
        wb = openpyxl.load_workbook(out_dir / "master_review_resolution_template.xlsx")
        assert "Summary" in wb.sheetnames
        assert "Resolution Items" in wb.sheetnames
        assert "Decision Options" in wb.sheetnames
        assert "Warnings" in wb.sheetnames
        
        # Xác minh dropdown validation trên Excel
        ws_items = wb["Resolution Items"]
        assert len(ws_items.data_validations.dataValidation) > 0
        wb.close()

        # Xác minh file CSV, Guide được sinh ra
        assert (out_dir / "master_review_resolution_template.csv").exists()
        assert (out_dir / "master_review_resolution_guide.md").exists()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_and_blocked_adapter_read():
    # Đảm bảo không đọc các tệp dữ liệu thô
    script_path = PROJECT_ROOT / "tools/feasibility/export_profile_master_review_resolution.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    assert "profile_bridge_items.json" not in content
    assert "blocked_items.json" not in content
