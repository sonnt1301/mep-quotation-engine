# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.run_profile_master_match_dry_run import run_master_match_pipeline

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_master_match_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_master_match_empty_input():
    tmp_path = make_tmp_dir("empty")
    sim_records_json = tmp_path / "simulated_material_records.json"
    sim_summary_json = tmp_path / "simulated_write_summary.json"
    out_dir = tmp_path / "out"

    sim_records_json.write_text("[]", encoding="utf-8")
    sim_summary_json.write_text("{}", encoding="utf-8")

    try:
        manifest = run_master_match_pipeline(sim_records_json, sim_summary_json, out_dir)

        # Input rỗng thì proposed_status = BLOCKED_NO_SIMULATION_RECORDS
        assert manifest["summary"]["proposed_status"] == "BLOCKED_NO_SIMULATION_RECORDS"
        assert manifest["summary"]["input_simulated_records_count"] == 0
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_real_write"] is False

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_master_match_logic_cases():
    tmp_path = make_tmp_dir("logic")
    sim_records_json = tmp_path / "simulated_material_records.json"
    sim_summary_json = tmp_path / "simulated_write_summary.json"
    out_dir = tmp_path / "out"

    # Tạo simulated records kiểm thử 3 cases: exact match, code match khác giá, no match
    sim_data = [
        {
            "simulation_record_id": "SIM_REC_0001",
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "description": "Biến tần ACS310",
            "unit": "cái",
            "quantity": 1.0,
            "unit_price": 5000000,
            "amount": 5000000.0,
            "currency": "VND",
            "write_key": "ABB|ACS310|P3|5000000|BIENTAN",
            "simulated_result": "WOULD_INSERT"
        },
        {
            "simulation_record_id": "SIM_REC_0002",
            "write_candidate_id": "WRITE_CAND_0002",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "Biến tần ACS355",
            "unit": "cái",
            "quantity": 1.0,
            "unit_price": 7000000,  # Master đang lưu giá 7.500.000 để test trùng mã khác giá
            "amount": 7000000.0,
            "currency": "VND",
            "write_key": "ABB|ACS355|P3|7000000|BIENTAN",
            "simulated_result": "WOULD_INSERT"
        },
        {
            "simulation_record_id": "SIM_REC_0003",
            "write_candidate_id": "WRITE_CAND_0003",
            "supplier_code": "CHINT",
            "material_code": "NXM-63S",
            "description": "MCCB Chint mới",
            "unit": "cái",
            "quantity": 1.0,
            "unit_price": 800000,
            "amount": 800000.0,
            "currency": "VND",
            "write_key": "CHINT|NXM-63S|P3|800000|MCCB",
            "simulated_result": "WOULD_INSERT"
        }
    ]

    sim_records_json.write_text(json.dumps(sim_data, ensure_ascii=False), encoding="utf-8")
    sim_summary_json.write_text("{}", encoding="utf-8")

    try:
        manifest = run_master_match_pipeline(sim_records_json, sim_summary_json, out_dir)

        # Trạng thái proposed_status bị block do cần master review
        assert manifest["summary"]["proposed_status"] == "BLOCKED_NEEDS_MASTER_REVIEW"
        assert manifest["summary"]["input_simulated_records_count"] == 3
        
        # Test case 1: EXACT MATCH -> WOULD_SKIP (do trùng hoàn toàn)
        res1 = manifest["match_results"][0]
        assert res1["match_status"] == "EXACT_MATCH"
        assert res1["recommended_action"] == "WOULD_SKIP"
        
        # Test case 2: Trùng mã nhưng khác giá -> NEEDS_MASTER_REVIEW
        res2 = manifest["match_results"][1]
        assert res2["match_status"] == "POSSIBLE_UPDATE"
        assert res2["recommended_action"] == "NEEDS_MASTER_REVIEW"
        
        # Test case 3: NO_MATCH -> WOULD_INSERT
        res3 = manifest["match_results"][2]
        assert res3["match_status"] == "NO_MATCH"
        assert res3["recommended_action"] == "WOULD_INSERT"

        # Đảm bảo flags an toàn = False
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_real_write"] is False

        # Kiểm tra Excel đủ 4 sheet
        wb = openpyxl.load_workbook(out_dir / "master_match_review.xlsx")
        assert "Summary" in wb.sheetnames
        assert "Match Results" in wb.sheetnames
        assert "Possible Duplicates" in wb.sheetnames
        assert "Master Index Fixture" in wb.sheetnames
        wb.close()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_and_blocked_adapter_read():
    # Kiểm tra xem run_profile_master_match_dry_run.py có đọc file items json gốc hoặc blocked items của adapter không
    script_path = PROJECT_ROOT / "tools/feasibility/run_profile_master_match_dry_run.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    assert "profile_bridge_items.json" not in content
    assert "blocked_items.json" not in content
