# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.export_profile_final_business_signoff import run_business_signoff_pipeline

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_signoff_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_signoff_blocked_by_approval_chain():
    tmp_path = make_tmp_dir("chain_blocked")
    plan_items_path = tmp_path / "final_write_plan_items.json"
    chain_status_path = tmp_path / "approval_chain_status.json"
    out_dir = tmp_path / "out"

    # Approval chain bị block (Proposed status = CHAIN_BLOCKED)
    chain_data = {
        "top_level_status": "CHAIN_BLOCKED",
        "chain_summary": {
            "total_blockers_count": 2,
            "commit_gate_status": "PENDING_HUMAN_APPROVAL",
            "simulation_status": "BLOCKED_BY_COMMIT_GATE",
            "master_match_status": "BLOCKED_NO_SIMULATION_RECORDS",
            "master_resolution_status": "NO_MASTER_REVIEW_REQUIRED",
            "final_write_plan_status": "FINAL_WRITE_PLAN_BLOCKED",
            "total_candidates_count": 3
        }
    }
    plan_data = [
        {
            "final_plan_item_id": "PLAN_0001",
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "MCCB",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 5000000,
            "amount": 5000000.0,
            "final_planned_action": "PLAN_INSERT",
            "risk_level": "LOW",
            "provenance": "PDF Page 3",
            "evidence_text": "ACS355",
            "matched_master_record_id": None
        }
    ]

    plan_items_path.write_text(json.dumps(plan_data), encoding="utf-8")
    chain_status_path.write_text(json.dumps(chain_data), encoding="utf-8")

    try:
        manifest = run_business_signoff_pipeline(
            plan_items_path=plan_items_path,
            plan_summary_path=Path(""),
            plan_risks_path=Path(""),
            chain_status_path=chain_status_path,
            chain_blockers_path=Path(""),
            sim_records_path=Path(""),
            match_results_path=Path(""),
            output_dir=out_dir
        )

        assert manifest["summary"]["proposed_status"] == "BLOCKED_BY_APPROVAL_CHAIN"
        assert manifest["ready_for_execution"] is False
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["signoff_items"][0]["human_decision"] == "PENDING"

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_signoff_clean_chain_and_excel_qa():
    tmp_path = make_tmp_dir("clean_signoff")
    plan_items_path = tmp_path / "final_write_plan_items.json"
    chain_status_path = tmp_path / "approval_chain_status.json"
    out_dir = tmp_path / "out"

    # Approval chain sạch (Proposed status = CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY)
    chain_data = {
        "top_level_status": "CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY",
        "chain_summary": {
            "total_blockers_count": 0,
            "commit_gate_status": "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY",
            "simulation_status": "SIMULATION_READY_FOR_REVIEW",
            "master_match_status": "MASTER_MATCH_READY_FOR_REVIEW",
            "master_resolution_status": "NO_MASTER_REVIEW_REQUIRED",
            "final_write_plan_status": "FINAL_WRITE_PLAN_READY_FOR_HUMAN_REVIEW",
            "total_candidates_count": 3
        }
    }
    plan_data = [
        {
            "final_plan_item_id": "PLAN_0001",
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "MCCB",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 5000000,
            "amount": 5000000.0,
            "final_planned_action": "PLAN_INSERT",
            "risk_level": "LOW",
            "provenance": "PDF Page 3",
            "evidence_text": "ACS355",
            "matched_master_record_id": None
        },
        {
            "final_plan_item_id": "PLAN_0002",
            "write_candidate_id": "WRITE_CAND_0002",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "description": "Biến tần ACS310",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 7000000,
            "amount": 7000000.0,
            "final_planned_action": "PLAN_INSERT",
            "risk_level": "LOW",
            "provenance": "PDF Page 3",
            "evidence_text": "ACS310",
            "matched_master_record_id": None
        },
        {
            "final_plan_item_id": "PLAN_0003",
            "write_candidate_id": "WRITE_CAND_0003",
            "supplier_code": "LS",
            "material_code": "MC-9B",
            "description": "Contactor",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 250000,
            "amount": 250000.0,
            "final_planned_action": "PLAN_INSERT",
            "risk_level": "LOW",
            "provenance": "PDF Page 4",
            "evidence_text": "MC-9B",
            "matched_master_record_id": None
        }
    ]

    plan_items_path.write_text(json.dumps(plan_data), encoding="utf-8")
    chain_status_path.write_text(json.dumps(chain_data), encoding="utf-8")

    try:
        manifest = run_business_signoff_pipeline(
            plan_items_path=plan_items_path,
            plan_summary_path=Path(""),
            plan_risks_path=Path(""),
            chain_status_path=chain_status_path,
            chain_blockers_path=Path(""),
            sim_records_path=Path(""),
            match_results_path=Path(""),
            output_dir=out_dir
        )

        assert manifest["summary"]["proposed_status"] == "FINAL_BUSINESS_SIGNOFF_PENDING"
        assert manifest["summary"]["total_signoff_items"] == 3
        assert manifest["ready_for_execution"] is False
        assert manifest["ready_for_write_to_main_pipeline"] is False

        # Kiểm tra Excel QA
        wb = openpyxl.load_workbook(out_dir / "final_business_signoff_template.xlsx")
        
        expected_sheets = ["Summary", "Sign-off Items", "Decision Options", "Source Evidence", "Approval Preconditions"]
        assert list(wb.sheetnames) == expected_sheets

        for sheet_name in expected_sheets:
            ws = wb[sheet_name]
            assert ws.cell(row=1, column=1).value is not None
            if sheet_name != "Summary":
                assert ws.freeze_panes == "A2"
                assert ws.auto_filter.ref is not None

        # Data Validation trong Sign-off Items tồn tại
        ws_items = wb["Sign-off Items"]
        assert len(ws_items.data_validations.dataValidation) > 0
        # 1 header + 3 rows = 4 rows
        assert ws_items.max_row == 4
        
        wb.close()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_signoff_stale_source_mismatch():
    tmp_path = make_tmp_dir("stale_mismatch")
    plan_items_path = tmp_path / "final_write_plan_items.json"
    chain_status_path = tmp_path / "approval_chain_status.json"
    out_dir = tmp_path / "out"

    chain_data = {
        "top_level_status": "CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY",
        "chain_summary": {
            "total_blockers_count": 0,
            "commit_gate_status": "APPROVED",
            "simulation_status": "SUCCESS",
            "master_match_status": "SUCCESS",
            "master_resolution_status": "READY",
            "final_write_plan_status": "READY",
            "total_candidates_count": 1
        }
    }
    plan_data = [
        {
            "final_plan_item_id": "PLAN_0001",
            "write_candidate_id": "WRITE_CAND_0001",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "final_planned_action": "PLAN_INSERT"
        }
    ]

    plan_items_path.write_text(json.dumps(plan_data), encoding="utf-8")
    chain_status_path.write_text(json.dumps(chain_data), encoding="utf-8")

    try:
        # Chạy lần 1 để sinh manifest cũ
        run_business_signoff_pipeline(
            plan_items_path=plan_items_path,
            plan_summary_path=Path(""),
            plan_risks_path=Path(""),
            chain_status_path=chain_status_path,
            chain_blockers_path=Path(""),
            sim_records_path=Path(""),
            match_results_path=Path(""),
            output_dir=out_dir
        )

        # Thay đổi dữ liệu plan để tạo mismatch hash (stale)
        plan_data.append({
            "final_plan_item_id": "PLAN_0002",
            "write_candidate_id": "WRITE_CAND_0002",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "final_planned_action": "PLAN_INSERT"
        })
        plan_items_path.write_text(json.dumps(plan_data), encoding="utf-8")

        # Chạy lần 2 để kiểm chứng
        manifest = run_business_signoff_pipeline(
            plan_items_path=plan_items_path,
            plan_summary_path=Path(""),
            plan_risks_path=Path(""),
            chain_status_path=chain_status_path,
            chain_blockers_path=Path(""),
            sim_records_path=Path(""),
            match_results_path=Path(""),
            output_dir=out_dir
        )

        assert manifest["summary"]["proposed_status"] == "STALE_SIGNOFF_SOURCE"

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_signoff_read_only_regression(monkeypatch):
    import subprocess
    import os
    import hashlib

    # Mock chặn mọi subprocess calls
    def mock_forbidden(*args, **kwargs):
        raise RuntimeError("Forbidden: subprocess/os call detected!")
    monkeypatch.setattr(subprocess, "run", mock_forbidden)
    monkeypatch.setattr(subprocess, "Popen", mock_forbidden)
    monkeypatch.setattr(os, "system", mock_forbidden)

    tmp_path = make_tmp_dir("readonly_regression")
    
    gate_manifest = tmp_path / "profile_commit_gate_manifest.json"
    sim_summary = tmp_path / "simulated_write_summary.json"
    sim_records = tmp_path / "simulated_material_records.json"
    match_summary = tmp_path / "master_match_summary.json"
    plan_summary = tmp_path / "final_write_plan_summary.json"
    plan_items = tmp_path / "final_write_plan_items.json"
    chain_status = tmp_path / "approval_chain_status.json"
    out_dir = tmp_path / "out"

    # Ghi dữ liệu fixture approved / 3 plan insert sạch
    gate_data = {"commit_gate_status": "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY", "blocking_reasons": []}
    sim_sum_data = {"summary": {"simulation_status": "SIMULATION_READY_FOR_REVIEW"}}
    sim_rec_data = [{"write_candidate_id": f"WRITE_CAND_000{i}", "simulation_record_id": f"SIM_000{i}"} for i in range(1, 4)]
    match_sum_data = {"summary": {"proposed_status": "MASTER_MATCH_READY_FOR_REVIEW"}}
    plan_sum_data = {"summary": {"proposed_status": "FINAL_WRITE_PLAN_READY_FOR_HUMAN_REVIEW"}}
    plan_its_data = [
        {
            "final_plan_item_id": f"PLAN_000{i}",
            "write_candidate_id": f"WRITE_CAND_000{i}",
            "supplier_code": "ABB",
            "material_code": f"ACS35{i}",
            "final_planned_action": "PLAN_INSERT",
            "risk_level": "LOW"
        } for i in range(1, 4)
    ]
    chain_data = {
        "top_level_status": "CHAIN_READY_FOR_HUMAN_APPROVAL_REPLAY",
        "chain_summary": {
            "total_blockers_count": 0,
            "commit_gate_status": "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY",
            "simulation_status": "SIMULATION_READY_FOR_REVIEW",
            "master_match_status": "MASTER_MATCH_READY_FOR_REVIEW",
            "master_resolution_status": "NO_MASTER_REVIEW_REQUIRED",
            "final_write_plan_status": "FINAL_WRITE_PLAN_READY_FOR_HUMAN_REVIEW",
            "total_candidates_count": 3
        }
    }

    gate_manifest.write_text(json.dumps(gate_data), encoding="utf-8")
    sim_summary.write_text(json.dumps(sim_sum_data), encoding="utf-8")
    sim_records.write_text(json.dumps(sim_rec_data), encoding="utf-8")
    match_summary.write_text(json.dumps(match_sum_data), encoding="utf-8")
    plan_summary.write_text(json.dumps(plan_sum_data), encoding="utf-8")
    plan_items.write_text(json.dumps(plan_its_data), encoding="utf-8")
    chain_status.write_text(json.dumps(chain_data), encoding="utf-8")

    # Tính hash trước khi chạy
    upstream_paths = [gate_manifest, sim_summary, sim_records, match_summary, plan_summary, plan_items, chain_status]
    def get_hash(p):
        return hashlib.sha256(p.read_bytes()).hexdigest()
    pre_hashes = {str(p): get_hash(p) for p in upstream_paths}

    try:
        manifest = run_business_signoff_pipeline(
            plan_items_path=plan_items,
            plan_summary_path=plan_summary,
            plan_risks_path=Path(""),
            chain_status_path=chain_status,
            chain_blockers_path=Path(""),
            sim_records_path=sim_records,
            match_results_path=Path(""),
            output_dir=out_dir
        )

        # 1. Trùng khớp hash trước/sau không đổi (Read-only)
        post_hashes = {str(p): get_hash(p) for p in upstream_paths}
        for p in upstream_paths:
            assert pre_hashes[str(p)] == post_hashes[str(p)], f"File has been modified: {p}"

        # 2. Tạo đúng 3 signoff items pending
        assert manifest["summary"]["total_signoff_items"] == 3
        assert manifest["summary"]["pending_count"] == 3
        assert manifest["summary"]["proposed_status"] == "FINAL_BUSINESS_SIGNOFF_PENDING"

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_and_blocked_adapter_read():
    script_path = PROJECT_ROOT / "tools/feasibility/export_profile_final_business_signoff.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    assert "profile_bridge_items.json" not in content
    assert "blocked_items.json" not in content
