import os
import json
import shutil
import tempfile
import hashlib
from pathlib import Path
from datetime import datetime, timezone
import pytest
import openpyxl

from mep_quotation.spec.models import (
    NormalizedQuotationModel,
    NormalizedItemModel,
    ExportSummaryModel,
    ParserWarningModel,
    ReviewDecisionsFileModel,
    ReviewDecisionModel
)
from mep_quotation.package.builder import create_empty_package
from mep_quotation.package.writer import write_json_file
from mep_quotation.package.integrity import validate_package_integrity
from mep_quotation.excel_export.export_service import export_excel

def _get_sha256(file_path: Path) -> str:
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            sha.update(chunk)
    return sha.hexdigest()

@pytest.fixture
def temp_project_dir():
    temp_dir = tempfile.mkdtemp()
    yield Path(temp_dir)
    shutil.rmtree(temp_dir)

@pytest.fixture
def sample_package(temp_project_dir):
    """Tạo một package mock hoàn thiện đầy đủ từ Phase 1 đến Phase 11."""
    # 1. Tạo package.json trống và thư mục package bằng create_empty_package
    from mep_quotation.package.builder import create_empty_package
    package_dir = create_empty_package(temp_project_dir, "AUT", "2026-06-20", seq=1)
    
    # original.pdf giả lập
    pdf_path = package_dir / "source" / "original.pdf"
    import fitz
    doc = fitz.open()
    doc.new_page()
    doc.save(str(pdf_path))
    doc.close()

    # metadata.json
    meta_data = {
        "schema_version": "1.0",
        "file_name": "original.pdf",
        "file_size": pdf_path.stat().st_size,
        "sha256": _get_sha256(pdf_path),
        "page_count": 1,
        "pdf_version": "1.4",
        "encrypted": False,
        "imported_at": "2026-06-20T00:00:00Z",
        "warnings": []
    }
    with open(package_dir / "source" / "metadata.json", "w", encoding="utf-8") as f:
        json.dump(meta_data, f)

    # corrections.json
    corr_dir = package_dir / "corrections"
    corr_dir.mkdir(parents=True, exist_ok=True)
    with open(corr_dir / "corrections.json", "w", encoding="utf-8") as f:
        json.dump({"schema_version": "1.0", "quotation_id": "AUT_20260620_001", "corrections": []}, f)

    # quotation.md
    md_content = "mock\nmock\nmock\nmock\n"
    md_path = package_dir / "text" / "quotation.md"
    md_path.parent.mkdir(parents=True, exist_ok=True)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    # raw_text.json
    raw_text_path = package_dir / "source" / "raw_text.json"
    with open(raw_text_path, "w", encoding="utf-8") as f:
        json.dump({"schema_version": "1.0", "quotation_id": "AUT_20260620_001", "source_pdf": "source/original.pdf", "source_sha256": "dummy", "extraction_engine": "pymupdf", "page_count": 1, "pages": [{"page_number": 1, "has_text": True, "character_count": len(md_content), "text": md_content}], "generated_at": "2026-06-20T00:00:00Z"}, f)

    # quotation_text.json
    with open(package_dir / "text" / "quotation_text.json", "w", encoding="utf-8") as f:
        json.dump({"schema_version": "1.0", "quotation_id": "AUT_20260620_001", "source_raw_text": "source/raw_text.json", "source_sha256": _get_sha256(raw_text_path), "page_count": 1, "total_characters": len(md_content), "pages_with_text": 1, "markdown_path": "text/quotation.md", "pages": [{"page_number": 1, "has_text": True, "character_count": len(md_content), "start_offset": 0, "end_offset": len(md_content)}], "generated_at": "2026-06-20T00:00:00Z"}, f)

    # line_candidates.json
    line_cand_file = package_dir / "parsed" / "line_candidates.json"
    line_cand_file.parent.mkdir(parents=True, exist_ok=True)
    line_cand_manifest = {
        "schema_version": "1.0",
        "quotation_id": "AUT_20260620_001",
        "source_text_manifest": "text/quotation_text.json",
        "source_markdown": "text/quotation.md",
        "source_sha256": _get_sha256(md_path),
        "parser_name": "rule",
        "parser_version": "1.0",
        "candidate_count": 4,
        "candidates": [
            {"candidate_id": "AUT_20260620_001_LINECAND_0001", "page_number": 1, "line_number": 1, "raw_line": "mock", "confidence": 0.9, "warnings": [], "evidence": {"source_path": "text/quotation.md", "start_offset": 0, "end_offset": 4, "text": "mock"}},
            {"candidate_id": "AUT_20260620_001_LINECAND_0002", "page_number": 1, "line_number": 2, "raw_line": "mock", "confidence": 0.9, "warnings": [], "evidence": {"source_path": "text/quotation.md", "start_offset": 5, "end_offset": 9, "text": "mock"}},
            {"candidate_id": "AUT_20260620_001_LINECAND_0003", "page_number": 1, "line_number": 3, "raw_line": "mock", "confidence": 0.9, "warnings": [], "evidence": {"source_path": "text/quotation.md", "start_offset": 10, "end_offset": 14, "text": "mock"}},
            {"candidate_id": "AUT_20260620_001_LINECAND_0004", "page_number": 1, "line_number": 4, "raw_line": "mock", "confidence": 0.9, "warnings": [], "evidence": {"source_path": "text/quotation.md", "start_offset": 15, "end_offset": 19, "text": "mock"}}
        ],
        "warnings": [],
        "generated_at": "2026-06-20T00:00:00Z"
    }
    with open(line_cand_file, "w", encoding="utf-8") as f:
        json.dump(line_cand_manifest, f)

    # row_candidates.json
    row_cand_file = package_dir / "parsed" / "row_candidates.json"
    row_manifest = {
        "schema_version": "1.0",
        "quotation_id": "AUT_20260620_001",
        "source_line_candidates": "parsed/line_candidates.json",
        "source_sha256": _get_sha256(line_cand_file),
        "source_text_manifest": "text/quotation_text.json",
        "assembler_name": "rule",
        "assembler_version": "1.0",
        "row_count": 4,
        "rows": [
            {"row_id": "AUT_20260620_001_ROWCAND_0001", "page_number": 1, "start_line_number": 1, "end_line_number": 2, "source_candidate_ids": ["AUT_20260620_001_LINECAND_0001"], "description_candidate": "Item 1", "brand_candidate": None, "unit_candidate": "cái", "quantity_candidate": 10.0, "unit_price_candidate": 5000.0, "currency_candidate": "VND", "evidence_text": "mock", "start_offset": 0, "end_offset": 4, "confidence": 0.9, "warnings": []},
            {"row_id": "AUT_20260620_001_ROWCAND_0002", "page_number": 1, "start_line_number": 2, "end_line_number": 3, "source_candidate_ids": ["AUT_20260620_001_LINECAND_0002"], "description_candidate": "Item 2", "brand_candidate": None, "unit_candidate": "m", "quantity_candidate": 2.0, "unit_price_candidate": 10000.0, "currency_candidate": "VND", "evidence_text": "mock", "start_offset": 5, "end_offset": 9, "confidence": 0.8, "warnings": []},
            {"row_id": "AUT_20260620_001_ROWCAND_0003", "page_number": 1, "start_line_number": 3, "end_line_number": 4, "source_candidate_ids": ["AUT_20260620_001_LINECAND_0003"], "description_candidate": "Item 3", "brand_candidate": None, "unit_candidate": "bộ", "quantity_candidate": 5.0, "unit_price_candidate": 20000.0, "currency_candidate": "VND", "evidence_text": "mock", "start_offset": 10, "end_offset": 14, "confidence": 0.8, "warnings": []},
            {"row_id": "AUT_20260620_001_ROWCAND_0004", "page_number": 1, "start_line_number": 4, "end_line_number": 5, "source_candidate_ids": ["AUT_20260620_001_LINECAND_0004"], "description_candidate": "Item 4", "brand_candidate": None, "unit_candidate": "cái", "quantity_candidate": 1.0, "unit_price_candidate": 50000.0, "currency_candidate": None, "evidence_text": "mock", "start_offset": 15, "end_offset": 19, "confidence": 0.8, "warnings": []}
        ],
        "warnings": [],
        "generated_at": "2026-06-20T00:00:00Z"
    }
    with open(row_cand_file, "w", encoding="utf-8") as f:
        json.dump(row_manifest, f)

    # item_candidates.json
    item_cand_file = package_dir / "parsed" / "item_candidates.json"
    item_manifest = {
        "schema_version": "1.0",
        "quotation_id": "AUT_20260620_001",
        "source_row_candidates": "parsed/row_candidates.json",
        "source_sha256": _get_sha256(row_cand_file),
        "source_text_manifest": "text/quotation_text.json",
        "builder_name": "rule",
        "builder_version": "1.0",
        "item_count": 4,
        "items": [
            {"item_candidate_id": "AUT_20260620_001_ITEMCAND_0001", "source_row_id": "AUT_20260620_001_ROWCAND_0001", "page_number": 1, "start_line_number": 1, "end_line_number": 2, "description_candidate": "Item 1", "brand_candidate": None, "unit_candidate": "cái", "quantity_candidate": 10.0, "unit_price_candidate": 5000.0, "amount_candidate": 50000.0, "currency_candidate": "VND", "raw_evidence_text": "mock", "start_offset": 0, "end_offset": 4, "confidence": 0.9, "warnings": []},
            {"item_candidate_id": "AUT_20260620_001_ITEMCAND_0002", "source_row_id": "AUT_20260620_001_ROWCAND_0002", "page_number": 1, "start_line_number": 2, "end_line_number": 3, "description_candidate": "Item 2", "brand_candidate": None, "unit_candidate": "m", "quantity_candidate": 2.0, "unit_price_candidate": 10000.0, "amount_candidate": 20000.0, "currency_candidate": "VND", "raw_evidence_text": "mock", "start_offset": 5, "end_offset": 9, "confidence": 0.8, "warnings": []},
            {"item_candidate_id": "AUT_20260620_001_ITEMCAND_0003", "source_row_id": "AUT_20260620_001_ROWCAND_0003", "page_number": 1, "start_line_number": 3, "end_line_number": 4, "description_candidate": "Item 3", "brand_candidate": None, "unit_candidate": "bộ", "quantity_candidate": 5.0, "unit_price_candidate": 20000.0, "amount_candidate": 100000.0, "currency_candidate": "VND", "raw_evidence_text": "mock", "start_offset": 10, "end_offset": 14, "confidence": 0.8, "warnings": []},
            {"item_candidate_id": "AUT_20260620_001_ITEMCAND_0004", "source_row_id": "AUT_20260620_001_ROWCAND_0004", "page_number": 1, "start_line_number": 4, "end_line_number": 5, "description_candidate": "Item 4", "brand_candidate": None, "unit_candidate": "cái", "quantity_candidate": 1.0, "unit_price_candidate": 50000.0, "amount_candidate": 50000.0, "currency_candidate": None, "raw_evidence_text": "mock", "start_offset": 15, "end_offset": 19, "confidence": 0.8, "warnings": []}
        ],
        "warnings": [],
        "generated_at": "2026-06-20T00:00:00Z"
    }
    with open(item_cand_file, "w", encoding="utf-8") as f:
        json.dump(item_manifest, f)

    # normalized_draft.json
    draft_dir = package_dir / "normalized"
    draft_dir.mkdir(parents=True, exist_ok=True)
    draft_file = draft_dir / "normalized_draft.json"
    draft_data = {
        "schema_version": "1.0",
        "quotation_id": "AUT_20260620_001",
        "supplier_code": "AUT",
        "quotation_date": "2026-06-20",
        "currency": "VND",
        "source_item_candidates": "parsed/item_candidates.json",
        "source_sha256": _get_sha256(item_cand_file),
        "draft_builder_name": "rule",
        "draft_builder_version": "1.0",
        "item_count": 4,
        "review_required_count": 4,
        "items": [
            {"draft_item_id": "AUT_20260620_001_DRAFTITEM_0001", "source_item_candidate_id": "AUT_20260620_001_ITEMCAND_0001", "source_row_id": "AUT_20260620_001_ROWCAND_0001", "page_number": 1, "start_line_number": 1, "end_line_number": 2, "description": "Item 1", "unit": "cái", "quantity": 10.0, "unit_price": 5000.0, "currency": "VND", "amount": 50000.0, "review_status": "needs_review", "review_reasons": [], "confidence": 0.9, "warnings": [], "evidence": {"raw_evidence_text": "mock", "start_offset": 0, "end_offset": 4}},
            {"draft_item_id": "AUT_20260620_001_DRAFTITEM_0002", "source_item_candidate_id": "AUT_20260620_001_ITEMCAND_0002", "source_row_id": "AUT_20260620_001_ROWCAND_0002", "page_number": 1, "start_line_number": 2, "end_line_number": 3, "description": "Item 2", "unit": None, "quantity": None, "unit_price": 10000.0, "currency": "VND", "amount": None, "review_status": "needs_review", "review_reasons": [], "confidence": 0.8, "warnings": [], "evidence": {"raw_evidence_text": "mock", "start_offset": 5, "end_offset": 9}},
            {"draft_item_id": "AUT_20260620_001_DRAFTITEM_0003", "source_item_candidate_id": "AUT_20260620_001_ITEMCAND_0003", "source_row_id": "AUT_20260620_001_ROWCAND_0003", "page_number": 1, "start_line_number": 3, "end_line_number": 4, "description": "Item 3", "unit": "bộ", "quantity": 5.0, "unit_price": 20000.0, "currency": "VND", "amount": 100000.0, "review_status": "needs_review", "review_reasons": [], "confidence": 0.8, "warnings": [], "evidence": {"raw_evidence_text": "mock", "start_offset": 10, "end_offset": 14}},
            {"draft_item_id": "AUT_20260620_001_DRAFTITEM_0004", "source_item_candidate_id": "AUT_20260620_001_ITEMCAND_0004", "source_row_id": "AUT_20260620_001_ROWCAND_0004", "page_number": 1, "start_line_number": 4, "end_line_number": 5, "description": "Item 4", "unit": "cái", "quantity": 1.0, "unit_price": 50000.0, "currency": None, "amount": 50000.0, "review_status": "needs_review", "review_reasons": [], "confidence": 0.8, "warnings": [], "evidence": {"raw_evidence_text": "mock", "start_offset": 15, "end_offset": 19}}
        ],
        "warnings": [],
        "generated_at": "2026-06-20T00:00:00Z"
    }
    with open(draft_file, "w", encoding="utf-8") as f:
        json.dump(draft_data, f)

    # Đăng ký các files vào package.json
    pkg_json_path = package_dir / "package.json"
    from mep_quotation.package.loader import load_package_json
    pkg_model = load_package_json(package_dir)
    pkg_model.files.line_candidates = "parsed/line_candidates.json"
    pkg_model.files.row_candidates = "parsed/row_candidates.json"
    pkg_model.files.item_candidates = "parsed/item_candidates.json"
    pkg_model.files.normalized_draft = "normalized/normalized_draft.json"
    with open(pkg_json_path, "w", encoding="utf-8") as f:
        json.dump(pkg_model.model_dump(mode="json"), f)

    # 3. Tạo review decisions thực tế bằng API Phase 10
    from mep_quotation.review.review_service import record_review_decision
    from mep_quotation.spec.models import ReviewFieldOverridesModel
    
    # Item 1: approved
    record_review_decision(
        package_dir,
        draft_item_id="AUT_20260620_001_DRAFTITEM_0001",
        decision_type="approved",
        reviewer="test_reviewer",
        reason="approved details"
    )
    
    # Item 2: edited (có Formula Injection và control characters)
    overrides = ReviewFieldOverridesModel(
        description="=1+2 Formula Injection",
        brand=None,
        unit=None,
        quantity=None,
        unit_price=10000.0,
        amount=None,
        currency="VND"
    )
    record_review_decision(
        package_dir,
        draft_item_id="AUT_20260620_001_DRAFTITEM_0002",
        decision_type="edited",
        reviewer="test_reviewer",
        reason="edited details",
        field_overrides=overrides
    )
    
    # Item 3: rejected
    record_review_decision(
        package_dir,
        draft_item_id="AUT_20260620_001_DRAFTITEM_0003",
        decision_type="rejected",
        reviewer="test_reviewer",
        reason="rejected details"
    )

    # 4. Xuất normalized.json chính thức bằng API Phase 11
    from mep_quotation.normalized_export.export_service import export_normalized
    export_normalized(package_dir, overwrite=True)
    
    # 5. Load normalized.json vừa xuất ra để tiêm thêm warnings và control characters
    norm_path = package_dir / "normalized" / "normalized.json"
    with open(norm_path, "r", encoding="utf-8") as f:
        norm_data = json.load(f)
    
    # Thêm warnings ở file-level và item-level
    norm_data["warnings"] = [{"code": "file_w", "message": "File warning example"}]
    norm_data["items"][0]["warnings"] = [{"code": "w_test", "message": "Item warning example"}]
    norm_data["items"][0]["evidence_text"] = "CV-3X2.5 100m"
    norm_data["items"][0]["reviewer"] = "test_reviewer"
    
    # Tiêm XML control characters vào evidence_text của item 2
    norm_data["items"][1]["evidence_text"] = "Control \x00 character \x1f test"
    norm_data["items"][1]["reviewer"] = "test_reviewer"
    
    # Ghi lại tệp normalized.json đã tiêm dữ liệu
    from mep_quotation.spec.models import NormalizedQuotationModel
    write_json_file(norm_path, NormalizedQuotationModel.model_validate(norm_data))
    
    return package_dir

def test_export_excel_success(sample_package):
    """Kiểm thử xuất Excel thành công, kiểm tra cấu trúc workbook, các ô, định dạng và manifest."""
    # Tính SHA256 của các source files trước khi xuất bản
    norm_path = sample_package / "normalized" / "normalized.json"
    decisions_path = sample_package / "review" / "review_decisions.json"
    
    sha_norm_before = _get_sha256(norm_path)
    sha_decisions_before = _get_sha256(decisions_path)
    
    # Thực hiện xuất Excel
    excel_path = export_excel(sample_package, overwrite=False)
    
    # 1. Xác nhận Phase 12 không sửa đổi bất kỳ tệp tin nguồn nào
    assert _get_sha256(norm_path) == sha_norm_before
    assert _get_sha256(decisions_path) == sha_decisions_before
    
    # 2. Kiểm tra sự tồn tại của các file xuất ra
    manifest_path = sample_package / "exports" / "export_manifest.json"
    assert excel_path.exists()
    assert manifest_path.exists()
    
    # 3. Load và validate manifest.json
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest_data = json.load(f)
        
    assert manifest_data["quotation_id"] == "AUT_20260620_001"
    assert manifest_data["supplier_code"] == "AUT"
    assert manifest_data["quotation_date"] == "2026-06-20"
    assert manifest_data["source_normalized"] == "normalized/normalized.json"
    assert manifest_data["source_normalized_sha256"] == sha_norm_before
    assert manifest_data["export_file_sha256"] == _get_sha256(excel_path)
    assert manifest_data["sheet_count"] == 4
    
    sheets = manifest_data["sheets"]
    assert len(sheets) == 4
    assert sheets[0]["name"] == "Summary"
    assert sheets[0]["row_count"] == 16 # 10 fields mặc định + 6 fields review summary
    assert sheets[1]["name"] == "Items"
    assert sheets[1]["row_count"] == 2 # 2 items
    assert sheets[2]["name"] == "Warnings"
    assert sheets[2]["row_count"] == 2 # 1 file warning + 1 item warning
    assert sheets[3]["name"] == "Trace"
    assert sheets[3]["row_count"] == 2
    
    # 4. Load và validate cấu trúc Excel workbook
    wb = openpyxl.load_workbook(excel_path)
    assert wb.sheetnames == ["Summary", "Items", "Warnings", "Trace"]
    
    # Kiểm tra Summary sheet
    ws_sum = wb["Summary"]
    assert ws_sum.cell(row=1, column=1).value == "Field"
    assert ws_sum.cell(row=1, column=2).value == "Value"
    assert ws_sum.cell(row=2, column=1).value == "Quotation ID"
    assert ws_sum.cell(row=2, column=2).value == "AUT_20260620_001"
    assert ws_sum.cell(row=5, column=1).value == "Currency"
    assert ws_sum.cell(row=5, column=2).value == "VND"
    assert ws_sum.cell(row=6, column=1).value == "Item Count"
    assert ws_sum.cell(row=6, column=2).value == 2
    
    # Kiểm tra Items sheet
    ws_items = wb["Items"]
    # Header bold
    assert ws_items.cell(row=1, column=1).font.bold is True
    # Row 2 (Item 1)
    assert ws_items.cell(row=2, column=1).value == "AUT_20260620_001_ITEM_0001"
    assert ws_items.cell(row=2, column=3).value == "Item 1"
    assert isinstance(ws_items.cell(row=2, column=6).value, (int, float)) # quantity
    assert ws_items.cell(row=2, column=6).value == 10.0
    assert ws_items.cell(row=2, column=7).value == 5000.0
    assert ws_items.cell(row=2, column=9).value == 50000.0
    assert ws_items.cell(row=2, column=15).value == "[w_test] Item warning example" # warnings
    
    # Row 3 (Item 2 - Test Formula Injection & Null Fields)
    # Tránh Formula Injection bằng nháy đơn ở đầu ô
    assert ws_items.cell(row=3, column=3).value == "'=1+2 Formula Injection"
    # Null fields được ghi thành ô trống
    assert ws_items.cell(row=3, column=4).value is None # brand
    assert ws_items.cell(row=3, column=6).value is None # quantity
    assert ws_items.cell(row=3, column=9).value is None # amount
    
    # Kiểm tra Warnings sheet
    ws_warn = wb["Warnings"]
    assert ws_warn.cell(row=2, column=1).value == "file"
    assert ws_warn.cell(row=2, column=2).value in (None, "")
    assert ws_warn.cell(row=2, column=3).value == "file_w"
    
    assert ws_warn.cell(row=3, column=1).value == "item"
    assert ws_warn.cell(row=3, column=2).value == "AUT_20260620_001_ITEM_0001"
    assert ws_warn.cell(row=3, column=3).value == "w_test"
    
    # Kiểm tra Trace sheet & Sanitize control characters
    ws_trace = wb["Trace"]
    assert ws_trace.cell(row=3, column=1).value == "AUT_20260620_001_ITEM_0002"
    # Lọc bỏ các ký tự \x00 và \x1f
    assert ws_trace.cell(row=3, column=5).value == "Control  character  test"
    
    # 5. Chạy validate package integrity thành công
    validate_package_integrity(sample_package)

def test_export_excel_overwrite_rule(sample_package):
    """Kiểm thử cản ghi đè khi overwrite=False và ghi đè hoàn tất khi overwrite=True."""
    # Thực hiện xuất Excel lần đầu tiên
    export_excel(sample_package, overwrite=False)
    
    excel_path = sample_package / "exports" / "quotation.xlsx"
    manifest_path = sample_package / "exports" / "export_manifest.json"
    
    assert excel_path.exists()
    assert manifest_path.exists()
    
    # Lần 2: overwrite=False -> Phải fail rõ ràng
    with pytest.raises(ValueError, match="Excel export files already exist"):
        export_excel(sample_package, overwrite=False)
        
    # Lần 3: xóa tệp manifest nhưng để lại tệp Excel -> overwrite=False vẫn phải fail
    os.remove(manifest_path)
    with pytest.raises(ValueError, match="Excel export files already exist"):
        export_excel(sample_package, overwrite=False)
        
    # Lần 4: overwrite=True -> Ghi đè thành công
    export_excel(sample_package, overwrite=True)
    assert excel_path.exists()
    assert manifest_path.exists()

def test_export_excel_mismatch_item_count(sample_package):
    """Kiểm thử export fail khi item_count bị lệch."""
    norm_path = sample_package / "normalized" / "normalized.json"
    with open(norm_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    # Sửa item_count lệch với thực tế items
    data["item_count"] = 10 
    write_json_file(norm_path, NormalizedQuotationModel.model_validate(data))
    
    with pytest.raises(ValueError, match="item_count mismatch"):
        export_excel(sample_package, overwrite=True)

def test_export_excel_package_json_update_fail(sample_package):
    """Kiểm thử ném lỗi rõ ràng khi cập nhật package.json lỗi sau khi ghi Excel/manifest."""
    from unittest.mock import patch
    from mep_quotation.excel_export import export_service
    
    original_write = export_service.write_json_file
    
    def mock_write(path, model):
        if Path(path).name == "package.json":
            raise IOError("Disk full or permission denied")
        return original_write(path, model)
        
    with patch("mep_quotation.excel_export.export_service.write_json_file", side_effect=mock_write):
        with pytest.raises(RuntimeError, match="Excel export failed during package.json update"):
            export_excel(sample_package, overwrite=True)

def test_export_excel_missing_exports_dir_success(sample_package):
    """Kiểm thử thư mục exports/ chưa tồn tại thì tự động tạo và không fail."""
    exports_dir = sample_package / "exports"
    if exports_dir.exists():
        shutil.rmtree(exports_dir)
        
    excel_path = export_excel(sample_package, overwrite=False)
    assert excel_path.exists()
