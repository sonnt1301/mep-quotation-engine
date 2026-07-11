# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.run_profile_write_simulation import run_write_simulation

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_write_simulation_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_simulation_blocked_by_commit_gate():
    tmp_path = make_tmp_dir("blocked_gate")
    gate_manifest = tmp_path / "profile_commit_gate_manifest.json"
    candidates_json = tmp_path / "write_candidate_items.json"
    candidates_summary = tmp_path / "write_candidate_summary.json"
    out_dir = tmp_path / "out"

    # Gate ở trạng thái PENDING
    gate_data = {
        "commit_gate_status": "PENDING_HUMAN_APPROVAL",
        "ready_for_write_to_main_pipeline": False,
        "blocking_reasons": []
    }
    # Candidates có 1 item
    candidates_data = [{
        "write_candidate_id": "WRITE_CANDIDATE_0001",
        "supplier_code": "CHINT",
        "material_code": "NXM-63S",
        "description": "MCCB Chint",
        "unit": "cái",
        "quantity": 1,
        "unit_price": 800000,
        "amount": 800000,
        "currency": "VND",
        "page_number": 3,
        "write_key": "CHINT|NXM-63S|P3|800000|MCCB",
        "human_decision": "APPROVE",
        "human_note": "",
        "provenance": "Source PDF, Page 3",
        "evidence_text": "Chint NXM 800,000",
        "proposed_action": "INSERT_CANDIDATE"
    }]
    cand_sum = {
        "summary": {
            "candidate_items_count": 1,
            "duplicate_write_key_count": 0
        }
    }

    gate_manifest.write_text(json.dumps(gate_data, ensure_ascii=False), encoding="utf-8")
    candidates_json.write_text(json.dumps(candidates_data, ensure_ascii=False), encoding="utf-8")
    candidates_summary.write_text(json.dumps(cand_sum, ensure_ascii=False), encoding="utf-8")

    try:
        manifest = run_write_simulation(gate_manifest, candidates_json, candidates_summary, out_dir)

        # Trạng thái bắt buộc là BLOCKED_BY_COMMIT_GATE
        assert manifest["summary"]["simulation_status"] == "BLOCKED_BY_COMMIT_GATE"
        assert manifest["summary"]["simulated_records_count"] == 0
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_real_write"] is False
        
        # Vẫn sinh đầy đủ file preview rỗng
        assert (out_dir / "simulated_material_records.json").exists()
        assert (out_dir / "simulated_commit_log.json").exists()
        assert (out_dir / "simulated_rollback_plan.json").exists()
        assert (out_dir / "simulated_write_summary.json").exists()
        assert (out_dir / "simulated_write_review.xlsx").exists()
        assert (out_dir / "simulated_write_report.md").exists()
        
    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_simulation_success_when_gate_approved():
    tmp_path = make_tmp_dir("approved_gate")
    gate_manifest = tmp_path / "profile_commit_gate_manifest.json"
    candidates_json = tmp_path / "write_candidate_items.json"
    candidates_summary = tmp_path / "write_candidate_summary.json"
    out_dir = tmp_path / "out"

    # Gate ở trạng thái APPROVED
    gate_data = {
        "commit_gate_status": "APPROVED_FOR_NEXT_PHASE_DESIGN_ONLY",
        "ready_for_write_to_main_pipeline": False,
        "blocking_reasons": []
    }
    
    candidates_data = [
        {
            "write_candidate_id": "WRITE_CANDIDATE_0001",
            "supplier_code": "CHINT",
            "material_code": "NXM-63S",
            "description": "MCCB Chint",
            "unit": "cái",
            "quantity": 1,
            "unit_price": 800000,
            "amount": 800000,
            "currency": "VND",
            "page_number": 3,
            "write_key": "CHINT|NXM-63S|P3|800000|MCCB",
            "human_decision": "APPROVE",
            "human_note": "",
            "provenance": "Source PDF, Page 3",
            "evidence_text": "Chint NXM 800,000",
            "proposed_action": "INSERT_CANDIDATE"
        },
        {
            "write_candidate_id": "WRITE_CANDIDATE_0002",
            "supplier_code": "LS",
            "material_code": "MC-9B",
            "description": "Contactor LS",
            "unit": "cái",
            "quantity": 2,
            "unit_price": 250000,
            "amount": 500000,
            "currency": "VND",
            "page_number": 4,
            "write_key": "LS|MC-9B|P4|250000|CONTACTOR",
            "human_decision": "APPROVE",
            "human_note": "",
            "provenance": "Source PDF, Page 4",
            "evidence_text": "MC-9B 250k",
            "proposed_action": "INSERT_CANDIDATE"
        }
    ]
    
    cand_sum = {
        "summary": {
            "candidate_items_count": 2,
            "duplicate_write_key_count": 0
        }
    }

    gate_manifest.write_text(json.dumps(gate_data, ensure_ascii=False), encoding="utf-8")
    candidates_json.write_text(json.dumps(candidates_data, ensure_ascii=False), encoding="utf-8")
    candidates_summary.write_text(json.dumps(cand_sum, ensure_ascii=False), encoding="utf-8")

    try:
        manifest = run_write_simulation(gate_manifest, candidates_json, candidates_summary, out_dir)

        # Trạng thái phải là READY
        assert manifest["summary"]["simulation_status"] == "SIMULATION_READY_FOR_REVIEW"
        assert manifest["summary"]["simulated_records_count"] == 2
        assert manifest["summary"]["would_insert_count"] == 2
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_real_write"] is False
        
        # Kiểm tra nội dung simulated records
        records = manifest["simulated_records"]
        assert len(records) == 2
        assert records[0]["simulated_result"] == "WOULD_INSERT"
        assert records[0]["ready_for_real_write"] is False
        
        # Kiểm tra warning not_checked_against_master_database
        log_entries = manifest["commit_log"]["log_entries"]
        assert "not_checked_against_master_database" in log_entries[0]["message"]
        
        # Kiểm tra rollback plan
        rollback_plan = manifest["rollback_plan"]
        assert rollback_plan["rollback_available"] is True
        assert len(rollback_plan["rollback_actions"]) == 2
        
        # Kiểm tra Excel đủ 4 sheet
        wb = openpyxl.load_workbook(out_dir / "simulated_write_review.xlsx")
        assert "Summary" in wb.sheetnames
        assert "Simulated Records" in wb.sheetnames
        assert "Commit Log" in wb.sheetnames
        assert "Rollback Plan" in wb.sheetnames
        wb.close()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_and_adapter_blocked_read():
    # Kiểm tra xem run_profile_write_simulation.py có đọc file items json gốc hoặc blocked items của adapter không
    script_path = PROJECT_ROOT / "tools/feasibility/run_profile_write_simulation.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    # Script không được phép chứa tham chiếu trực tiếp đến file bridge gốc hoặc blocked items của adapter
    assert "profile_bridge_items.json" not in content
    assert "blocked_items.json" not in content
