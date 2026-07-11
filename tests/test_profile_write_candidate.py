# -*- coding: utf-8 -*-
import json
import os
import shutil
import uuid
from pathlib import Path
import pytest
import openpyxl

from tools.feasibility.export_profile_write_candidate import run_candidate_pipeline

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def make_tmp_dir(name):
    candidates_dir = [
        PROJECT_ROOT / "scratch" / "pytest_tmp",
        Path("C:/tmp"),
    ]
    dirname = f"mep_profile_write_candidate_{name}_{uuid.uuid4().hex}"
    for base in candidates_dir:
        try:
            path = base / dirname
            path.mkdir(parents=True, exist_ok=True)
            return path
        except OSError:
            continue
    raise OSError("No writable temporary directory found.")


def test_write_candidate_pipeline_success():
    tmp_path = make_tmp_dir("success")
    preview_json = tmp_path / "normalized_items_preview.json"
    summary_json = tmp_path / "profile_write_adapter_summary.json"
    out_dir = tmp_path / "out"

    # Chuẩn bị dữ liệu preview giả lập có 2 items hợp lệ (không trùng)
    preview_data = [
        {
            "item_id": "ITEM_0001",
            "supplier_code": "ABB",
            "material_code": "ACS310",
            "description": "Biến tần ACS310",
            "unit": "cái",
            "quantity": 1.0,
            "unit_price": 5000000,
            "amount": 5000000.0,
            "currency": "VND",
            "page_number": 3,
            "write_key": "ABB|ACS310|P3|5000000|BIENTAN",
            "human_decision": "APPROVE",
            "human_note": "ok",
            "provenance": "Source PDF, Page 3",
            "evidence_text": "Biến tần ACS310 5,000,000"
        },
        {
            "item_id": "ITEM_0002",
            "supplier_code": "ABB",
            "material_code": "ACS355",
            "description": "Biến tần ACS355",
            "unit": "cái",
            "quantity": 2.0,
            "unit_price": 7000000,
            "amount": 14000000.0,
            "currency": "VND",
            "page_number": 3,
            "write_key": "ABB|ACS355|P3|7000000|BIENTAN",
            "human_decision": "EDIT_AND_APPROVE",
            "human_note": "Sửa số lượng",
            "provenance": "Source PDF, Page 3",
            "evidence_text": "Biến tần ACS355 7,000,000"
        }
    ]
    
    preview_json.write_text(json.dumps(preview_data, ensure_ascii=False), encoding="utf-8")
    summary_json.write_text(json.dumps({"input_bridge_items_count": 2}, ensure_ascii=False), encoding="utf-8")

    try:
        manifest = run_candidate_pipeline(preview_json, summary_json, out_dir)

        # 1. Kiểm tra an toàn: Không set ready_for_write_to_main_pipeline = True
        assert manifest["ready_for_write_to_main_pipeline"] is False
        assert manifest["summary"]["ready_for_write_to_main_pipeline"] is False
        assert manifest["write_target"] == "write_candidate_preview_only"
        
        # 2. Kiểm tra counts
        assert manifest["summary"]["candidate_items_count"] == 2
        assert manifest["summary"]["skipped_items_count"] == 0
        assert manifest["summary"]["duplicate_write_key_count"] == 0
        
        # 3. Trạng thái proposed_status
        assert manifest["summary"]["proposed_status"] == "READY_FOR_HUMAN_COMMIT_REVIEW"
        
        # 4. Kiểm tra sự sinh ra của các tệp tin đầu ra
        assert (out_dir / "write_candidate_items.json").exists()
        assert (out_dir / "write_candidate_summary.json").exists()
        assert (out_dir / "write_candidate_review.csv").exists()
        assert (out_dir / "write_candidate_review.xlsx").exists()
        assert (out_dir / "write_candidate_commit_plan.md").exists()
        
        # 5. Kiểm tra Excel đủ 3 sheet (linh hoạt tên sheet Warnings)
        wb = openpyxl.load_workbook(out_dir / "write_candidate_review.xlsx")
        assert "Summary" in wb.sheetnames
        assert "Write Candidates" in wb.sheetnames
        assert any("Warnings" in s and "Duplicates" in s for s in wb.sheetnames)
        wb.close()

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_write_candidate_with_duplicate_write_keys():
    tmp_path = make_tmp_dir("duplicates")
    preview_json = tmp_path / "normalized_items_preview.json"
    summary_json = tmp_path / "profile_write_adapter_summary.json"
    out_dir = tmp_path / "out"

    # Dữ liệu chứa 2 items bị trùng write_key
    preview_data = [
        {
            "item_id": "ITEM_0001",
            "supplier_code": "LS",
            "material_code": "MC-9B",
            "description": "Contactor MC-9B",
            "unit": "cái",
            "quantity": 1.0,
            "unit_price": 250000,
            "amount": 250000.0,
            "currency": "VND",
            "page_number": 4,
            "write_key": "LS|MC-9B|P4|250000|CONTACTOR",
            "human_decision": "APPROVE",
            "human_note": "",
            "provenance": "Source PDF, Page 4",
            "evidence_text": "MC-9B 250k"
        },
        {
            "item_id": "ITEM_0002",
            "supplier_code": "LS",
            "material_code": "MC-9B",
            "description": "Contactor MC-9B trùng",
            "unit": "cái",
            "quantity": 1.0,
            "unit_price": 250000,
            "amount": 250000.0,
            "currency": "VND",
            "page_number": 4,
            "write_key": "LS|MC-9B|P4|250000|CONTACTOR",
            "human_decision": "APPROVE",
            "human_note": "",
            "provenance": "Source PDF, Page 4",
            "evidence_text": "MC-9B 250k"
        }
    ]

    preview_json.write_text(json.dumps(preview_data, ensure_ascii=False), encoding="utf-8")
    summary_json.write_text(json.dumps({"input_bridge_items_count": 2}, ensure_ascii=False), encoding="utf-8")

    try:
        manifest = run_candidate_pipeline(preview_json, summary_json, out_dir)

        # Trùng write_key phải được phát hiện và mark SKIP_CANDIDATE cho dòng thứ 2
        assert manifest["summary"]["candidate_items_count"] == 1
        assert manifest["summary"]["skipped_items_count"] == 1
        assert manifest["summary"]["duplicate_write_key_count"] == 1
        
        # proposed_action của dòng trùng thứ 2 phải là SKIP_CANDIDATE
        candidates = manifest["candidates"]
        assert candidates[0]["proposed_action"] == "INSERT_CANDIDATE"
        assert candidates[1]["proposed_action"] == "SKIP_CANDIDATE"
        assert "Trùng khóa ghi" in candidates[1]["human_note"]

        # Trạng thái proposed_status phải bị chặn
        assert manifest["summary"]["proposed_status"] == "BLOCKED_NEEDS_REVIEW"

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_write_candidate_empty_input():
    tmp_path = make_tmp_dir("empty")
    preview_json = tmp_path / "normalized_items_preview.json"
    summary_json = tmp_path / "profile_write_adapter_summary.json"
    out_dir = tmp_path / "out"

    # Ghi danh sách preview rỗng
    preview_json.write_text("[]", encoding="utf-8")
    summary_json.write_text("{}", encoding="utf-8")

    try:
        manifest = run_candidate_pipeline(preview_json, summary_json, out_dir)

        # preview rỗng thì proposed_status = BLOCKED_NEEDS_REVIEW
        assert manifest["summary"]["candidate_items_count"] == 0
        assert manifest["summary"]["proposed_status"] == "BLOCKED_NEEDS_REVIEW"
        assert len(manifest["candidates"]) == 0

    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def test_no_raw_bridge_items_read():
    # Kiểm tra xem export_profile_write_candidate.py có đọc file ITEMS_JSON gốc trực tiếp không.
    script_path = PROJECT_ROOT / "tools/feasibility/export_profile_write_candidate.py"
    assert script_path.exists()
    
    with open(script_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    # Script không được phép chứa tham chiếu trực tiếp đến file items json của bridge
    assert "profile_bridge_items.json" not in content
