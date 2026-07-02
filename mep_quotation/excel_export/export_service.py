import os
import json
import hashlib
from pathlib import Path
from datetime import datetime, timezone
import openpyxl

from mep_quotation.spec.models import (
    QuotationPackageModel,
    NormalizedQuotationModel,
    ExcelExportContextModel,
    ExcelExportSheetModel,
    ExcelExportManifestModel,
    AuditLogEntryModel
)
from mep_quotation.package.writer import write_json_file
from mep_quotation.excel_export.workbook_builder import build_excel_workbook
from mep_quotation.excel_export.export_manifest import build_excel_export_manifest

def _get_sha256(file_path: Path) -> str:
    """Tính toán chữ ký băm SHA256 của tệp tin."""
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            sha.update(chunk)
    return sha.hexdigest()

def _write_audit_log(package_path: Path, event: str, level: str, details: dict, quotation_id: str):
    """Ghi nhật ký sự kiện audit log cục bộ vào logs/processing.log.jsonl."""
    log_file = package_path / "logs" / "processing.log.jsonl"
    entry = AuditLogEntryModel(
        timestamp=datetime.now(timezone.utc),
        level=level,
        event=event,
        quotation_id=quotation_id,
        details=details
    )
    # Ghi nối đuôi dòng log
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(entry.model_dump_json() + "\n")

def export_excel(package_path: Path, overwrite: bool = False) -> Path:
    """Điều phối dịch vụ xuất dữ liệu báo giá chính thức sang Excel và ghi manifest."""
    package_path = Path(package_path)
    
    # 1. Load và validate package.json
    package_json_path = package_path / "package.json"
    if not package_json_path.exists():
        raise FileNotFoundError(f"package.json not found in package directory: {package_path}")
        
    with open(package_json_path, "r", encoding="utf-8") as f:
        package_data = json.load(f)
    package = QuotationPackageModel.model_validate(package_data)
    quotation_id = package.quotation_id
    
    # Ghi log bắt đầu
    _write_audit_log(package_path, "excel_export_started", "INFO", {"overwrite": overwrite}, quotation_id)
    
    try:
        # 2. Load và validate normalized/normalized.json
        normalized_path = package_path / "normalized" / "normalized.json"
        if not normalized_path.exists():
            raise FileNotFoundError(f"normalized.json not found at {normalized_path}")
            
        with open(normalized_path, "r", encoding="utf-8") as f:
            normalized_data = json.load(f)
        normalized = NormalizedQuotationModel.model_validate(normalized_data)
        
        # 3. Xác định và kiểm tra item_count
        actual_item_count = len(normalized.items)
        defined_item_count = normalized.item_count
        
        if defined_item_count is not None:
            if defined_item_count != actual_item_count:
                raise ValueError(
                    f"item_count mismatch: normalized.item_count ({defined_item_count}) "
                    f"does not match actual items size ({actual_item_count})"
                )
        
        item_count = actual_item_count
        
        # 4. Kiểm tra overwrite rule cho cả file Excel và manifest
        excel_path = package_path / "exports" / "quotation.xlsx"
        manifest_path = package_path / "exports" / "export_manifest.json"
        
        if not overwrite:
            if excel_path.exists() or manifest_path.exists():
                raise ValueError(
                    "Excel export files already exist. Set overwrite=True to replace them."
                )
                
        # 5. Tạo thư mục exports/ nếu chưa tồn tại
        exports_dir = package_path / "exports"
        exports_dir.mkdir(parents=True, exist_ok=True)
        
        # 6. Tính source_normalized_sha256 và khởi tạo context
        source_normalized_sha256 = _get_sha256(normalized_path)
        exported_at = datetime.now(timezone.utc)
        
        context = ExcelExportContextModel(
            source_normalized_sha256=source_normalized_sha256,
            exported_at=exported_at
        )
        
        # 7. Xây dựng Excel workbook
        wb = build_excel_workbook(normalized, context)
        _write_audit_log(package_path, "excel_workbook_built", "INFO", {}, quotation_id)
        
        # 8. Ghi file Excel bằng Atomic Write
        temp_excel_path = exports_dir / "quotation.xlsx.tmp"
        wb.save(temp_excel_path)
        wb.close()
        
        # Đổi tên nguyên tử
        if temp_excel_path.exists():
            os.replace(temp_excel_path, excel_path)
            
        _write_audit_log(package_path, "excel_workbook_written", "INFO", {"path": "exports/quotation.xlsx"}, quotation_id)
        
        # 9. Load lại và validate workbook
        wb_check = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb_check.sheetnames
        
        # Kiểm tra đủ 4 sheet và đúng thứ tự
        expected_sheets = ["Summary", "Items", "Warnings", "Trace"]
        if sheet_names != expected_sheets:
            wb_check.close()
            raise ValueError(
                f"Workbook validation failed: expected sheets {expected_sheets}, but got {sheet_names}"
            )
            
        # Kiểm tra row count
        ws_items = wb_check["Items"]
        # Đếm số dòng trong sheet Items (read_only=True yêu cầu duyệt qua các dòng)
        items_row_count = sum(1 for _ in ws_items.iter_rows()) - 1
        
        ws_trace = wb_check["Trace"]
        trace_row_count = sum(1 for _ in ws_trace.iter_rows()) - 1
        
        wb_check.close()
        
        if items_row_count != item_count:
            raise ValueError(
                f"Workbook validation failed: Items sheet row count ({items_row_count}) "
                f"does not match expected item count ({item_count})"
            )
        if trace_row_count != item_count:
            raise ValueError(
                f"Workbook validation failed: Trace sheet row count ({trace_row_count}) "
                f"does not match expected item count ({item_count})"
            )
            
        _write_audit_log(package_path, "excel_workbook_validated", "INFO", {}, quotation_id)
        
        # 10. Tính export_file_sha256
        export_file_sha256 = _get_sha256(excel_path)
        
        # 11. Đếm row count cho Warnings sheet
        # Warnings sheet có 1 header row + các file warning + các item warning
        # Ở đây ta có thể tính trực tiếp dòng warnings để ghi manifest
        file_warnings_count = len(normalized.warnings)
        item_warnings_count = sum(len(item.warnings) for item in normalized.items)
        warnings_row_count = file_warnings_count + item_warnings_count
        
        # Tóm tắt số dòng trong Summary sheet:
        # Summary có 1 header row + 10 dòng mặc định + 6 dòng review summary (nếu có)
        summary_rows_count = 10
        if normalized.export_summary:
            summary_rows_count += 6
            
        sheets_metadata = [
            ExcelExportSheetModel(name="Summary", row_count=summary_rows_count),
            ExcelExportSheetModel(name="Items", row_count=item_count),
            ExcelExportSheetModel(name="Warnings", row_count=warnings_row_count),
            ExcelExportSheetModel(name="Trace", row_count=item_count)
        ]
        
        # 12. Build manifest.json và ghi Atomic Write
        manifest = build_excel_export_manifest(
            package_path=package_path,
            normalized=normalized,
            export_file_sha256=export_file_sha256,
            workbook_sheets=sheets_metadata,
            export_context=context
        )
        
        temp_manifest_path = exports_dir / "export_manifest.json.tmp"
        write_json_file(temp_manifest_path, manifest)
        
        if temp_manifest_path.exists():
            os.replace(temp_manifest_path, manifest_path)
            
        _write_audit_log(package_path, "excel_export_manifest_written", "INFO", {"path": "exports/export_manifest.json"}, quotation_id)
        
        # 13. Cập nhật package.json
        try:
            package.files.excel_export = "exports/quotation.xlsx"
            package.files.excel_export_manifest = "exports/export_manifest.json"
            package.updated_at = datetime.now(timezone.utc)
            write_json_file(package_json_path, package)
        except Exception as e:
            # Ghi nhận audit log thất bại và ném lỗi dừng ngay lập tức
            _write_audit_log(package_path, "excel_export_failed", "ERROR", {"error": f"Failed to update package.json: {str(e)}"}, quotation_id)
            raise RuntimeError(f"Excel export failed during package.json update: {str(e)}")
            
        # Ghi log hoàn tất thành công
        _write_audit_log(package_path, "excel_export_completed", "INFO", {
            "excel_path": "exports/quotation.xlsx",
            "manifest_path": "exports/export_manifest.json"
        }, quotation_id)
        
        return excel_path
        
    except Exception as e:
        # Ghi nhận audit log lỗi
        _write_audit_log(package_path, "excel_export_failed", "ERROR", {"error": str(e)}, quotation_id)
        raise e
