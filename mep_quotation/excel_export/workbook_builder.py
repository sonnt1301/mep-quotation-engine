import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from mep_quotation.spec.models import NormalizedQuotationModel, ExcelExportContextModel

# Regex tìm các ký tự điều khiển ASCII < 32 không hợp lệ cho XML Excel 
# Ngoại trừ tab (\t), newline (\n), carriage return (\r)
XML_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def sanitize_excel_text(val) -> str:
    """Lọc sạch các ký tự điều khiển XML không hợp lệ trong Excel."""
    if val is None:
        return ""
    val_str = str(val)
    return XML_CONTROL_CHARS.sub("", val_str)

def escape_formula(val) -> str:
    """Escape chống Formula Injection cho các ô text bắt đầu bằng =, +, -, @."""
    if val is None:
        return ""
    val_str = sanitize_excel_text(val)
    if val_str and val_str[0] in ('=', '+', '-', '@'):
        return f"'{val_str}"
    return val_str

def apply_auto_width_and_styles(ws, max_width=60):
    """Áp dụng font in đậm cho dòng header, auto filter và auto column width."""
    # 1. Định dạng header dòng 1
    header_font = Font(name="Calibri", size=11, bold=True)
    for cell in ws[1]:
        cell.font = header_font
        
    # 2. Bật auto filter nếu có nhiều hơn 1 dòng
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
        
    # 3. Tính toán và áp dụng độ rộng cột tự động
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Lấy độ dài chuỗi hiển thị thực tế
            val = cell.value
            if val is not None:
                # Nếu là ô đã escape formula bằng nháy đơn ở đầu, độ dài thực tế không tính nháy đơn
                val_str = str(val)
                if val_str.startswith("'") and len(val_str) > 1 and val_str[1] in ('=', '+', '-', '@'):
                    val_str = val_str[1:]
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), max_width)

def build_excel_workbook(
    normalized: NormalizedQuotationModel,
    export_context: ExcelExportContextModel
) -> openpyxl.Workbook:
    """Pure function xây dựng Excel workbook từ dữ liệu normalized và export context."""
    wb = openpyxl.Workbook()
    
    # openpyxl mặc định tạo 1 sheet "Sheet", ta sẽ đổi tên nó thành "Summary"
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # ----------------------------------------------------
    # Sheet 1: Summary
    # ----------------------------------------------------
    # Xác định Currency theo quy tắc
    summary_currency = ""
    if normalized.currency:
        summary_currency = normalized.currency
    else:
        # Lấy unique currency từ items
        currencies = {item.currency.upper() for item in normalized.items if item.currency}
        if len(currencies) == 1:
            summary_currency = list(currencies)[0]
        elif len(currencies) > 1:
            summary_currency = "MULTIPLE"
            
    # Xác định Item Count
    item_count = len(normalized.items)
    
    summary_data = [
        ("Quotation ID", escape_formula(normalized.quotation_id)),
        ("Supplier Code", escape_formula(normalized.supplier_code)),
        ("Quotation Date", escape_formula(normalized.quotation_date)),
        ("Currency", escape_formula(summary_currency)),
        ("Item Count", item_count),
        ("Source Normalized Path", escape_formula(export_context.source_normalized_sha256 if False else "normalized/normalized.json")),
        ("Source Normalized SHA256", escape_formula(export_context.source_normalized_sha256)),
        ("Exported At", export_context.exported_at.isoformat().replace("+00:00", "Z")),
        ("Exporter Name", escape_formula(export_context.exporter_name)),
        ("Exporter Version", escape_formula(export_context.exporter_version)),
    ]
    
    # Thêm export summary nếu có
    if normalized.export_summary:
        summary_data.extend([
            ("Draft Item Count", normalized.export_summary.draft_item_count),
            ("Approved Count", normalized.export_summary.approved_count),
            ("Edited Count", normalized.export_summary.edited_count),
            ("Rejected Count", normalized.export_summary.rejected_count),
            ("Unreviewed Count", normalized.export_summary.unreviewed_count),
            ("Exported Item Count", normalized.export_summary.exported_item_count),
        ])
        
    ws_summary.append(["Field", "Value"])
    for field, val in summary_data:
        # Field luôn là text literal, val có thể là int hoặc text literal
        ws_summary.append([escape_formula(field), val if isinstance(val, (int, float)) else escape_formula(val)])
        
    # Auto-width Summary
    header_font = Font(name="Calibri", size=11, bold=True)
    ws_summary["A1"].font = header_font
    ws_summary["B1"].font = header_font
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
        
    # ----------------------------------------------------
    # Sheet 2: Items
    # ----------------------------------------------------
    ws_items = wb.create_sheet(title="Items")
    ws_items.freeze_panes = "A2"
    
    headers_items = [
        "item_id", "material_code", "description", "brand", "unit", 
        "quantity", "unit_price", "currency", "amount", "page_number", 
        "confidence", "reviewer", "source_draft_item_id", 
        "source_review_decision_id", "warnings"
    ]
    ws_items.append([escape_formula(h) for h in headers_items])
    
    for item in normalized.items:
        # Định dạng warnings: "[code] message; [code2] message2"
        warning_parts = []
        for w in item.warnings:
            if w.code and w.message:
                warning_parts.append(f"[{w.code}] {w.message}")
            elif w.code:
                warning_parts.append(f"[{w.code}]")
            elif w.message:
                warning_parts.append(w.message)
        warning_str = "; ".join(warning_parts) if warning_parts else None
        
        row_cells = [
            escape_formula(item.item_id),
            escape_formula(item.material_code),
            escape_formula(item.description),
            escape_formula(item.brand),
            escape_formula(item.unit),
            item.quantity, # numeric
            item.unit_price, # numeric
            escape_formula(item.currency),
            item.amount, # numeric
            item.page_number, # numeric
            item.confidence, # numeric
            escape_formula(item.reviewer),
            escape_formula(item.source_draft_item_id),
            escape_formula(item.source_review_decision_id),
            escape_formula(warning_str)
        ]
        ws_items.append(row_cells)
        
    apply_auto_width_and_styles(ws_items, max_width=60)
    
    # ----------------------------------------------------
    # Sheet 3: Warnings
    # ----------------------------------------------------
    ws_warnings = wb.create_sheet(title="Warnings")
    headers_warnings = ["level", "item_id", "code", "message"]
    ws_warnings.append([escape_formula(h) for h in headers_warnings])
    
    # File-level warnings
    for w in normalized.warnings:
        ws_warnings.append([
            "file",
            "",
            escape_formula(w.code),
            escape_formula(w.message)
        ])
        
    # Item-level warnings
    for item in normalized.items:
        for w in item.warnings:
            ws_warnings.append([
                "item",
                escape_formula(item.item_id),
                escape_formula(w.code),
                escape_formula(w.message)
            ])
            
    apply_auto_width_and_styles(ws_warnings, max_width=60)
    
    # ----------------------------------------------------
    # Sheet 4: Trace
    # ----------------------------------------------------
    ws_trace = wb.create_sheet(title="Trace")
    headers_trace = ["item_id", "source_draft_item_id", "source_review_decision_id", "page_number", "evidence_text"]
    ws_trace.append([escape_formula(h) for h in headers_trace])
    
    for item in normalized.items:
        ws_trace.append([
            escape_formula(item.item_id),
            escape_formula(item.source_draft_item_id),
            escape_formula(item.source_review_decision_id),
            item.page_number, # numeric
            escape_formula(item.evidence_text)
        ])
        
    apply_auto_width_and_styles(ws_trace, max_width=80) # trace cho rộng tối đa 80
    
    return wb
