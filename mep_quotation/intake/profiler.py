import os
import re
import hashlib
import mimetypes
import json
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Optional, Tuple, Dict, Any

import fitz  # PyMuPDF
import openpyxl
from PIL import Image

from mep_quotation.spec.models import (
    SourceProfileModel,
    TechnicalReadabilityModel,
    SourceDateCandidateModel,
    SourceRole,
    RecommendedNextAction,
    WarningModel,
    QuotationPackageModel
)
from mep_quotation.pdf.checksum import calculate_sha256

def resolve_source_file(package_path: Path, pkg: QuotationPackageModel) -> Path:
    """
    Xác định tệp nguồn duy nhất để tiến hành phân tích theo độ ưu tiên:
    1. Ưu tiên files.source_pdf nếu có khai báo và tệp tồn tại thực tế.
    2. Quét source/original.* (loại trừ file lock tạm thời ~$*).
    
    Ném ValueError rõ ràng nếu không tìm thấy hoặc có nhiều file nguồn xung đột.
    """
    # 1. Kiểm tra files.source_pdf trong package metadata
    if pkg.files and pkg.files.source_pdf:
        pdf_path = package_path / pkg.files.source_pdf
        if pdf_path.exists() and pdf_path.is_file():
            return pdf_path

    # 2. Quét source/original.*
    source_dir = package_path / "source"
    if not source_dir.exists():
        raise FileNotFoundError(f"Không tìm thấy thư mục source: {source_dir}")

    candidates = []
    for file in source_dir.iterdir():
        if file.is_file():
            name_lower = file.name.lower()
            if name_lower.startswith("original.") and not file.name.startswith("~$"):
                candidates.append(file)

    if len(candidates) == 1:
        return candidates[0]
    elif len(candidates) == 0:
        raise ValueError(f"Không tìm thấy tệp source/original.* nào trong package {package_path}")
    else:
        raise ValueError(
            f"Tìm thấy nhiều tệp source/original.*: {[f.name for f in candidates]} "
            f"và không có cấu hình metadata chính thức để phân giải."
        )

def detect_mime_and_type(file_path: Path) -> Tuple[str, str, str]:
    """Trả về (file_extension, detected_file_type, detected_mime_type)."""
    ext = file_path.suffix.lower()
    
    # Mime type mapping mặc định
    mime_type, _ = mimetypes.guess_type(str(file_path))
    if not mime_type:
        if ext == ".pdf":
            mime_type = "application/pdf"
        elif ext in (".xlsx", ".xlsm"):
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif ext == ".xls":
            mime_type = "application/vnd.ms-excel"
        elif ext == ".csv":
            mime_type = "text/csv"
        elif ext in (".png", ".jpg", ".jpeg", ".webp"):
            mime_type = f"image/{ext[1:]}"
            if ext == ".jpg":
                mime_type = "image/jpeg"
        else:
            mime_type = "application/octet-stream"

    # Xử lý file type
    if ext == ".pdf":
        file_type = "pdf"
    elif ext == ".xlsx":
        file_type = "excel_xlsx"
    elif ext == ".xlsm":
        file_type = "excel_xlsm"
    elif ext == ".xls":
        file_type = "excel_xls"
    elif ext == ".csv":
        file_type = "csv"
    elif ext in (".png", ".jpg", ".jpeg"):
        file_type = "image"
    elif ext == ".webp":
        file_type = "webp"
    else:
        file_type = "unsupported"

    return ext, file_type, mime_type

def parse_date_candidates(text: str) -> List[SourceDateCandidateModel]:
    """Dò tìm các ứng viên ngày tháng bằng Regex trên probe text."""
    candidates = []
    
    # Chuẩn hóa khoảng trắng
    text_clean = re.sub(r"\s+", " ", text)
    
    # Regex 1: YYYY-MM-DD
    matches_iso = re.finditer(r"\b(\d{4})[-/](\d{2})[-/](\d{2})\b", text_clean)
    for m in matches_iso:
        y, m_val, d = m.groups()
        try:
            # Kiểm tra ngày hợp lệ
            dt = datetime(int(y), int(m_val), int(d))
            candidates.append({
                "date": dt.strftime("%Y-%m-%d"),
                "pos": m.start(),
                "evidence": text_clean[max(0, m.start()-50):min(len(text_clean), m.end()+50)].strip()
            })
        except ValueError:
            pass

    # Regex 2: DD/MM/YYYY hoặc DD-MM-YYYY
    matches_dmy = re.finditer(r"\b(\d{1,2})[-/](\d{1,2})[-/](\d{4})\b", text_clean)
    for m in matches_dmy:
        d, m_val, y = m.groups()
        try:
            dt = datetime(int(y), int(m_val), int(d))
            candidates.append({
                "date": dt.strftime("%Y-%m-%d"),
                "pos": m.start(),
                "evidence": text_clean[max(0, m.start()-50):min(len(text_clean), m.end()+50)].strip()
            })
        except ValueError:
            pass

    # Regex 3: ngày ... tháng ... năm ...
    matches_text = re.finditer(r"ngày\s+(\d{1,2})\s+tháng\s+(\d{1,2})\s+năm\s+(\d{4})", text_clean, re.IGNORECASE)
    for m in matches_text:
        d, m_val, y = m.groups()
        try:
            dt = datetime(int(y), int(m_val), int(d))
            candidates.append({
                "date": dt.strftime("%Y-%m-%d"),
                "pos": m.start(),
                "evidence": text_clean[max(0, m.start()-50):min(len(text_clean), m.end()+50)].strip()
            })
        except ValueError:
            pass

    # Loại bỏ trùng ngày, ưu tiên ngày có chỉ dẫn rõ
    unique_dates: Dict[str, SourceDateCandidateModel] = {}
    for c in candidates:
        date_str = c["date"]
        evidence = c["evidence"]
        pos = c["pos"]
        
        # Xác định date_type và confidence bằng ngữ cảnh
        context = text_clean[max(0, pos-80):min(len(text_clean), pos+80)].lower()
        
        date_type = "issue_date_candidate"
        confidence = 0.5
        
        if "hiệu lực" in context or "áp dụng" in context or "effective" in context:
            date_type = "effective_date_candidate"
            confidence = 0.8
        elif "báo giá" in context or "ngày lập" in context or "quotation" in context or "quote date" in context:
            date_type = "quotation_date_candidate"
            confidence = 0.8
        elif "hết hạn" in context or "expiry" in context or "expired" in context:
            date_type = "expiry_date_candidate"
            confidence = 0.8
        elif "nhận" in context or "received" in context:
            date_type = "received_date_candidate"
            confidence = 0.7

        model_cand = SourceDateCandidateModel(
            date=date_str,
            date_type=date_type,
            source="text_probe",
            confidence=confidence,
            evidence=evidence
        )
        
        # Nếu trùng ngày, ưu tiên ngày có confidence cao hơn
        if date_str not in unique_dates or confidence > unique_dates[date_str].confidence:
            unique_dates[date_str] = model_cand

    # Sắp xếp theo confidence giảm dần và lấy tối đa 5 ngày ứng viên
    sorted_candidates = sorted(unique_dates.values(), key=lambda x: x.confidence, reverse=True)
    return sorted_candidates[:5]

def heuristic_source_role(text: str, file_name: str) -> Tuple[SourceRole, float]:
    """
    Xác định vai trò của tài liệu dựa trên rule-based keywords.
    Trả về (SourceRole, confidence).
    """
    text_lower = (text + " " + file_name).lower()
    
    # Từ khóa chỉ dẫn
    keywords = {
        SourceRole.supplier_quotation_candidate: [
            "báo giá", "kính gửi", "khách hàng", "quotation", "quote", "attention", "project", "dự án",
            "báo giá vật tư", "đơn giá vật tư", "chiết khấu báo giá"
        ],
        SourceRole.supplier_price_list_candidate: [
            "bảng giá", "bảng giá bán lẻ", "list giá", "price list", "pricelist", "mã sản phẩm", "đơn giá", 
            "giá bán", "bảng giá niêm yết"
        ],
        SourceRole.supplier_catalog_candidate: [
            "catalogue", "catalog", "hướng dẫn sử dụng", "mô tả kỹ thuật", "kích thước lắp đặt", "chức năng", 
            "specs", "catalog reference", "tài liệu lựa chọn"
        ],
        SourceRole.boq_candidate: [
            "bảng khối lượng", "khối lượng mời thầu", "tiên lượng", "boq", "bill of quantities", "bảng dự toán", 
            "dự toán thầu", "khối lượng vật tư thiết bị"
        ],
        SourceRole.purchase_order_candidate: [
            "purchase order", "đơn đặt hàng", "đơn hàng", "po"
        ],
        SourceRole.technical_datasheet_candidate: [
            "datasheet", "đặc tính kỹ thuật", "technical specification", "datasheets"
        ]
    }
    
    scores = {role: 0 for role in keywords.keys()}
    for role, words in keywords.items():
        for word in words:
            # Đếm số lần từ khóa xuất hiện
            matches = len(re.findall(re.escape(word), text_lower))
            scores[role] += matches

    # Tìm vai trò có điểm cao nhất
    max_role = SourceRole.unknown_document
    max_score = 0
    
    for role, score in scores.items():
        if score > max_score:
            max_score = score
            max_role = role

    # Phân tích độ tự tin
    if max_score == 0:
        return SourceRole.unknown_document, 0.0
        
    # Check tranh chấp (mixed document)
    contenders = [role for role, score in scores.items() if score == max_score and score > 0]
    if len(contenders) > 1:
        return SourceRole.mixed_document_candidate, 0.5

    # Tính confidence
    # Nếu điểm vượt trội, gán confidence cao
    confidence = 0.6 + min(0.3, max_score * 0.05)
    return max_role, round(confidence, 2)

def profile_source_file(package_path: Path, pkg: QuotationPackageModel) -> SourceProfileModel:
    """Thực hiện profiling chi tiết cho file nguồn duy nhất trong package."""
    package_path = Path(package_path).resolve()
    
    # 1. Resolve source file
    source_file_path = resolve_source_file(package_path, pkg)
    
    # 2. Get basic file metadata
    file_name = source_file_path.name
    file_size_bytes = source_file_path.stat().st_size
    source_sha256 = calculate_sha256(source_file_path)
    source_rel_path = str(source_file_path.relative_to(package_path)).replace("\\", "/")
    
    ext, file_type, mime_type = detect_mime_and_type(source_file_path)
    
    # Khởi tạo các biến chứa kết quả phân tích
    is_supported = True
    has_native_text = False
    native_text_probe_char_count = 0
    text_density = "none"
    is_scanned = False
    requires_ocr = False
    
    page_count = None
    sheet_count = None
    image_width = None
    image_height = None
    
    probe_text = ""
    warnings: List[WarningModel] = []
    
    # Kiểm tra file size
    if file_size_bytes > 50 * 1024 * 1024: # 50MB
        warnings.append(WarningModel(code="large_file", message="File nguồn có dung lượng lớn (>50MB)."))
        
    # 3. Phân tích cụ thể theo từng định dạng
    if file_type == "pdf":
        try:
            # Kiểm tra mã hóa bằng pypdf
            from pypdf import PdfReader
            pdf_reader = PdfReader(source_file_path)
            encrypted = pdf_reader.is_encrypted
        except Exception:
            encrypted = False
            
        if encrypted:
            requires_ocr = False
            is_scanned = False
            has_native_text = False
            warnings.append(WarningModel(code="password_protected_file", message="Tệp PDF bị khóa mật khẩu (password protected)."))
            recommended_action = RecommendedNextAction.reject_or_hold
        else:
            # Đọc bằng fitz (PyMuPDF) để lấy probe text
            try:
                doc = fitz.open(source_file_path)
                page_count = len(doc)
                
                # Trích xuất 3 trang đầu làm probe
                probe_pages = min(3, page_count)
                for idx in range(probe_pages):
                    page = doc[idx]
                    page_text = page.get_text() or ""
                    probe_text += page_text + "\n"
                    
                doc.close()
            except Exception as e:
                warnings.append(WarningModel(code="corrupted_file", message=f"Không thể đọc tệp PDF. Có khả năng tệp bị hỏng: {e}"))
                probe_text = ""
                page_count = 0
                
            native_text_probe_char_count = len(probe_text.strip())
            has_native_text = native_text_probe_char_count > 0
            
            # Đánh giá độ đậm đặc chữ
            if native_text_probe_char_count == 0:
                text_density = "none"
                is_scanned = True
                requires_ocr = True
                warnings.append(WarningModel(code="no_native_text_detected", message="Không phát hiện văn bản gốc. Tệp tin có thể là PDF quét ảnh."))
                warnings.append(WarningModel(code="scan_pdf_requires_ocr", message="Tệp PDF dạng quét ảnh, yêu cầu chạy OCR để nhận diện."))
                recommended_action = RecommendedNextAction.run_pdf_ocr_pipeline_later
            else:
                is_scanned = False
                requires_ocr = False
                if native_text_probe_char_count < 100:
                    text_density = "none"
                elif native_text_probe_char_count < 1000:
                    text_density = "low"
                    warnings.append(WarningModel(code="low_text_density", message="Mật độ chữ rất thấp, khả năng nhận diện có thể bị hạn chế."))
                elif native_text_probe_char_count < 5000:
                    text_density = "medium"
                else:
                    text_density = "high"
                recommended_action = RecommendedNextAction.run_pdf_native_pipeline
                
    elif file_type == "excel_xlsx":
        try:
            wb = openpyxl.load_workbook(source_file_path, read_only=True)
            sheet_count = len(wb.sheetnames)
            sheet_names = wb.sheetnames
            
            # Quét nhẹ dòng đầu và cột đầu của sheet active làm probe text
            active_sheet = wb.active
            temp_rows = []
            row_idx = 0
            for row in active_sheet.iter_rows(max_row=10, max_col=10, values_only=True):
                row_vals = [str(v) for v in row if v is not None]
                if row_vals:
                    temp_rows.append(" ".join(row_vals))
            probe_text = "\n".join(temp_rows)
            wb.close()
            
            has_native_text = True
            native_text_probe_char_count = len(probe_text)
            text_density = "high" if native_text_probe_char_count > 100 else "medium"
            
            if sheet_count > 1:
                warnings.append(WarningModel(code="multi_sheet_excel", message=f"Excel chứa nhiều trang tính ({sheet_count} sheets)."))
                
            recommended_action = RecommendedNextAction.run_excel_intake_pipeline_later
        except Exception as e:
            warnings.append(WarningModel(code="corrupted_file", message=f"Không thể đọc tệp Excel (.xlsx): {e}"))
            recommended_action = RecommendedNextAction.reject_or_hold

    elif file_type == "image":
        try:
            img = Image.open(source_file_path)
            image_width, image_height = img.size
            has_native_text = False
            requires_ocr = True
            text_density = "none"
            warnings.append(WarningModel(code="image_requires_ocr", message="File nguồn là ảnh, yêu cầu chạy OCR để nhận diện."))
            recommended_action = RecommendedNextAction.run_image_ocr_pipeline_later
        except Exception as e:
            warnings.append(WarningModel(code="corrupted_file", message=f"Không thể đọc tệp ảnh: {e}"))
            recommended_action = RecommendedNextAction.reject_or_hold

    else:
        # csv, xls, xlsm, webp hoặc unsupported
        is_supported = False
        requires_ocr = False
        text_density = "none"
        
        # Ghi nhận warning cho XLS/XLSM/CSV/WEBP
        if ext in (".xls", ".xlsm", ".csv", ".webp"):
            warnings.append(WarningModel(
                code="limited_support", 
                message=f"Định dạng {ext.upper()} chỉ được hỗ trợ giới hạn nhận diện loại file, chưa có bộ đọc sâu."
            ))
        else:
            warnings.append(WarningModel(
                code="unsupported_file_type", 
                message=f"Định dạng tệp {ext.upper()} chưa được hỗ trợ trong hệ thống."
            ))
            
        recommended_action = RecommendedNextAction.unsupported_file_type

    # 4. Xác định Document Role qua Heuristic
    source_role, role_confidence = heuristic_source_role(probe_text, file_name)
    
    # 5. Dò tìm ngày tháng ứng viên
    date_candidates = parse_date_candidates(probe_text)
    
    # 6. Hậu xử lý các warning & review requirements
    # Bổ sung cảnh báo nếu không có ngày báo giá
    has_quote_date = any(d.date_type == "quotation_date_candidate" for d in date_candidates)
    if is_supported and not has_quote_date:
        warnings.append(WarningModel(code="missing_quotation_date", message="Không phát hiện được ngày báo giá rõ ràng từ nội dung."))
        
    # Cảnh báo nếu vai trò tài liệu là unknown
    if source_role == SourceRole.unknown_document:
        warnings.append(WarningModel(code="unknown_document_role", message="Không thể nhận diện tự động vai trò tài liệu của tệp tin này."))
        
    # Bắt buộc rà soát con người nếu confidence của role thấp hoặc mixed document
    requires_human_profile_review = False
    if not is_supported:
        requires_human_profile_review = True
    elif source_role in (SourceRole.unknown_document, SourceRole.mixed_document_candidate) or role_confidence < 0.5:
        requires_human_profile_review = True
        warnings.append(WarningModel(code="manual_profile_required", message="Độ tự tin nhận diện thấp, yêu cầu con người rà soát xác thực."))
        recommended_action = RecommendedNextAction.manual_profile_required

    # Build model TechnicalReadability
    readability = TechnicalReadabilityModel(
        is_supported_file_type=is_supported,
        has_native_text=has_native_text,
        native_text_probe_char_count=native_text_probe_char_count,
        text_density_level=text_density,
        is_scanned_candidate=is_scanned,
        requires_ocr=requires_ocr,
        page_count=page_count,
        sheet_count=sheet_count,
        image_width=image_width,
        image_height=image_height
    )

    # Build model SourceProfile
    profile = SourceProfileModel(
        schema_version="1.0",
        quotation_id=pkg.quotation_id,
        source_file=source_rel_path,
        source_sha256=source_sha256,
        file_name=file_name,
        file_extension=ext,
        detected_file_type=file_type,
        detected_mime_type=mime_type,
        file_size_bytes=file_size_bytes,
        source_role=source_role,
        source_role_confidence=role_confidence,
        technical_readability=readability,
        date_candidates=date_candidates,
        recommended_next_action=recommended_action,
        requires_human_profile_review=requires_human_profile_review,
        warnings=warnings,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )

    return profile
